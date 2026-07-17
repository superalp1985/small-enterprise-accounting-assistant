#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Isolated full-year acceptance test for the packaged accounting application."""

import json
import uuid
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from account_catalog import load_account_catalog
from finance_exporter import export_finance_workbook
from finance_store import FinanceDataStore


SOLO_TEMPLATE_LABELS = (
    "咨询与专业服务",
    "软件与信息技术",
    "电商与网络零售",
    "内容创作与广告设计",
    "本地维修与安装服务",
)


def _assert(condition: Any, message: str):
    if not condition:
        raise AssertionError(message)


def _month(year: int, number: int) -> str:
    return f"{year:04d}-{number:02d}"


def _voucher_date(year: int, month: int, day: int) -> str:
    return date(year, month, min(day, 28)).isoformat()


def _formula_errors(path: Path) -> List[str]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    errors: List[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(
                    marker in value for marker in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!")
                ):
                    errors.append(f"{sheet.title}!{cell.coordinate}:{value}")
    workbook.close()
    return errors


def run_full_cycle_acceptance(root: Path, version: str = "dev",
                              catalog_path: Optional[Path] = None) -> Dict[str, Any]:
    """Exercise a complete year in a disposable account set.

    The caller supplies the root. A unique child directory is always created so
    this acceptance run cannot overwrite a real account set or a previous run.
    """
    root = Path(root).resolve()
    run_id = f"run-{date.today().isoformat()}-{uuid.uuid4().hex[:8]}"
    run_root = root / run_id
    data_dir = run_root / "data" / "small_enterprise"
    output_dir = run_root / "output"
    output_dir.mkdir(parents=True, exist_ok=False)

    catalog = load_account_catalog(catalog_path)
    templates = catalog.get("templates", {})
    for label in SOLO_TEMPLATE_LABELS:
        _assert(label in templates, f"缺少一人公司行业模板：{label}")
        _assert(templates[label].get("solo_company_template"), f"模板未标记为一人公司模板：{label}")

    store = FinanceDataStore(
        data_dir, "enterprise", "小企业会计", "小企业会计准则"
    )
    stage_results: List[Dict[str, Any]] = []
    tax_accrual_periods: List[str] = []
    archived_periods: List[str] = []
    close_vouchers: List[str] = []
    bank_match_counts: Dict[str, int] = {}
    workbook_path = output_dir / "一人公司全周期全科目验收.xlsx"

    try:
        settings = store.get_settings()
        settings["company"].update({
            "name": "一人公司全周期验收有限公司",
            "credit_code": "91110101MA00000001",
            "taxpayer_type": "小规模纳税人",
            "industry": "软件和信息技术服务业",
            "legal_representative": "测试经营者",
            "finance_contact": "测试经营者",
        })
        settings["accounting"].update({
            "account_template": "完整66科目",
            "opening_date": "2026-01-01",
            "auto_backup": True,
        })
        settings["tax"].update({
            "vat_filing_frequency": "按季",
            "cit_filing_frequency": "按季",
            "stamp_duty_filing_frequency": "按季",
            "small_low_profit": True,
            "average_employees": 1,
            "average_assets": 2000000,
            "restricted_industry": False,
        })
        store.save_settings(settings)
        stage_results.append({"stage": "建账与资格", "ok": True})

        store.upsert_opening_balance({
            "period": "2026-01", "subject": "1002 银行存款",
            "debit_balance": 2000000,
        })
        store.upsert_opening_balance({
            "period": "2026-01", "subject": "3001 实收资本",
            "credit_balance": 2000000,
        })
        _assert(store.opening_balance_totals("2026-01")["difference"] == 0, "期初余额不平")

        coverage_vouchers: Dict[str, Dict[str, Any]] = {}
        for index, account in enumerate(store.all_accounts()):
            month_number = index % 12 + 1
            day = index // 12 + 2
            amount = float(100 + index)
            subject = f"{account['code']} {account['name']}"
            credit_normal = str(account.get("normal_balance", "")).startswith("贷")
            clearing = "3001 实收资本" if account["code"] == "1002" else "1002 银行存款"
            description = f"全科目覆盖测试-{account['code']}"
            target_line = {"科目": subject, "摘要": description}
            clearing_line = {"科目": clearing, "摘要": description}
            if credit_normal:
                target_line["贷方"] = amount
                clearing_line["借方"] = amount
            else:
                target_line["借方"] = amount
                clearing_line["贷方"] = amount
            added = store.add_voucher_lines(
                [target_line, clearing_line],
                voucher_date=_voucher_date(2026, month_number, day),
            )
            coverage_vouchers[str(account["code"])] = {
                "voucher_no": added[0]["voucher_no"],
                "date": added[0]["date"],
                "lines": [target_line, clearing_line],
            }

        covered_codes = {
            str(row.get("subject_code", "")) for row in store.list_vouchers()
        }
        expected_codes = {str(row["code"]) for row in store.all_accounts()}
        _assert(covered_codes >= expected_codes, "全66科目凭证覆盖不完整")
        stage_results.append({
            "stage": "全66科目过账", "ok": True,
            "covered": len(expected_codes), "voucher_groups": len(coverage_vouchers),
        })

        roundtrip = coverage_vouchers["1001"]
        result = store.unpost_voucher(roundtrip["voucher_no"])
        _assert(result["status"] == store.UNPOSTED_STATUS, "普通凭证反过账失败")
        store.replace_voucher_group(
            roundtrip["voucher_no"], roundtrip["lines"], roundtrip["date"]
        )
        _assert(
            not [
                row for row in store.list_vouchers(include_unposted=True)
                if row["voucher_no"] == roundtrip["voucher_no"]
                and row.get("status") == store.UNPOSTED_STATUS
            ],
            "反过账凭证重新入账失败",
        )

        asset = store.upsert_fixed_asset({
            "asset_name": "验收用办公电脑", "category": "电子设备",
            "purchase_date": "2026-01-10", "original_cost": 12000,
            "residual_rate": 0.05, "useful_months": 36,
            "depreciation_start_period": "2026-02",
        })
        _assert(asset.get("id"), "固定资产卡片创建失败")

        normal_invoice_no = "FC202601001"
        for month_number in range(1, 13):
            period = _month(2026, month_number)
            sale_date = _voucher_date(2026, month_number, 10)
            expense_date = _voucher_date(2026, month_number, 12)
            sale_summary = f"{period}技术服务收款"
            expense_summary = f"{period}办公订阅付款"
            store.add_voucher_lines([
                {"科目": "1002 银行存款", "借方": 10000, "摘要": sale_summary},
                {"科目": "5001 主营业务收入", "贷方": 10000, "摘要": sale_summary},
            ], voucher_date=sale_date)
            store.add_voucher_lines([
                {"科目": "5602 管理费用", "借方": 1000, "摘要": expense_summary},
                {"科目": "1002 银行存款", "贷方": 1000, "摘要": expense_summary},
            ], voucher_date=expense_date)
            store.upsert_invoice({
                "invoice_no": normal_invoice_no if month_number == 1 else f"FC2026{month_number:02d}001",
                "invoice_date": sale_date,
                "invoice_type": "销项", "document_type": "正常发票",
                "invoice_form": "普通发票", "price_tax_mode": "含税",
                "total_amount": 10100,
            })
            store.import_bank_transactions([
                {"date": sale_date, "direction": "收入", "amount": 10000, "summary": sale_summary},
                {"date": expense_date, "direction": "支出", "amount": 1000, "summary": expense_summary},
            ])
            bank_result = store.auto_reconcile_bank_transactions(period)
            bank_match_counts[period] = int(bank_result.get("matched", 0))
            _assert(bank_result.get("unmatched", 0) == 0, f"{period}银行自动对账存在未匹配流水")

            payroll = store.upsert_payroll({
                "period": period, "employee_name": "唯一股东兼员工",
                "gross_salary": 6000, "social_personal": 500,
                "housing_personal": 0, "income_tax": 15,
                "social_company": 1000, "housing_company": 0,
                "pay_date": _voucher_date(2026, month_number, 28),
            })
            store.post_payroll_voucher(payroll["id"])
            if month_number >= 2:
                store.post_depreciation_voucher(period)

            if month_number == 6:
                store.upsert_invoice({
                    "invoice_no": "FC202606RED", "original_invoice_no": normal_invoice_no,
                    "invoice_date": sale_date, "invoice_type": "销项",
                    "document_type": "红字发票", "invoice_form": "普通发票",
                    "price_tax_mode": "含税", "total_amount": 1010,
                })
            if month_number == 9:
                store.upsert_invoice({
                    "invoice_date": sale_date, "document_type": "未开票收入",
                    "price_tax_mode": "含税", "total_amount": 2020,
                })
            if month_number == 12:
                store.upsert_invoice({
                    "invoice_no": "FC202612SPECIAL", "invoice_date": sale_date,
                    "invoice_type": "销项", "document_type": "正常发票",
                    "invoice_form": "增值税专用发票",
                    "price_tax_mode": "含税", "total_amount": 1010,
                })

            if month_number in (3, 6, 9, 12):
                store.upsert_stamp_duty_item({
                    "period": period, "item": "买卖合同",
                    "taxable_amount": 10000, "rate": 0.0003,
                })
                if month_number == 12:
                    store.upsert_tax_adjustment({
                        "period": period, "direction": "调增", "amount": 500,
                        "category": "申报前验收调整", "basis": "全周期测试",
                    })
                preview = store.tax_accrual_preview(period)
                if preview.get("can_post") and preview.get("lines"):
                    store.post_tax_accrual_voucher(period)
                    tax_accrual_periods.append(period)
                    if month_number == 3:
                        store.unpost_tax_accrual_voucher(period)
                        store.post_tax_accrual_voucher(period)

            close_voucher = store.post_profit_close_voucher(period)
            close_vouchers.append(close_voucher)
            if month_number == 1:
                store.unpost_profit_close_voucher(period)
                close_vouchers[-1] = store.post_profit_close_voucher(period)

            _assert(not store.voucher_balance_issues(period), f"{period}存在不平衡凭证")
            store.set_period_status(period, "已归档", "全周期验收归档")
            archived_periods.append(period)
            if month_number == 1:
                store.reopen_archived_period(period, "测试归档撤销")
                store.set_period_status(period, "已归档", "重新完成归档")

        _assert(len(close_vouchers) == 12, "12个月损益结转覆盖不完整")
        _assert(len(archived_periods) == 12, "12个月归档覆盖不完整")
        _assert(tax_accrual_periods, "季度税费计提未生成任何凭证")
        _assert(all(value >= 2 for value in bank_match_counts.values()), "银行对账覆盖不完整")
        stage_results.append({
            "stage": "年度业务全周期", "ok": True,
            "months": 12, "period_closes": len(close_vouchers),
            "tax_accrual_periods": tax_accrual_periods,
        })

        iit = store.individual_income_tax_summary("2026-12")
        stamp = store.stamp_duty_summary("2026-12")
        annual_cit = store.annual_cit_summary(2026)
        _assert(iit.get("rows"), "个人所得税累计预扣准备表无数据")
        _assert(stamp.get("items"), "印花税准备表无数据")
        _assert("taxable_income" in annual_cit, "企业所得税年度汇算准备数据缺失")
        stage_results.append({
            "stage": "全年税务准备", "ok": True,
            "iit_rows": len(iit["rows"]), "stamp_rows": len(stamp["items"]),
        })

        backup = store.create_backup("全周期验收恢复点", kind="acceptance")
        before_name = store.get_settings()["company"]["name"]
        changed = store.get_settings()
        changed["company"]["name"] = "不应保留的测试名称"
        store.save_settings(changed)
        store.restore_backup(backup)
        _assert(store.get_settings()["company"]["name"] == before_name, "备份恢复后资料不一致")
        integrity = store.integrity_check()
        _assert(integrity.get("ok"), "SQLite完整性检查失败")
        _assert(integrity.get("journal_mode") == "wal", "SQLite未使用WAL模式")
        stage_results.append({"stage": "容灾恢复", "ok": True, **integrity})

        export_finance_workbook(store, workbook_path, "2026-12")
        workbook = load_workbook(workbook_path, data_only=False, read_only=True)
        sheet_names = list(workbook.sheetnames)
        report_basis = workbook["报表取数底稿"]
        inventory_formula = next(
            (
                str(row[14].value)
                for row in report_basis.iter_rows()
                if row[12].value == 9 and row[13].value == "存货"
            ),
            "",
        )
        workbook.close()
        errors = _formula_errors(workbook_path)
        _assert(len(sheet_names) == 30, f"导出工作表数量异常：{len(sheet_names)}")
        _assert(not errors, "导出工作簿存在明显公式错误")
        required_inventory_codes = ("4001", "4101", "4401", "4403")
        missing_inventory_codes = [
            code for code in required_inventory_codes
            if f'"{code}"' not in inventory_formula
        ]
        _assert(
            not missing_inventory_codes,
            "资产负债表存货取数遗漏成本科目：" + ",".join(missing_inventory_codes),
        )
        stage_results.append({
            "stage": "年末报表导出", "ok": True,
            "workbook": str(workbook_path), "sheets": len(sheet_names),
            "formula_errors": len(errors),
            "inventory_cost_codes": list(required_inventory_codes),
        })

        posted_codes = {
            str(row.get("subject_code", "")) for row in store.list_vouchers()
        }
        missing_codes = sorted(expected_codes - posted_codes)
        _assert(not missing_codes, "年末账套缺少已过账科目：" + ",".join(missing_codes))
        return {
            "ok": True,
            "version": version,
            "run_id": run_id,
            "run_root": str(run_root),
            "account_catalog_count": len(store.all_accounts()),
            "covered_account_count": len(expected_codes),
            "missing_account_codes": [],
            "solo_templates": list(SOLO_TEMPLATE_LABELS),
            "months_processed": 12,
            "archived_periods": archived_periods,
            "tax_accrual_periods": tax_accrual_periods,
            "bank_match_counts": bank_match_counts,
            "workbook": str(workbook_path),
            "workbook_size": workbook_path.stat().st_size,
            "stages": stage_results,
        }
    finally:
        store.close()


def write_acceptance_report(path: Path, report: Dict[str, Any]):
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
