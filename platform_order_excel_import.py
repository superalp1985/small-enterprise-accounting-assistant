#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import common platform order exports into the batch review queue."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional

from openpyxl import load_workbook


class PlatformOrderExcelImportError(ValueError):
    """Raised when an Excel workbook is not a supported order export."""


HEADER_ALIASES = {
    "order_no": ("订单编号", "订单号", "订单ID", "订单id"),
    "payment_detail": ("支付详情", "支付信息", "付款详情"),
    "gross_amount": ("总金额", "订单金额", "商品总金额"),
    "paid_amount": ("买家实付金额", "买家实付款", "实付金额"),
    "status": ("订单状态", "交易状态"),
    "logistics_no": ("物流单号", "运单号"),
    "logistics_company": ("物流公司", "承运商"),
    "date": ("下单时间", "支付时间", "成交时间", "订单时间", "日期"),
}

SUCCESS_MARKERS = ("成功", "完成", "已结算", "已收货")
CLOSED_MARKERS = ("关闭", "取消", "交易失败", "退款成功")
PLATFORM_MARKERS = (
    "抖音", "快手", "淘宝", "天猫", "拼多多", "京东", "小红书", "视频号",
    "微信小店", "B站", "哔哩哔哩", "知乎", "小店",
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float):
        if value.is_integer():
            return format(value, ".0f")
        return format(value, ".15g")
    return str(value).strip()


def _money(value: Any) -> float:
    text = _text(value).replace(",", "").replace("¥", "").replace("￥", "")
    if not text:
        return 0.0
    try:
        return float(Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return 0.0


def _date_text(value: Any, fallback: str) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = _text(value)
    candidates = [
        re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text),
        re.search(r"(20\d{2})(\d{2})(\d{2})(?:\d{6})?", text),
    ]
    for match in candidates:
        if not match:
            continue
        try:
            return date(*(int(part) for part in match.groups()[:3])).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return fallback


def _header_map(row: Iterable[Any]) -> Dict[str, int]:
    return {_text(value).replace("\n", "").replace(" ", ""): index for index, value in enumerate(row)}


def _find_index(headers: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for alias in aliases:
        index = headers.get(alias.replace(" ", ""))
        if index is not None:
            return index
    return None


def _value(row: tuple[Any, ...], headers: Dict[str, int], aliases: Iterable[str]) -> Any:
    index = _find_index(headers, aliases)
    return row[index] if index is not None and index < len(row) else None


def _find_order_sheet(workbook):
    for sheet in workbook.worksheets:
        first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = _header_map(first)
        if all(_find_index(headers, HEADER_ALIASES[key]) is not None for key in ("order_no", "gross_amount", "status")):
            return sheet, headers
    return None, None


def _platform_name(text: str) -> str:
    for marker in PLATFORM_MARKERS:
        if marker in text:
            return marker
    return "平台"


def _status_state(status: str) -> tuple[bool, str, List[str]]:
    compact = status.replace(" ", "")
    warnings: List[str] = []
    if any(marker in compact for marker in CLOSED_MARKERS):
        return True, "不可入账", [f"订单状态为“{status or '未填写'}”，不进入入账队列"]
    if any(marker in compact for marker in SUCCESS_MARKERS):
        return False, "待处理", warnings
    warnings.append(f"订单状态为“{status or '未填写'}”，请人工确认后再入账")
    return True, "不可入账", warnings


def read_platform_order_workbook(
    source: Path,
    company_name: str = "",
    company_industry: str = "",
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> List[Dict[str, Any]]:
    """Return one review draft per order row from a platform export."""
    source = Path(source).resolve()
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise PlatformOrderExcelImportError("仅支持平台导出的 .xlsx 或 .xlsm 文件")
    try:
        # Platform exports are small enough to load eagerly.  The read-only
        # worksheet iterator can keep the XML parser alive while the Tk review
        # queue is being drained, which is unstable in the frozen Windows build.
        workbook = load_workbook(
            source,
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise PlatformOrderExcelImportError(f"无法读取平台订单 Excel：{exc}") from exc

    sheet, headers = _find_order_sheet(workbook)
    if sheet is None or headers is None:
        workbook.close()
        raise PlatformOrderExcelImportError(
            "未找到订单编号、总金额和订单状态列，请选择购物平台订单导出文件"
        )

    fallback_date = datetime.fromtimestamp(source.stat().st_mtime).strftime("%Y-%m-%d")
    drafts: List[Dict[str, Any]] = []
    try:
        total_rows = max(int(getattr(sheet, "max_row", 0) or 0) - 1, 1)
    except (TypeError, ValueError):
        total_rows = 1
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        if progress_callback and (row_index == 1 or row_index % 25 == 0 or row_index >= total_rows):
            progress_callback(min(row_index, total_rows), total_rows, "正在读取平台订单")
        row_number = row_index + 1
        if not any(value not in (None, "") for value in row):
            continue
        order_no = _text(_value(row, headers, HEADER_ALIASES["order_no"]))
        if not order_no:
            continue
        payment_detail = _text(_value(row, headers, HEADER_ALIASES["payment_detail"]))
        status = _text(_value(row, headers, HEADER_ALIASES["status"]))
        gross_amount = _money(_value(row, headers, HEADER_ALIASES["gross_amount"]))
        paid_amount = _money(_value(row, headers, HEADER_ALIASES["paid_amount"]))
        amount = gross_amount or paid_amount
        if amount <= 0 and paid_amount > 0:
            amount = paid_amount
        warnings: List[str] = []
        if gross_amount > 0 and paid_amount > 0 and abs(gross_amount - paid_amount) >= 0.01:
            warnings.append("总金额与买家实付金额不一致，请核对优惠、退款和平台分摊")
        non_postable, item_status, status_warnings = _status_state(status)
        warnings.extend(status_warnings)
        platform = _platform_name(payment_detail)
        description = payment_detail or f"{platform}订单 {order_no}"
        invoice_date = _date_text(
            _value(row, headers, HEADER_ALIASES["date"]),
            _date_text(payment_detail, fallback_date),
        )
        drafts.append({
            "file_name": f"{source.name} / {order_no}",
            "filepath": str(source),
            "source_type": "platform_excel",
            "source_row": row_number,
            "source_reference": order_no,
            "platform": platform,
            "status": item_status,
            "non_postable": non_postable,
            "amount": round(amount, 2),
            "net_amount": round(amount, 2),
            "tax_amount": 0.0,
            "total_amount": round(amount, 2),
            "signed_total_amount": round(amount, 2),
            "signed_net_amount": round(amount, 2),
            "signed_tax_amount": 0.0,
            "gross_amount": round(gross_amount, 2),
            "paid_amount": round(paid_amount, 2),
            "description": description,
            "item_descriptions": [payment_detail] if payment_detail else [],
            "tax_categories": [],
            "company_industry": company_industry,
            "invoice_date": invoice_date,
            "invoice_code": "",
            "invoice_no": order_no,
            "seller": company_name,
            "seller_tax_id": "",
            "buyer": "",
            "buyer_tax_id": "",
            "counterparty": f"{platform}订单客户",
            "invoice_type": "销项",
            "document_type": "正常发票",
            "invoice_form": "平台订单",
            "invoice_status": status,
            "risk_level": "",
            "direction": "贷方",
            "counter_subject": "1012 其他货币资金-平台待结算款",
            "matched_subject": "",
            "match_score": 0.0,
            "match_type": "",
            "confidence": 1.0,
            "needs_review": True,
            "manual_review_required": True,
            "warnings": warnings,
            "logistics_no": _text(_value(row, headers, HEADER_ALIASES["logistics_no"])),
            "logistics_company": _text(_value(row, headers, HEADER_ALIASES["logistics_company"])),
        })

    workbook.close()
    if not drafts:
        raise PlatformOrderExcelImportError("平台订单 Excel 中没有可导入的订单行")
    return drafts
