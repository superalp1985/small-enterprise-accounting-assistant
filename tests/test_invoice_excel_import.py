#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook
import xlrd

from invoice_excel_import import InvoiceExcelImportError, read_tax_invoice_workbook


BASE_HEADERS = [
    "序号", "发票代码", "发票号码", "数电发票号码",
    "销方识别号", "销方名称", "购方识别号", "购买方名称",
    "开票日期", "金额", "税额", "价税合计", "发票来源",
    "发票票种", "发票状态", "是否正数发票", "发票风险等级",
]
DETAIL_HEADERS = [
    "序号", "发票代码", "发票号码", "数电发票号码",
    "销方识别号", "销方名称", "购方识别号", "购买方名称",
    "开票日期", "货物或应税劳务名称", "金额", "税额", "价税合计",
]


class TaxInvoiceExcelImportTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.path = Path(self.temp.name) / "tax-export.xlsx"

    def tearDown(self):
        self.temp.cleanup()

    def _write_workbook(self):
        workbook = Workbook()
        detail = workbook.active
        detail.title = "信息汇总表"
        detail.append(DETAIL_HEADERS)
        detail.append([
            1, "", "", "26952000003230149426", "SELLER001", "销售公司",
            "BUYER001", "购买公司", "2026-07-29 21:38:43",
            "*茶*茶叶礼盒", 78.05, 10.15, 88.20,
        ])
        detail.append([
            2, "", "", "26952000003230149426", "SELLER001", "销售公司",
            "BUYER001", "购买公司", "2026-07-29 21:38:43",
            "*茶*茶叶礼盒", -26.55, -3.45, -30.00,
        ])
        detail.append([
            3, "011001", "00012345", "", "BUYER001", "购买公司",
            "CUSTOMER1", "客户公司", "2026/07/28",
            "*技术服务*咨询服务", 100, 1, 101,
        ])

        base = workbook.create_sheet("发票基础信息")
        base.append(BASE_HEADERS)
        base.append([
            1, "", "", "26952000003230149426", "SELLER001", "销售公司",
            "BUYER001", "购买公司", "2026-07-29 21:38:43",
            51.50, 6.70, 58.20, "电子发票服务平台", "数电普通发票",
            "正常", "是", "正常",
        ])
        base.append([
            2, "011001", "00012345", "", "BUYER001", "购买公司",
            "CUSTOMER1", "客户公司", "2026/07/28",
            100, 1, 101, "税控系统", "增值税普通发票",
            "正常", "是", "正常",
        ])
        base.append([
            3, "", "", "26952000003230149427", "SELLER002", "其他公司",
            "BUYER001", "购买公司", "2026-07-27",
            -50, -0.5, -50.5, "电子发票服务平台", "数电普通发票",
            "已红冲-全额", "否", "正常",
        ])
        workbook.save(self.path)

    def test_groups_discount_rows_and_preserves_invoice_identifiers(self):
        self._write_workbook()
        rows = read_tax_invoice_workbook(
            self.path, company_tax_id="BUYER001", company_industry="零售业"
        )

        self.assertEqual(len(rows), 3)
        digital = rows[0]
        self.assertEqual(digital["invoice_no"], "26952000003230149426")
        self.assertEqual(digital["invoice_date"], "2026-07-29")
        self.assertEqual(digital["tax_categories"], ["茶"])
        self.assertEqual(digital["item_descriptions"], ["*茶*茶叶礼盒"])
        self.assertAlmostEqual(digital["net_amount"], 51.50)
        self.assertAlmostEqual(digital["tax_amount"], 6.70)
        self.assertAlmostEqual(digital["amount"], 58.20)
        self.assertEqual(digital["invoice_type"], "进项")

        legacy = rows[1]
        self.assertEqual(legacy["invoice_code"], "011001")
        self.assertEqual(legacy["invoice_no"], "00012345")
        self.assertEqual(legacy["invoice_type"], "销项")
        self.assertEqual(legacy["direction"], "贷方")

        red = rows[2]
        self.assertTrue(red["non_postable"])
        self.assertEqual(red["status"], "不可入账")
        self.assertEqual(red["document_type"], "红字发票")
        self.assertAlmostEqual(red["amount"], 50.50)

    def test_rejects_unrelated_workbook(self):
        workbook = Workbook()
        workbook.active.append(["姓名", "电话"])
        workbook.save(self.path)
        with self.assertRaises(InvoiceExcelImportError):
            read_tax_invoice_workbook(self.path)

    def test_reads_legacy_xls_tax_export(self):
        class FakeSheet:
            def __init__(self, name, rows):
                self.name = name
                self._rows = rows
                self.nrows = len(rows)
                self.ncols = max(len(row) for row in rows)

            def cell(self, row, column):
                value = self._rows[row][column] if column < len(self._rows[row]) else ""
                return SimpleNamespace(value=value, ctype=xlrd.XL_CELL_TEXT)

        class FakeBook:
            datemode = 0

            def __init__(self):
                self.released = False
                self.sheets = {
                    "发票基础信息1": FakeSheet("发票基础信息1", [
                        BASE_HEADERS,
                        [
                            1, "", "", "26952000003230149426",
                            "SELLER001", "销售公司", "BUYER001", "购买公司",
                            "2026-07-29", 78.05, 10.15, 88.20,
                            "电子发票服务平台", "数电普通发票", "正常", "是", "正常",
                        ],
                    ]),
                }

            def sheet_names(self):
                return list(self.sheets)

            def sheet_by_name(self, name):
                return self.sheets[name]

            def release_resources(self):
                self.released = True

        book = FakeBook()
        legacy_path = Path(self.temp.name) / "tax-export.xls"
        with patch("invoice_excel_import.xlrd.open_workbook", return_value=book) as opener:
            rows = read_tax_invoice_workbook(legacy_path, company_tax_id="BUYER001")

        opener.assert_called_once_with(str(legacy_path.resolve()), on_demand=True)
        self.assertTrue(book.released)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["invoice_no"], "26952000003230149426")
        self.assertEqual(rows[0]["invoice_type"], "进项")
        self.assertEqual(rows[0]["amount"], 88.20)
