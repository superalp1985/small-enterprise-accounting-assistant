#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from finance_exporter import export_finance_workbook
from finance_store import FinanceDataStore
from legal_notice import LEGAL_NOTICE_SUMMARY, POLICY_PRESET_REVIEW_DATE, POLICY_SOURCES
from management_dialogs import company_profile_errors, normalize_credit_code


class FinanceStoreTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.store = FinanceDataStore(
            self.root / "small_enterprise",
            "enterprise",
            "小企业会计",
            "小企业会计准则",
        )

    def tearDown(self):
        self.store.close()
        self.temp.cleanup()

    def test_export_directory_defaults_are_portable(self):
        settings = self.store.get_settings()
        self.assertEqual(settings["export"]["default_dir"], "exports")
        self.assertEqual(
            self.store.resolve_export_dir(settings["export"]["default_dir"]),
            self.store.data_dir / "exports",
        )

        legacy = self.store.data_dir / "exports"
        settings["export"]["default_dir"] = str(legacy)
        self.store.save_settings(settings)
        self.assertEqual(
            self.store.get_settings()["export"]["default_dir"],
            "exports",
        )

    def test_multiple_drafts_are_upserted_together(self):
        saved = self.store.add_drafts([
            {"type": "batch", "description": "票据一"},
            {"type": "batch", "description": "票据二"},
        ])
        self.assertEqual(len(saved), 2)
        self.assertNotEqual(saved[0]["id"], saved[1]["id"])

        updated = self.store.add_drafts([
            {**saved[0], "description": "票据一已修改"},
            {"type": "manual", "description": "手工草稿"},
        ])
        self.assertEqual(len(updated), 2)
        drafts = self.store.list_drafts()
        self.assertEqual(len(drafts), 3)
        by_id = {row["id"]: row for row in drafts}
        self.assertEqual(by_id[saved[0]["id"]]["description"], "票据一已修改")

        self.store.delete_drafts([saved[0]["id"], saved[1]["id"]])
        remaining = self.store.list_drafts()
        self.assertEqual(len(remaining), 1)
        self.assertEqual(remaining[0]["type"], "manual")

    def test_company_profile_validation_and_fixed_product_scope(self):
        self.assertEqual(
            company_profile_errors({"company": {"name": "", "credit_code": ""}}),
            ["请填写企业名称", "请填写统一社会信用代码/税号"],
        )
        self.assertEqual(
            normalize_credit_code(" 91110101ma00000001 "),
            "91110101MA00000001",
        )

        settings = self.store.get_settings()
        settings["company"].update({
            "name": "  测试科技有限公司  ",
            "credit_code": " 91110101ma00000001 ",
            "taxpayer_type": "一般纳税人",
            "currency": "美元",
        })
        settings["tax"].update({
            "small_low_profit": False,
            "invoice_required": False,
            "input_vat_deductible": True,
        })
        settings["accounting"].update({
            "standard": "企业会计准则",
            "fiscal_year_start": "04-01",
            "auto_backup": False,
        })
        self.store.save_settings(settings)

        loaded = self.store.get_settings()
        self.assertEqual(company_profile_errors(loaded), [])
        self.assertEqual(loaded["company"]["name"], "测试科技有限公司")
        self.assertEqual(loaded["company"]["credit_code"], "91110101MA00000001")
        self.assertEqual(loaded["company"]["taxpayer_type"], "小规模纳税人")
        self.assertEqual(loaded["company"]["currency"], "人民币")
        self.assertTrue(loaded["tax"]["small_low_profit"])
        self.assertTrue(loaded["tax"]["invoice_required"])
        self.assertFalse(loaded["tax"]["input_vat_deductible"])
        self.assertEqual(loaded["accounting"]["standard"], "小企业会计准则")
        self.assertEqual(loaded["accounting"]["fiscal_year_start"], "01-01")
        self.assertTrue(loaded["accounting"]["auto_backup"])

    def _seed_balanced_ledger(self):
        self.store.add_voucher_lines([
            {"科目": "1002 银行存款", "借方": 100, "摘要": "服务收款", "source": "manual"},
            {"科目": "5001 主营业务收入", "贷方": 100, "摘要": "服务收款", "source": "manual"},
        ], voucher_date="2026-07-01")
        self.store.add_voucher_lines([
            {"科目": "5602 管理费用", "借方": 40, "摘要": "办公支出", "source": "manual"},
            {"科目": "1002 银行存款", "贷方": 40, "摘要": "办公支出", "source": "manual"},
        ], voucher_date="2026-07-02")

    def test_balanced_voucher_and_atomic_json_persistence(self):
        added = self.store.add_voucher_lines([
            {"科目": "5602 管理费用", "借方": 88.8, "摘要": "软件订阅"},
            {"科目": "1002 银行存款", "贷方": 88.8, "摘要": "软件订阅"},
        ], voucher_date="2026-07-03")
        self.assertEqual(len(added), 2)
        self.assertEqual(added[0]["voucher_no"], added[1]["voucher_no"])
        self.assertEqual(sum(row["debit"] for row in added), 88.8)
        self.assertEqual(sum(row["credit"] for row in added), 88.8)
        self.assertFalse((self.store.data_dir / ".ledger.json.tmp").exists())

        with self.assertRaises(ValueError):
            self.store.add_voucher_lines([
                {"科目": "5602 管理费用", "借方": 10},
                {"科目": "1002 银行存款", "贷方": 9},
            ], voucher_date="2026-07-03")

    def test_batch_invoice_and_voucher_posting_is_atomic(self):
        entries = [
            {
                "voucher_date": "2026-07-05",
                "lines": [
                    {"科目": "5602 管理费用", "借方": 58.2, "摘要": "茶叶"},
                    {"科目": "1002 银行存款", "贷方": 58.2, "摘要": "茶叶"},
                ],
                "invoice": {
                    "invoice_no": "26952000003230149426",
                    "invoice_date": "2026-07-05",
                    "amount": 51.5, "tax_amount": 6.7, "total_amount": 58.2,
                },
            },
            {
                "voucher_date": "2026-07-05",
                "lines": [
                    {"科目": "5602 管理费用", "借方": 101, "摘要": "服务费"},
                    {"科目": "1002 银行存款", "贷方": 101, "摘要": "服务费"},
                ],
                "invoice": {
                    "invoice_code": "011001", "invoice_no": "00012345",
                    "invoice_date": "2026-07-05",
                    "amount": 100, "tax_amount": 1, "total_amount": 101,
                },
            },
        ]
        posted = self.store.post_invoice_vouchers(entries)
        self.assertEqual(len(posted), 2)
        self.assertEqual(len(self.store.list_vouchers()), 4)
        self.assertEqual(len(self.store.list_invoices()), 2)
        self.assertNotEqual(posted[0]["voucher_no"], posted[1]["voucher_no"])

        ledger_before = self.store.list_vouchers(include_unposted=True)
        invoices_before = self.store.list_invoices()
        invalid = [entries[0], {**entries[1], "lines": [
            {"科目": "5602 管理费用", "借方": 10},
            {"科目": "1002 银行存款", "贷方": 9},
        ]}]
        with self.assertRaises(ValueError):
            self.store.post_invoice_vouchers(invalid)
        self.assertEqual(self.store.list_vouchers(include_unposted=True), ledger_before)
        self.assertEqual(self.store.list_invoices(), invoices_before)

    def test_profile_isolation_and_duplicate_invoice_validation(self):
        other = FinanceDataStore(self.root / "governmental", "governmental", "行政事业单位会计")
        self._seed_balanced_ledger()
        self.assertEqual(len(self.store.list_vouchers()), 4)
        self.assertEqual(other.list_vouchers(), [])

        duplicate = {
            "id": "a", "invoice_code": "011001", "invoice_no": "00012345",
            "invoice_date": "2026-07-02", "amount": 10, "tax_amount": 1,
        }
        self.store._write_json(self.store.invoices_path, [duplicate, {**duplicate, "id": "b"}])
        codes = {issue["code"] for issue in self.store.validate("2026-07")}
        self.assertIn("DUPLICATE_INVOICE", codes)

    def test_configurable_tax_calculation(self):
        settings = self.store.get_settings()
        settings["company"].update({
            "name": "测试科技有限公司", "credit_code": "91110101MA00000001",
            "taxpayer_type": "小规模纳税人",
        })
        settings["tax"].update({
            "vat_rate": 0.01, "surcharge_rate": 0.06,
            "cit_rate": 0.05, "input_vat_deductible": False,
        })
        self.store.save_settings(settings)
        self._seed_balanced_ledger()
        self.store.add_voucher_lines([
            {"科目": "2221 应交税费", "借方": 30, "摘要": "缴纳税费"},
            {"科目": "1002 银行存款", "贷方": 30, "摘要": "缴纳税费"},
        ], voucher_date="2026-07-03")
        summary = self.store.tax_summary("2026-07")
        self.assertEqual(summary["revenue"], 100)
        self.assertEqual(summary["expenses"], 40)
        self.assertEqual(summary["profit"], 60)
        self.assertEqual(summary["vat_payable"], 0)
        self.assertEqual(summary["surcharge"], 0)
        self.assertEqual(summary["cit_payable"], 3)
        self.assertTrue(summary["vat"]["threshold_eligible"])
        self.assertEqual(summary["period"]["key"], "2026-Q3")

    def test_red_unbilled_invoice_price_tax_and_quarterly_vat(self):
        normal = self.store.upsert_invoice({
            "invoice_no": "N001", "invoice_date": "2026-07-02",
            "invoice_type": "销项", "document_type": "正常发票",
            "invoice_form": "普通发票", "price_tax_mode": "含税",
            "total_amount": 202000,
        })
        unbilled = self.store.upsert_invoice({
            "invoice_date": "2026-08-10", "document_type": "未开票收入",
            "price_tax_mode": "含税", "total_amount": 50500,
        })
        red = self.store.upsert_invoice({
            "invoice_no": "R001", "original_invoice_no": "N001",
            "invoice_date": "2026-08-20", "invoice_type": "销项",
            "document_type": "红字发票", "price_tax_mode": "含税",
            "total_amount": 10100,
        })
        special = self.store.upsert_invoice({
            "invoice_no": "S001", "invoice_date": "2026-09-01",
            "invoice_type": "销项", "invoice_form": "增值税专用发票",
            "price_tax_mode": "含税", "total_amount": 10100,
        })
        self.assertEqual(normal["amount"], 200000)
        self.assertEqual(unbilled["invoice_type"], "销项")
        self.assertEqual(unbilled["invoice_form"], "无票")
        self.assertEqual(red["amount"], -10000)
        self.assertEqual(red["tax_amount"], -100)
        self.assertEqual(special["tax_amount"], 100)

        summary = self.store.tax_summary("2026-08")
        self.assertEqual(summary["period"]["months"], ["2026-07", "2026-08", "2026-09"])
        self.assertEqual(summary["vat"]["sales"], 250000)
        self.assertTrue(summary["vat"]["threshold_eligible"])
        self.assertEqual(summary["vat"]["non_exempt_sales"], 10000)
        self.assertEqual(summary["vat_payable"], 100)

    def test_tax_adjustment_iit_stamp_duty_and_tax_accrual(self):
        self._seed_balanced_ledger()
        self.store.upsert_tax_adjustment({
            "period": "2026-07", "direction": "调增", "amount": 40,
            "category": "无票支出", "basis": "人工复核",
        })
        self.store.upsert_tax_adjustment({
            "period": "2026-07", "direction": "已预缴所得税", "amount": 1,
            "category": "以前季度已预缴",
        })
        cit = self.store.cit_prepayment_summary("2026-07")
        self.assertEqual(cit["taxable_income"], 100)
        self.assertEqual(cit["current_tax"], 5)
        self.assertEqual(cit["cit_payable"], 4)

        self.store.upsert_payroll({
            "period": "2026-01", "employee_name": "张三", "gross_salary": 10000,
            "social_personal": 500, "income_tax": 135,
        })
        iit = self.store.individual_income_tax_summary("2026-01")
        self.assertEqual(iit["rows"][0]["current_withholding"], 135)
        self.assertEqual(iit["rows"][0]["difference"], 0)

        self.store.upsert_stamp_duty_item({
            "period": "2026-07", "item": "买卖合同",
            "taxable_amount": 100000, "rate": 0.0003,
        })
        stamp = self.store.stamp_duty_summary("2026-08")
        self.assertEqual(stamp["period"]["key"], "2026-Q3")
        self.assertEqual(stamp["stamp_duty_payable"], 15)

        preview = self.store.tax_accrual_preview("2026-09")
        self.assertFalse(preview["posted"])
        self.assertEqual(sum(row.get("debit", 0) for row in preview["lines"]), 4)
        voucher_no = self.store.post_tax_accrual_voucher("2026-09")
        self.assertTrue(voucher_no)
        self.assertTrue(self.store.tax_accrual_preview("2026-09")["posted"])
        self.assertEqual(self.store.tax_summary("2026-07")["cit_payable"], 4)
        self.assertEqual(self.store.unpost_tax_accrual_voucher("2026-09"), voucher_no)

    def test_policy_presets_are_visible_editable_and_validated(self):
        settings = self.store.get_settings()
        tax = settings["tax"]
        self.assertEqual(tax["policy_reference_date"], POLICY_PRESET_REVIEW_DATE)
        self.assertEqual(tax["policy_effective_through"], "2027-12-31")
        self.assertEqual(tax["vat_monthly_exemption_threshold"], 100000)
        self.assertEqual(tax["vat_quarterly_exemption_threshold"], 300000)
        self.assertEqual(tax["cit_taxable_income_limit"], 3000000)
        self.assertEqual(tax["cit_employee_limit"], 300)
        self.assertEqual(tax["cit_asset_limit"], 50000000)
        self.assertEqual(tax["average_employees"], 1)
        self.assertEqual(tax["iit_monthly_deduction"], 5000)
        self.assertEqual(tax["stamp_duty_relief_rate"], 0.5)
        self.assertEqual(tax["default_price_tax_mode"], "含税")

        tax["vat_rate"] = 0.03
        tax["vat_monthly_exemption_threshold"] = 80000
        self.store.save_settings(settings)
        saved = self.store.get_settings()["tax"]
        self.assertEqual(saved["vat_rate"], 0.03)
        self.assertEqual(saved["vat_monthly_exemption_threshold"], 80000)

        saved_settings = self.store.get_settings()
        saved_settings["tax"]["cit_employee_limit"] = -1
        with self.assertRaisesRegex(ValueError, "从业人数上限"):
            self.store.save_settings(saved_settings)
        saved_settings = self.store.get_settings()
        saved_settings["tax"]["stamp_duty_relief_rate"] = 1.1
        with self.assertRaisesRegex(ValueError, "印花税减征系数"):
            self.store.save_settings(saved_settings)

    def test_policy_sources_are_release_managed(self):
        settings = self.store.get_settings()
        settings["policy_sources"] = [
            {"title": "过期来源", "url": "https://example.invalid"}
        ]
        self.store.save_settings(settings)

        loaded = self.store.get_settings()
        self.assertEqual(loaded["policy_sources"], POLICY_SOURCES)
        self.assertIsNot(loaded["policy_sources"], POLICY_SOURCES)
        self.assertIsNot(loaded["policy_sources"][0], POLICY_SOURCES[0])

    def test_backup_validation_and_restore(self):
        settings = self.store.get_settings()
        settings["company"]["name"] = "备份前公司"
        self.store.save_settings(settings)
        backup = self.store.create_backup("单元测试")
        self.assertEqual(self.store.validate_backup(backup)["profile_key"], "enterprise")

        settings["company"]["name"] = "已修改公司"
        self.store.save_settings(settings)
        self.store.restore_backup(backup)
        self.assertEqual(self.store.get_settings()["company"]["name"], "备份前公司")

    def test_sqlite_wal_integrity_startup_backup_rotation_and_json_repair(self):
        self._seed_balanced_ledger()
        integrity = self.store.integrity_check()
        self.assertTrue(integrity["ok"])
        self.assertEqual(integrity["journal_mode"], "wal")

        for _ in range(7):
            self.store.startup_safety_check(keep=5)
        automatic = [
            row for row in self.store.list_backups()
            if row.get("kind") == "auto_startup"
        ]
        self.assertEqual(len(automatic), 5)
        with zipfile.ZipFile(Path(automatic[0]["path"])) as archive:
            self.assertIn("accounting.db", archive.namelist())

        self.store.ledger_path.write_text("{损坏", encoding="utf-8")
        self.assertEqual(len(self.store.list_vouchers()), 4)
        with open(self.store.ledger_path, encoding="utf-8") as handle:
            self.assertEqual(len(json.load(handle)), 4)
        self.assertTrue(self.store.integrity_check()["ok"])

    def test_voucher_unpost_keeps_audit_copy_and_excludes_reports(self):
        self._seed_balanced_ledger()
        before = self.store.tax_summary("2026-07")
        self.assertEqual(before["profit"], 60)
        result = self.store.unpost_voucher("202607-0002")
        self.assertEqual(result["status"], "已反过账")
        self.assertEqual(self.store.tax_summary("2026-07")["profit"], 100)
        raw = [
            row for row in self.store.list_vouchers(include_unposted=True)
            if row["voucher_no"] == "202607-0002"
        ]
        self.assertEqual({row["status"] for row in raw}, {"已反过账"})
        self.assertIn(
            "UNPOSTED_VOUCHER",
            {issue["code"] for issue in self.store.validate("2026-07")},
        )
        self.store.replace_voucher_group("202607-0002", [
            {"科目": "5602 管理费用", "借方": 40, "摘要": "重新入账"},
            {"科目": "1002 银行存款", "贷方": 40, "摘要": "重新入账"},
        ], "2026-07-02")
        self.assertEqual(self.store.tax_summary("2026-07"), before)

    def test_opening_date_and_forced_archive_balance_gate(self):
        settings = self.store.get_settings()
        settings["accounting"]["opening_date"] = "2026-07-01"
        self.store.save_settings(settings)
        with self.assertRaises(ValueError):
            self.store.add_voucher_lines([
                {"科目": "1002 银行存款", "借方": 10},
                {"科目": "5001 主营业务收入", "贷方": 10},
            ], voucher_date="2026-06-30")

        self.store._write_json(self.store.ledger_path, [{
            "id": "bad-1", "voucher_no": "202607-0999", "line_no": 1,
            "date": "2026-07-31", "period": "2026-07", "description": "错误凭证",
            "subject": "1002 银行存款", "debit": 100, "credit": 0,
            "status": "已记账", "source": "manual",
        }])
        issues = self.store.voucher_balance_issues("2026-07")
        self.assertEqual(issues[0]["voucher_no"], "202607-0999")
        self.assertEqual(issues[0]["difference"], 100)
        with self.assertRaises(ValueError):
            self.store.set_period_status("2026-07", "已归档")

    def test_archived_period_blocks_financial_mutation(self):
        self._seed_balanced_ledger()
        self.store.set_period_status("2026-07", "已归档", "申报资料已归档")
        with self.assertRaises(ValueError):
            self.store.add_voucher_lines([
                {"科目": "5602 管理费用", "借方": 10},
                {"科目": "1002 银行存款", "贷方": 10},
            ], voucher_date="2026-07-10")
        with self.assertRaises(ValueError):
            self.store.delete_voucher_numbers(["202607-0001"])
        with self.assertRaises(ValueError):
            self.store.upsert_opening_balance({
                "period": "2026-07", "subject": "1002 银行存款",
                "debit_balance": 100,
            })
        with self.assertRaises(ValueError):
            self.store.import_bank_transactions([{
                "date": "2026-07-10", "direction": "支出", "amount": 10,
            }])
        with self.assertRaises(ValueError):
            self.store.upsert_payroll({
                "period": "2026-07", "employee_name": "张三", "gross_salary": 100,
            })
        with self.assertRaises(ValueError):
            self.store.upsert_fixed_asset({
                "asset_name": "测试电脑", "purchase_date": "2026-07-10",
                "original_cost": 3000, "residual_rate": 0.05, "useful_months": 36,
                "depreciation_start_period": "2026-08",
            })

        self.store.reopen_archived_period("2026-07", "重新打开")
        self.store.delete_voucher_numbers(["202607-0001"])
        self.assertEqual(len(self.store.list_vouchers()), 2)

    def test_opening_balance_totals_and_validation(self):
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "1002 银行存款", "debit_balance": 1000,
        })
        equity = self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "3001 实收资本", "credit_balance": 1000,
        })
        self.assertEqual(self.store.opening_balance_totals("2026-01"), {
            "debit": 1000.0, "credit": 1000.0, "difference": 0.0,
        })
        self.assertNotIn(
            "OPENING_UNBALANCED",
            {issue["code"] for issue in self.store.validate("2026-07")},
        )

        self.store.upsert_opening_balance({
            **equity, "subject": "3001 实收资本", "credit_balance": 900,
        })
        self.assertIn(
            "OPENING_UNBALANCED",
            {issue["code"] for issue in self.store.validate("2026-07")},
        )

    def test_account_balances_roll_forward_current_and_ytd(self):
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "1002 银行存款", "debit_balance": 1000,
        })
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "3001 实收资本", "credit_balance": 1000,
        })
        self.store.add_voucher_lines([
            {"科目": "1002 银行存款", "借方": 500, "摘要": "六月销售收款"},
            {"科目": "5001 主营业务收入", "贷方": 500, "摘要": "六月销售收款"},
        ], voucher_date="2026-06-20")
        self.store.add_voucher_lines([
            {"科目": "5602 管理费用", "借方": 100, "摘要": "七月办公费"},
            {"科目": "1002 银行存款", "贷方": 100, "摘要": "七月办公费"},
        ], voucher_date="2026-07-10")

        rows = self.store.account_balances("2026-07")
        bank = next(row for row in rows if row["subject_code"] == "1002")
        self.assertEqual(bank["year_opening_debit"], 1000)
        self.assertEqual(bank["opening_debit"], 1500)
        self.assertEqual(bank["period_debit"], 0)
        self.assertEqual(bank["period_credit"], 100)
        self.assertEqual(bank["ytd_debit"], 500)
        self.assertEqual(bank["ytd_credit"], 100)
        self.assertEqual(bank["ending_debit"], 1400)

        cash_flow = self.store.cash_flow_summary("2026-07")
        self.assertEqual(cash_flow["month_net_change"], -100)
        self.assertEqual(cash_flow["ytd_net_change"], 400)
        self.assertEqual(cash_flow["period_opening_cash"], 1500)
        self.assertEqual(cash_flow["ending_cash"], 1400)
        self.assertEqual(cash_flow["month_cash_difference"], 0)
        self.assertEqual(cash_flow["ytd_cash_difference"], 0)

    def test_cash_flow_manual_override_and_review_resolution(self):
        self.store.add_voucher_lines([
            {"科目": "5602 管理费用", "借方": 40, "摘要": "综合付款"},
            {"科目": "1601 固定资产", "借方": 60, "摘要": "综合付款"},
            {"科目": "1002 银行存款", "贷方": 100, "摘要": "综合付款"},
        ], voucher_date="2026-07-08")
        self.store.import_bank_transactions([{
            "date": "2026-07-08", "direction": "支出", "amount": 100,
            "summary": "综合付款",
        }])
        self.store.auto_reconcile_bank_transactions("2026-07")
        transaction = self.store.list_bank_transactions("2026-07")[0]

        automatic = self.store.cash_flow_summary("2026-07")
        self.assertEqual(automatic["needs_review_count"], 1)
        self.assertIn(
            "CASH_FLOW_REVIEW",
            {issue["code"] for issue in self.store.validate("2026-07")},
        )

        self.store.set_bank_cash_flow_category(
            transaction["id"], "investing_asset_payment"
        )
        overridden = self.store.cash_flow_summary("2026-07")
        self.assertEqual(overridden["needs_review_count"], 0)
        self.assertEqual(
            overridden["month_entries"][0]["category"],
            "investing_asset_payment",
        )
        self.assertEqual(
            overridden["month_entries"][0]["classification_source"], "手工"
        )
        with self.assertRaises(ValueError):
            self.store.set_bank_cash_flow_category(
                transaction["id"], "operating_sales_receipt"
            )

    def test_cash_flow_excludes_internal_cash_transfer(self):
        self.store.add_voucher_lines([
            {"科目": "1001 库存现金", "借方": 200, "摘要": "银行提现"},
            {"科目": "1002 银行存款", "贷方": 200, "摘要": "银行提现"},
        ], voucher_date="2026-07-05")
        cash_flow = self.store.cash_flow_summary("2026-07")
        self.assertEqual(cash_flow["month_entries"], [])
        self.assertEqual(cash_flow["month_net_change"], 0)
        self.assertEqual(cash_flow["month_cash_difference"], 0)

    def test_month_end_checklist_reports_readiness(self):
        settings = self.store.get_settings()
        settings["company"].update({
            "name": "测试科技有限公司",
            "credit_code": "91110101MA00000001",
            "taxpayer_type": "小规模纳税人",
        })
        self.store.save_settings(settings)
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "1002 银行存款", "debit_balance": 1000,
        })
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "3001 实收资本", "credit_balance": 1000,
        })
        self._seed_balanced_ledger()
        checklist = self.store.month_end_checklist("2026-07")
        self.assertFalse(checklist["ready"])
        self.assertEqual(checklist["blocking_count"], 1)
        close_item = next(
            row for row in checklist["items"] if row["item"] == "损益结转"
        )
        self.assertEqual(close_item["status"], "待处理")

        self.store.post_profit_close_voucher("2026-07")
        checklist = self.store.month_end_checklist("2026-07")
        self.assertTrue(checklist["ready"])
        self.assertEqual(checklist["blocking_count"], 0)
        self.assertEqual(
            {row["item"] for row in checklist["items"]},
            {
                "企业资料", "期初余额", "记账凭证", "银行对账", "工资社保",
                "固定资产折旧", "现金流量分类", "损益结转",
                "税务资格与期间", "税费计提", "试算平衡与报表勾稽",
            },
        )

    def test_profit_close_is_reversible_and_keeps_tax_summary(self):
        self._seed_balanced_ledger()
        summary_before = self.store.tax_summary("2026-07")
        preview = self.store.profit_close_preview("2026-07")
        self.assertFalse(preview["posted"])
        self.assertEqual(preview["income_total"], 100)
        self.assertEqual(preview["expense_total"], 40)
        self.assertEqual(preview["net_profit"], 60)
        self.assertEqual(len(preview["lines"]), 3)

        voucher_no = self.store.post_profit_close_voucher("2026-07")
        posted = self.store.profit_close_preview("2026-07")
        self.assertTrue(posted["posted"])
        self.assertEqual(posted["voucher_no"], voucher_no)
        self.assertEqual(posted["net_profit"], 60)
        self.assertEqual(self.store.tax_summary("2026-07"), summary_before)
        balances = {
            row["subject_code"]: row
            for row in self.store.account_balances("2026-07")
        }
        self.assertEqual(balances["5001"]["ending_credit"], 0)
        self.assertEqual(balances["5602"]["ending_debit"], 0)
        self.assertEqual(balances["3103"]["ending_credit"], 60)
        with self.assertRaises(ValueError):
            self.store.post_profit_close_voucher("2026-07")

        self.assertEqual(
            self.store.unpost_profit_close_voucher("2026-07"), voucher_no
        )
        restored = self.store.profit_close_preview("2026-07")
        self.assertFalse(restored["posted"])
        self.assertEqual(restored["net_profit"], 60)

    def test_bank_import_deduplication_and_reconciliation(self):
        self._seed_balanced_ledger()
        transaction = {
            "date": "2026-07-01", "direction": "收入", "amount": 100,
            "summary": "服务收款", "counterparty": "客户甲",
        }
        result = self.store.import_bank_transactions([transaction, transaction])
        self.assertEqual(result, {"imported": 1, "skipped": 1})
        matched = self.store.auto_reconcile_bank_transactions("2026-07")
        self.assertEqual(matched, {"matched": 1, "unmatched": 0})
        record = self.store.list_bank_transactions("2026-07")[0]
        self.assertEqual(record["voucher_no"], "202607-0001")

        self.store.import_bank_transactions([{
            "date": "2026-07-02", "direction": "支出", "amount": 30,
            "summary": "金额不一致",
        }])
        unmatched = next(
            row for row in self.store.list_bank_transactions("2026-07")
            if row["amount"] == 30
        )
        with self.assertRaises(ValueError):
            self.store.set_bank_match(unmatched["id"], "202607-0002")

    def test_payroll_voucher_is_balanced_and_not_duplicated(self):
        payroll = self.store.upsert_payroll({
            "period": "2026-07", "employee_name": "张三", "gross_salary": 10000,
            "social_personal": 800, "housing_personal": 400, "income_tax": 100,
            "social_company": 1600, "housing_company": 800, "pay_date": "2026-07-31",
        })
        self.assertEqual(payroll["net_salary"], 8700)
        voucher_no = self.store.post_payroll_voucher(payroll["id"])
        lines = [
            row for row in self.store.list_vouchers() if row["voucher_no"] == voucher_no
        ]
        self.assertEqual(sum(row["debit"] for row in lines), 12400)
        self.assertEqual(sum(row["credit"] for row in lines), 12400)
        with self.assertRaises(ValueError):
            self.store.post_payroll_voucher(payroll["id"])

    def test_depreciation_final_month_rounding_and_duplicate_prevention(self):
        asset = self.store.upsert_fixed_asset({
            "asset_name": "办公电脑", "category": "电子设备",
            "purchase_date": "2025-12-20", "original_cost": 1000,
            "residual_rate": 0.05, "useful_months": 36,
            "depreciation_start_period": "2026-01",
        })
        final_month = self.store.depreciation_schedule("2028-12")[0]
        self.assertEqual(final_month["monthly_depreciation"], 26.39)
        self.assertEqual(final_month["depreciation_amount"], 26.35)
        self.assertEqual(final_month["accumulated_depreciation"], 950)
        self.assertEqual(final_month["net_book_value"], 50)
        voucher_no = self.store.post_depreciation_voucher("2028-12")
        self.assertTrue(voucher_no)
        self.assertIn("2028-12", next(
            row for row in self.store.list_fixed_assets() if row["id"] == asset["id"]
        )["posted_periods"])
        with self.assertRaises(ValueError):
            self.store.post_depreciation_voucher("2028-12")

    def test_multi_sheet_export_structure_and_formulas(self):
        settings = self.store.get_settings()
        settings["company"].update({
            "name": "测试科技有限公司", "credit_code": "91110101MA00000001",
        })
        self.store.save_settings(settings)
        self._seed_balanced_ledger()
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "1002 银行存款", "debit_balance": 1000,
        })
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "3001 实收资本", "credit_balance": 1000,
        })
        self.store.import_bank_transactions([{
            "date": "2026-07-01", "direction": "收入", "amount": 100,
            "summary": "服务收款",
        }])
        self.store.upsert_payroll({
            "period": "2026-07", "employee_name": "张三", "gross_salary": 1000,
            "social_personal": 80, "income_tax": 20, "social_company": 160,
        })
        self.store.upsert_fixed_asset({
            "asset_name": "办公电脑", "purchase_date": "2026-06-20",
            "original_cost": 3600, "residual_rate": 0, "useful_months": 36,
            "depreciation_start_period": "2026-07",
        })
        target = self.root / "财税申报辅助.xlsx"
        export_finance_workbook(self.store, target, "2026-07")
        self.assertTrue(target.exists())

        workbook = load_workbook(target, data_only=False)
        self.assertEqual(workbook.sheetnames, [
            "月度总览", "使用说明", "科目目录", "期初余额", "记账凭证", "总分类账", "明细分类账",
            "现金银行日记账", "发票台账", "银行对账", "工资社保",
            "固定资产", "折旧明细", "科目余额表", "报表取数底稿",
            "会小企01资产负债表", "会小企02利润表", "会小企03现金流量表", "利润表辅助",
            "资产负债辅助", "税务期间", "纳税调整", "增值税测算", "个税测算",
            "印花税准备", "所得税预缴", "所得税测算", "年度汇算准备", "申报校验", "政策依据",
        ])
        self.assertTrue(str(workbook["科目余额表"]["C6"].value).startswith("=MAX((SUMIFS"))
        self.assertEqual(workbook["月度总览"]["A7"].value, "='利润表辅助'!B6")
        self.assertEqual(
            workbook["月度总览"]["C7"].value,
            "='利润表辅助'!B7+'利润表辅助'!B8",
        )
        self.assertTrue(
            str(workbook["月度总览"]["G7"].value).startswith("=A11+C11+E11+'印花税准备'!")
        )
        self.assertIn("暂不建议申报", workbook["月度总览"]["A4"].value)
        self.assertEqual(workbook["月度总览"]["A3"].value, LEGAL_NOTICE_SUMMARY)
        self.assertEqual(workbook["政策依据"]["A2"].value, LEGAL_NOTICE_SUMMARY)
        instruction_values = {
            cell.value for row in workbook["使用说明"].iter_rows() for cell in row
            if cell.value
        }
        self.assertIn(LEGAL_NOTICE_SUMMARY, instruction_values)
        self.assertTrue(str(workbook["科目余额表"]["E6"].value).startswith("=SUMIFS"))
        self.assertEqual(workbook["科目余额表"]["G6"].value, "=MAX(C6+E6-D6-F6,0)")
        self.assertEqual(workbook["工资社保"]["G6"].value, "=C6-D6-E6-F6")
        self.assertEqual(workbook["固定资产"]["H6"].value, "=ROUND(D6*(1-E6)/F6,2)")
        self.assertEqual(
            workbook["利润表辅助"]["B9"].value,
            "='会小企02利润表'!D34",
        )
        self.assertEqual(workbook["资产负债辅助"]["B11"].value, "=B6-B7-B10")
        self.assertEqual(workbook["所得税测算"]["B11"].value, "='所得税预缴'!B14")
        self.assertEqual(workbook["科目目录"].max_row, 71)
        self.assertEqual(workbook["税务期间"]["A6"].value, "增值税期间")
        self.assertEqual(workbook["所得税预缴"]["B14"].value, "=MAX(B12-B13,0)")
        self.assertEqual(workbook["年度汇算准备"]["B14"].value, "=B12-B13")
        self.assertTrue(
            str(workbook["会小企01资产负债表"]["C5"].value).startswith(
                "='报表取数底稿'!"
            )
        )
        self.assertTrue(
            str(workbook["会小企02利润表"]["C36"].value).startswith(
                "='报表取数底稿'!"
            )
        )
        self.assertTrue(
            str(workbook["会小企03现金流量表"]["C26"].value).startswith(
                "='报表取数底稿'!"
            )
        )
        self.assertEqual(workbook["银行对账"]["K5"].value, "现金流项目")
        self.assertEqual(workbook["银行对账"]["L5"].value, "分类来源")
        self.assertTrue(
            str(workbook["总分类账"]["C6"].value).startswith(
                "='科目余额表'!"
            )
        )
        self.assertTrue(
            str(workbook["明细分类账"]["M6"].value).startswith("=SUMIFS")
        )
        self.assertEqual(workbook["明细分类账"]["J6"].value, "=ABS(M6)")
        self.assertTrue(workbook["明细分类账"].column_dimensions["M"].hidden)
        self.assertEqual(workbook["现金银行日记账"]["J5"].value, "现金流项目")
        self.assertEqual(workbook["现金银行日记账"]["I6"].value, "=ABS(M6)")

        profit_labels = {
            cell.value for cell in workbook["会小企02利润表"]["A"] if cell.value
        }
        self.assertIn("营业税", profit_labels)
        self.assertIn(
            "教育费附加、矿产资源补偿费、排污费",
            profit_labels,
        )
        self.assertNotIn("研发费用", profit_labels)
        self.assertNotIn("企业利润表", workbook.sheetnames)

    def test_governmental_export_does_not_include_enterprise_statements(self):
        governmental = FinanceDataStore(
            self.root / "governmental", "governmental", "行政事业单位会计"
        )
        target = self.root / "事业单位财税辅助.xlsx"
        export_finance_workbook(governmental, target, "2026-07")
        workbook = load_workbook(target, data_only=False)
        self.assertNotIn("报表取数底稿", workbook.sheetnames)
        self.assertNotIn("会小企01资产负债表", workbook.sheetnames)
        self.assertNotIn("企业资产负债表", workbook.sheetnames)
        self.assertEqual(workbook["银行对账"].max_column, 11)

    def test_export_rolls_prior_activity_into_report_period_opening(self):
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "1002 银行存款", "debit_balance": 1000,
        })
        self.store.upsert_opening_balance({
            "period": "2026-01", "subject": "3001 实收资本", "credit_balance": 1000,
        })
        self.store.add_voucher_lines([
            {"科目": "1002 银行存款", "借方": 500, "摘要": "六月收款"},
            {"科目": "5001 主营业务收入", "贷方": 500, "摘要": "六月收款"},
        ], voucher_date="2026-06-20")
        self.store.add_voucher_lines([
            {"科目": "5602 管理费用", "借方": 100, "摘要": "七月费用"},
            {"科目": "1002 银行存款", "贷方": 100, "摘要": "七月费用"},
        ], voucher_date="2026-07-10")

        target = self.root / "七月滚动余额.xlsx"
        export_finance_workbook(self.store, target, "2026-07")
        workbook = load_workbook(target, data_only=False)
        journal_periods = {
            workbook["记账凭证"].cell(row, 2).value for row in range(6, 10)
        }
        self.assertEqual(journal_periods, {"2026-06", "2026-07"})
        bank_row = next(
            row for row in range(6, 10)
            if workbook["科目余额表"].cell(row, 1).value == "1002"
        )
        opening_formula = workbook["科目余额表"].cell(bank_row, 3).value
        current_formula = workbook["科目余额表"].cell(bank_row, 5).value
        self.assertIn('"<2026-07"', opening_formula)
        self.assertIn('"2026-07"', current_formula)


if __name__ == "__main__":
    unittest.main()
