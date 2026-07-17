#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from legal_notice import LEGAL_NOTICE_SUMMARY, policy_snapshot_text


DARK = "17365D"
BLUE = "1F4E78"
TEAL = "0F6B78"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FCE4D6"
GRAY = "E7E6E6"
TEXT = "202020"
INPUT_BLUE = "0000FF"
MONEY_FMT = '#,##0.00;[Red](#,##0.00);-'
PERCENT_FMT = '0.00%'
DATE_FMT = 'yyyy-mm-dd'
THIN_GRAY = Side(style="thin", color="D9E1F2")


def _safe_date(value):
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return text


def _money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _filtered(records: Iterable[Dict], period: Optional[str], date_key: str) -> List[Dict]:
    result = list(records)
    if period:
        result = [row for row in result if str(row.get(date_key, ""))[:7] == period]
    return result


def _sheet_base(ws, freeze: Optional[str] = None):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 90


def _title(ws, title: str, subtitle: str, end_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.font = Font(name="微软雅黑", size=16, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    cell = ws.cell(2, 1, subtitle)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.font = Font(name="微软雅黑", size=9, color=TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28


def _headers(ws, row: int, headers: Sequence[str]):
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row, col, value)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=DARK))
    ws.row_dimensions[row].height = 26


def _table(ws, start_row: int, end_row: int, end_col: int, name: str):
    if end_row <= start_row:
        return
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _set_widths(ws, widths: Sequence[float]):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _style_data(ws, start_row: int, end_row: int, end_col: int,
                money_cols: Sequence[int] = (), date_cols: Sequence[int] = ()):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=end_col):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=9, color=TEXT)
            cell.alignment = Alignment(
                horizontal="right" if cell.column in money_cols else "left",
                vertical="center", wrap_text=cell.column not in money_cols,
            )
            if cell.column in money_cols:
                cell.number_format = MONEY_FMT
            if cell.column in date_cols and cell.value:
                cell.number_format = DATE_FMT


def _formula_sum(sheet: str, cells: Sequence[str]) -> str:
    if not cells:
        return "=0"
    refs = [f"'{sheet}'!{cell}" for cell in cells]
    return "=SUM(" + ",".join(refs) + ")"


def _formula_net(sheet: str, positive_cells: Sequence[str],
                 negative_cells: Sequence[str]) -> str:
    positive = _formula_sum(sheet, positive_cells)[1:]
    negative = _formula_sum(sheet, negative_cells)[1:]
    return f"={positive}-{negative}"


def _enterprise_statement_names(store) -> Dict[str, str]:
    if store.is_small_enterprise_standard:
        return {
            "balance": "会小企01资产负债表",
            "profit": "会小企02利润表",
            "cash_flow": "会小企03现金流量表",
            "caption": "小企业会计准则（会小企报表）",
        }
    return {
        "balance": "企业资产负债表",
        "profit": "企业利润表",
        "cash_flow": "企业现金流量表",
        "caption": "企业会计准则通用格式",
    }


def _create_enterprise_standard_reports(sheets: Dict[str, object], store,
                                        company_name: str, period: str):
    """Create auditable statements for the ledger's fixed accounting standard."""
    account_rows = store.account_balances(period)
    cash_flow = store.cash_flow_summary(period)
    report_names = _enterprise_statement_names(store)
    small_standard = store.is_small_enterprise_standard
    production_cost_codes = (
        ("4001", "4101", "4401", "4403")
        if small_standard else ("5001",)
    )
    operating_revenue_codes = ("5001", "5051") if small_standard else ("6001", "6051")
    other_income_codes = ("5111", "5301") if small_standard else ("6111", "6301")
    operating_cost_codes = ("5401", "5402") if small_standard else ("6401", "6402")
    tax_surcharge_codes = ("5403",) if small_standard else ("6403",)
    selling_expense_codes = ("5601",) if small_standard else ("6601",)
    management_expense_codes = ("5602",) if small_standard else ("6602",)
    finance_expense_codes = ("5603",) if small_standard else ("6603",)
    nonoperating_expense_codes = ("5711",) if small_standard else ("6711",)
    income_tax_codes = ("5801",) if small_standard else ("6801",)
    equity_codes = (
        ("3001",), ("3002",), ("3101",), ("3103", "3104")
    ) if small_standard else (
        ("4001",), ("4002",), ("4101",), ("4103", "4104")
    )
    ws = sheets["报表取数底稿"]
    _sheet_base(ws, "A6")
    _title(
        ws, f"{report_names['caption']}财务报表取数底稿",
        f"企业：{company_name} | 期间：{period} | 标准报表均引用本底稿",
        17,
    )
    account_headers = [
        "科目编码", "科目名称", "年初借方", "年初贷方", "本月期初借方", "本月期初贷方",
        "本月借方", "本月贷方", "本年借方", "本年贷方", "期末借方", "期末贷方",
    ]
    _headers(ws, 5, account_headers)
    for row_index, item in enumerate(account_rows, start=6):
        values = [
            item.get("subject_code", ""), item.get("subject", ""),
            _money(item.get("year_opening_debit")), _money(item.get("year_opening_credit")),
            _money(item.get("opening_debit")), _money(item.get("opening_credit")),
            _money(item.get("period_debit")), _money(item.get("period_credit")),
            _money(item.get("ytd_debit")), _money(item.get("ytd_credit")),
            _money(item.get("ending_debit")), _money(item.get("ending_credit")),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    account_end = max(6, len(account_rows) + 5)
    _style_data(ws, 6, account_end, 12, money_cols=tuple(range(3, 13)))
    _set_widths(
        ws,
        [13, 28, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 8, 30, 17, 17, 44],
    )
    _table(ws, 5, len(account_rows) + 5, 12, "ReportAccountBalances")

    def total(col: str, codes=(), contains=()) -> str:
        terms = [
            f'SUMIFS(${col}$6:${col}${account_end},$A$6:$A${account_end},"{code}")'
            for code in codes
        ]
        terms.extend(
            f'SUMIFS(${col}$6:${col}${account_end},$B$6:$B${account_end},"*{word}*")'
            for word in contains
        )
        return "SUM(" + ",".join(terms) + ")" if terms else "0"

    def net(debit_col: str, credit_col: str, codes=(), contains=()) -> str:
        return (
            f"={total(debit_col, codes, contains)}-"
            f"{total(credit_col, codes, contains)}"
        )

    def credit_net(debit_col: str, credit_col: str, codes=(), contains=()) -> str:
        return (
            f"={total(credit_col, codes, contains)}-"
            f"{total(debit_col, codes, contains)}"
        )

    def gross(col: str, codes=(), contains=()) -> str:
        return f"={total(col, codes, contains)}"

    section_fill = PatternFill("solid", fgColor=TEAL)
    section_font = Font(name="微软雅黑", size=10, bold=True, color=WHITE)

    # 会小企 01 资产负债表取数。
    bs_header = account_end + 3
    for col, value in enumerate(["行次", "资产项目", "期末余额", "年初余额", "取数规则"], start=13):
        cell = ws.cell(bs_header, col, value)
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bs_map: Dict[int, int] = {}

    def put_bs(line: int, item: str, ending_formula: str,
               opening_formula: str, rule: str):
        row_index = bs_header + line
        bs_map[line] = row_index
        values = [line, item, ending_formula, opening_formula, rule]
        for offset, value in enumerate(values, start=13):
            ws.cell(row_index, offset, value)
        for col in (15, 16):
            ws.cell(row_index, col).number_format = MONEY_FMT

    asset_current = {
        1: ("货币资金", ("1001", "1002", "1012"), "库存现金、银行存款和其他货币资金"),
        2: ("短期投资", ("1101",), "短期投资"),
        3: ("应收票据", ("1121",), "应收票据"),
        4: ("应收账款", ("1122",), "应收账款净额"),
        5: ("预付账款", ("1123",), "预付账款"),
        6: ("应收股利", ("1131",), "应收股利"),
        7: ("应收利息", ("1132",), "应收利息"),
        8: ("其他应收款", ("1221",), "其他应收款"),
        9: ("存货", ("1401", "1402", "1403", "1404", "1405", "1407", "1408", "1411", "1421") + production_cost_codes, "存货相关科目净额"),
        10: ("其中：原材料", ("1403",), "原材料"),
        11: ("在产品等", production_cost_codes, "生产、制造及工程作业成本期末余额"),
        12: ("库存商品", ("1405",), "库存商品"),
        13: ("周转材料", ("1411",), "周转材料"),
        14: ("其他流动资产", ("1901",), "其他流动资产；自定义科目需复核"),
    }
    for line, (item, codes, rule) in asset_current.items():
        put_bs(line, item, net("K", "L", codes), net("C", "D", codes), rule)
    current_liability_codes = (
        "2001", "2201", "2202", "2203", "2211", "2221", "2231", "2232",
        "2241", "2402",
    )
    put_bs(
        14, "其他流动资产",
        f"={net('K', 'L', ('1901',))[1:]}+{total('K', current_liability_codes)}",
        f"={net('C', 'D', ('1901',))[1:]}+{total('C', current_liability_codes)}",
        "其他流动资产及流动负债科目的借方余额重分类",
    )
    put_bs(
        15, "流动资产合计",
        "=SUM(" + ",".join(f"O{bs_map[line]}" for line in (*range(1, 10), 14)) + ")",
        "=SUM(" + ",".join(f"P{bs_map[line]}" for line in (*range(1, 10), 14)) + ")",
        "行1至行9加行14",
    )
    noncurrent = {
        16: ("长期债券投资", ("1501",), (), "长期债券投资"),
        17: ("长期股权投资", ("1511",), (), "长期股权投资"),
        21: ("在建工程", ("1604",), (), "在建工程"),
        22: ("工程物资", ("1605",), (), "工程物资"),
        23: ("固定资产清理", ("1606",), (), "固定资产清理"),
        24: ("生产性生物资产", ("1621",), ("1622",), "原值减累计折旧"),
        25: ("无形资产", ("1701",), ("1702",), "原值减累计摊销"),
        26: ("开发支出", ("4301",) if small_standard else ("5301",), (), "研发支出中的资本化余额"),
        27: ("长期待摊费用", ("1801",), (), "长期待摊费用"),
        28: ("其他非流动资产", ("1902", "1903"), (), "其他非流动资产；自定义科目需复核"),
    }
    put_bs(18, "固定资产原价", gross("K", ("1601",)), gross("C", ("1601",)), "固定资产借方余额")
    put_bs(19, "减：累计折旧", gross("L", ("1602",)), gross("D", ("1602",)), "累计折旧贷方余额")
    put_bs(20, "固定资产账面价值", f"=O{bs_map[18]}-O{bs_map[19]}", f"=P{bs_map[18]}-P{bs_map[19]}", "行18减行19")
    for line, (item, positive_codes, negative_codes, rule) in noncurrent.items():
        ending = net("K", "L", positive_codes)
        opening = net("C", "D", positive_codes)
        if negative_codes:
            ending = f"={ending[1:]}-{total('L', negative_codes)}+{total('K', negative_codes)}"
            opening = f"={opening[1:]}-{total('D', negative_codes)}+{total('C', negative_codes)}"
        put_bs(line, item, ending, opening, rule)
    noncurrent_lines = (16, 17, 20, 21, 22, 23, 24, 25, 26, 27, 28)
    put_bs(
        29, "非流动资产合计",
        "=SUM(" + ",".join(f"O{bs_map[line]}" for line in noncurrent_lines) + ")",
        "=SUM(" + ",".join(f"P{bs_map[line]}" for line in noncurrent_lines) + ")",
        "非流动资产项目合计",
    )
    put_bs(30, "资产总计", f"=O{bs_map[15]}+O{bs_map[29]}", f"=P{bs_map[15]}+P{bs_map[29]}", "行15加行29")

    liabilities = {
        31: ("短期借款", ("2001",), "短期借款"),
        32: ("应付票据", ("2201",), "应付票据"),
        33: ("应付账款", ("2202",), "应付账款"),
        34: ("预收账款", ("2203",), "预收账款"),
        35: ("应付职工薪酬", ("2211",), "应付职工薪酬"),
        36: ("应交税费", ("2221",), "应交税费"),
        37: ("应付利息", ("2231",), "应付利息"),
        38: ("应付利润", ("2232",), "应付利润"),
        39: ("其他应付款", ("2241",), "其他应付款"),
        40: ("其他流动负债", ("2402",), "其他流动负债；自定义科目需复核"),
        42: ("长期借款", ("2501",), "长期借款"),
        43: ("长期应付款", ("2701",), "长期应付款"),
        44: ("递延收益", ("2401",), "递延收益"),
        45: ("其他非流动负债", ("2801",), "其他非流动负债；自定义科目需复核"),
    }
    for line, (item, codes, rule) in liabilities.items():
        put_bs(
            line, item, gross("L", codes), gross("D", codes),
            f"{rule}贷方余额；借方余额重分类至流动资产",
        )
    put_bs(
        41, "流动负债合计",
        "=SUM(" + ",".join(f"O{bs_map[line]}" for line in range(31, 41)) + ")",
        "=SUM(" + ",".join(f"P{bs_map[line]}" for line in range(31, 41)) + ")",
        "行31至行40",
    )
    put_bs(
        46, "非流动负债合计",
        "=SUM(" + ",".join(f"O{bs_map[line]}" for line in range(42, 46)) + ")",
        "=SUM(" + ",".join(f"P{bs_map[line]}" for line in range(42, 46)) + ")",
        "行42至行45",
    )
    put_bs(47, "负债合计", f"=O{bs_map[41]}+O{bs_map[46]}", f"=P{bs_map[41]}+P{bs_map[46]}", "行41加行46")
    equity = {
        48: ("实收资本（或股本）", equity_codes[0], "实收资本或股本"),
        49: ("资本公积", equity_codes[1], "资本公积"),
        50: ("盈余公积", equity_codes[2], "盈余公积"),
    }
    if small_standard:
        mapped_statement_codes = {
            code
            for _line, (_item, codes, _rule) in asset_current.items()
            for code in codes
        }
        mapped_statement_codes.update({"1601", "1602"})
        for _line, (_item, positive_codes, negative_codes, _rule) in noncurrent.items():
            mapped_statement_codes.update(positive_codes)
            mapped_statement_codes.update(negative_codes)
        mapped_statement_codes.update(current_liability_codes)
        for _line, (_item, codes, _rule) in liabilities.items():
            mapped_statement_codes.update(codes)
        for _line, (_item, codes, _rule) in equity.items():
            mapped_statement_codes.update(codes)
        mapped_statement_codes.update(equity_codes[3])
        expected_statement_codes = {
            str(account.get("code", ""))
            for account in store.all_accounts()
            if account.get("class") != "损益类"
        }
        missing_statement_codes = sorted(
            expected_statement_codes - mapped_statement_codes
        )
        if missing_statement_codes:
            raise ValueError(
                "会小企资产负债表取数遗漏官方科目："
                + "、".join(missing_statement_codes)
            )
    for line, (item, codes, rule) in equity.items():
        put_bs(line, item, credit_net("K", "L", codes), credit_net("C", "D", codes), rule)
    profit_loss_codes = (
        operating_revenue_codes + other_income_codes + operating_cost_codes
        + tax_surcharge_codes + selling_expense_codes + management_expense_codes
        + finance_expense_codes + nonoperating_expense_codes + income_tax_codes
    )
    undistributed_codes = equity_codes[3]
    ending_unclosed = credit_net("K", "L", profit_loss_codes)
    opening_unclosed = credit_net("C", "D", profit_loss_codes)
    ending_equity = credit_net("K", "L", undistributed_codes)
    opening_equity = credit_net("C", "D", undistributed_codes)
    put_bs(
        51, "未分配利润",
        f"={ending_equity[1:]}+{ending_unclosed[1:]}",
        f"={opening_equity[1:]}+{opening_unclosed[1:]}",
        "利润分配/本年利润加尚未结转的损益余额",
    )
    put_bs(
        52, "所有者权益（或股东权益）合计",
        "=SUM(" + ",".join(f"O{bs_map[line]}" for line in range(48, 52)) + ")",
        "=SUM(" + ",".join(f"P{bs_map[line]}" for line in range(48, 52)) + ")",
        "行48至行51",
    )
    put_bs(53, "负债和所有者权益总计", f"=O{bs_map[47]}+O{bs_map[52]}", f"=P{bs_map[47]}+P{bs_map[52]}", "行47加行52")

    bs_end = bs_header + 53
    for row in ws.iter_rows(min_row=bs_header + 1, max_row=bs_end, min_col=13, max_col=17):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=9, color=TEXT)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.conditional_formatting.add(
        f"O{bs_map[53]}",
        CellIsRule(
            operator="notBetween",
            formula=[f"O{bs_map[30]}-0.01", f"O{bs_map[30]}+0.01"],
            fill=PatternFill("solid", fgColor=LIGHT_RED),
        ),
    )

    # 会小企 02 利润表取数。
    profit_header = bs_end + 3
    for col, value in enumerate(["行次", "项目", "本年累计金额", "本月金额", "取数规则"], start=13):
        cell = ws.cell(profit_header, col, value)
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    profit_map: Dict[int, int] = {}

    def put_profit(line: int, item: str, ytd_formula: str,
                   month_formula: str, rule: str):
        row_index = profit_header + line
        profit_map[line] = row_index
        for offset, value in enumerate(
            [line, item, ytd_formula, month_formula, rule], start=13
        ):
            ws.cell(row_index, offset, value)
        for col in (15, 16):
            ws.cell(row_index, col).number_format = MONEY_FMT

    def movement(codes=(), contains=(), income=False):
        function = credit_net if income else net
        return (
            function("I", "J", codes, contains),
            function("G", "H", codes, contains),
        )

    ytd, month = movement(operating_revenue_codes, income=True)
    put_profit(1, "一、营业收入", ytd, month, "主营业务收入和其他业务收入")
    ytd, month = movement(operating_cost_codes)
    put_profit(2, "减：营业成本", ytd, month, "主营业务成本和其他业务成本")
    ytd, month = movement(tax_surcharge_codes)
    put_profit(3, "营业税金及附加", ytd, month, "税金及附加")
    profit_contains = {
        4: ("其中：消费税", ("消费税",)),
        5: ("营业税", ("营业税",)),
        6: ("城市维护建设税", ("城市维护建设税", "城建税")),
        7: ("资源税", ("资源税",)),
        8: ("土地增值税", ("土地增值税",)),
        9: ("城镇土地使用税、房产税、车船税、印花税", ("土地使用税", "房产税", "车船税", "印花税")),
        10: ("教育费附加、矿产资源补偿费、排污费", ("教育费附加", "矿产资源补偿费", "排污费")),
    }
    for line, (item, words) in profit_contains.items():
        ytd, month = movement(contains=words)
        put_profit(line, item, ytd, month, "按科目名称明细取数")
    ytd, month = movement(selling_expense_codes)
    put_profit(11, "销售费用", ytd, month, "销售费用")
    ytd, month = movement(contains=("商品维修费",))
    put_profit(12, "其中：商品维修费", ytd, month, "按科目名称明细取数")
    ytd, month = movement(contains=("广告费", "业务宣传费"))
    put_profit(13, "广告费和业务宣传费", ytd, month, "按科目名称明细取数")
    ytd, month = movement(management_expense_codes)
    put_profit(14, "管理费用", ytd, month, "管理费用")
    for line, item, words in (
        (15, "其中：开办费", ("开办费",)),
        (16, "业务招待费", ("业务招待费",)),
        (17, "研究费用", ("研发费", "研究费用")),
    ):
        ytd, month = movement(contains=words)
        put_profit(line, item, ytd, month, "按科目名称明细取数")
    ytd, month = movement(finance_expense_codes)
    put_profit(18, "财务费用", ytd, month, "财务费用")
    ytd, month = movement(contains=("利息费用", "利息支出"))
    put_profit(19, "其中：利息费用（收入以“-”号填列）", ytd, month, "按科目名称明细取数")
    ytd, month = movement(("5111",) if small_standard else ("6111",), income=True)
    put_profit(20, "加：投资收益（损失以“-”号填列）", ytd, month, "投资收益")
    put_profit(
        21, "二、营业利润（亏损以“-”号填列）",
        f"=O{profit_map[1]}-O{profit_map[2]}-O{profit_map[3]}-O{profit_map[11]}-O{profit_map[14]}-O{profit_map[18]}+O{profit_map[20]}",
        f"=P{profit_map[1]}-P{profit_map[2]}-P{profit_map[3]}-P{profit_map[11]}-P{profit_map[14]}-P{profit_map[18]}+P{profit_map[20]}",
        "按官方勾稽公式",
    )
    ytd, month = movement(("5301",) if small_standard else ("6301",), income=True)
    put_profit(22, "加：营业外收入", ytd, month, "营业外收入")
    ytd, month = movement(contains=("政府补助",), income=True)
    put_profit(23, "其中：政府补助", ytd, month, "按科目名称明细取数")
    ytd, month = movement(nonoperating_expense_codes)
    put_profit(24, "减：营业外支出", ytd, month, "营业外支出")
    loss_items = {
        25: ("其中：坏账损失", ("坏账损失",)),
        26: ("无法收回的长期债券投资损失", ("长期债券投资损失",)),
        27: ("无法收回的长期股权投资损失", ("长期股权投资损失",)),
        28: ("自然灾害等不可抗力因素造成的损失", ("自然灾害", "不可抗力")),
        29: ("税收滞纳金", ("税收滞纳金", "滞纳金")),
    }
    for line, (item, words) in loss_items.items():
        ytd, month = movement(contains=words)
        put_profit(line, item, ytd, month, "按科目名称明细取数")
    put_profit(
        30, "三、利润总额（亏损总额以“-”号填列）",
        f"=O{profit_map[21]}+O{profit_map[22]}-O{profit_map[24]}",
        f"=P{profit_map[21]}+P{profit_map[22]}-P{profit_map[24]}",
        "行21加行22减行24",
    )
    ytd, month = movement(income_tax_codes)
    put_profit(31, "减：所得税费用", ytd, month, "所得税费用")
    put_profit(
        32, "四、净利润（净亏损以“-”号填列）",
        f"=O{profit_map[30]}-O{profit_map[31]}",
        f"=P{profit_map[30]}-P{profit_map[31]}",
        "行30减行31",
    )
    profit_end = profit_header + 32

    # 会小企 03 现金流量表取数。
    cash_header = profit_end + 3
    for col, value in enumerate(["行次", "项目", "本年累计金额", "本月金额", "取数规则"], start=13):
        cell = ws.cell(cash_header, col, value)
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cash_map: Dict[int, int] = {}

    def put_cash(line: int, item: str, category: str = "", rule: str = ""):
        row_index = cash_header + line
        cash_map[line] = row_index
        ytd_value = _money(cash_flow["ytd"].get(category)) if category else None
        month_value = _money(cash_flow["month"].get(category)) if category else None
        for offset, value in enumerate(
            [line, item, ytd_value, month_value, rule], start=13
        ):
            ws.cell(row_index, offset, value)
        for col in (15, 16):
            ws.cell(row_index, col).number_format = MONEY_FMT

    cash_items = {
        1: ("销售产成品、商品、提供劳务收到的现金", "operating_sales_receipt"),
        2: ("收到其他与经营活动有关的现金", "operating_other_receipt"),
        3: ("购买原材料、商品、接受劳务支付的现金", "operating_purchase_payment"),
        4: ("支付的职工薪酬", "operating_payroll_payment"),
        5: ("支付的税费", "operating_tax_payment"),
        6: ("支付其他与经营活动有关的现金", "operating_other_payment"),
        8: ("收回短期投资、长期债券投资和长期股权投资收到的现金", "investing_recovery_receipt"),
        9: ("取得投资收益收到的现金", "investing_income_receipt"),
        10: ("处置固定资产、无形资产和其他非流动资产收回的现金净额", "investing_disposal_receipt"),
        11: ("短期投资、长期债券投资和长期股权投资支付的现金", "investing_investment_payment"),
        12: ("购建固定资产、无形资产和其他非流动资产支付的现金", "investing_asset_payment"),
        14: ("取得借款收到的现金", "financing_borrowing_receipt"),
        15: ("吸收投资者投资收到的现金", "financing_capital_receipt"),
        16: ("偿还借款本金支付的现金", "financing_principal_payment"),
        17: ("偿还借款利息支付的现金", "financing_interest_payment"),
        18: ("分配利润支付的现金", "financing_distribution_payment"),
    }
    for line, (item, category) in cash_items.items():
        put_cash(line, item, category, "现金/银行凭证分录分类汇总")
    put_cash(7, "经营活动产生的现金流量净额", rule="行1加行2减行3至行6")
    put_cash(13, "投资活动产生的现金流量净额", rule="行8至行10减行11和行12")
    put_cash(19, "筹资活动产生的现金流量净额", rule="行14加行15减行16至行18")
    put_cash(20, "四、现金净增加额", rule="行7加行13加行19")
    put_cash(21, "加：期初现金余额", rule="账面货币资金期初余额")
    put_cash(22, "五、期末现金余额", rule="行20加行21")
    for line, positive, negative in (
        (7, (1, 2), (3, 4, 5, 6)),
        (13, (8, 9, 10), (11, 12)),
        (19, (14, 15), (16, 17, 18)),
    ):
        row_index = cash_map[line]
        for col_letter, col in (("O", 15), ("P", 16)):
            positive_formula = "+".join(f"{col_letter}{cash_map[item]}" for item in positive)
            negative_formula = "-".join(f"{col_letter}{cash_map[item]}" for item in negative)
            ws.cell(row_index, col, f"={positive_formula}-{negative_formula}")
    for col_letter, col in (("O", 15), ("P", 16)):
        ws.cell(cash_map[20], col, f"={col_letter}{cash_map[7]}+{col_letter}{cash_map[13]}+{col_letter}{cash_map[19]}")
    ws.cell(cash_map[21], 15, cash_flow["year_opening_cash"])
    ws.cell(cash_map[21], 16, cash_flow["period_opening_cash"])
    ws.cell(cash_map[22], 15, f"=O{cash_map[20]}+O{cash_map[21]}")
    ws.cell(cash_map[22], 16, f"=P{cash_map[20]}+P{cash_map[21]}")
    cash_end = cash_header + 22

    for start, end in ((profit_header + 1, profit_end), (cash_header + 1, cash_end)):
        for row in ws.iter_rows(min_row=start, max_row=end, min_col=13, max_col=17):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=9, color=TEXT)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.auto_filter.ref = f"A5:L{account_end}"

    def standard_header(report_ws, title: str, form_no: str, columns: int,
                        headers: Sequence[str], landscape=False):
        _sheet_base(report_ws, "A5")
        report_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        cell = report_ws.cell(1, 1, title)
        cell.font = Font(name="微软雅黑", size=16, bold=True, color=TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        report_ws.row_dimensions[1].height = 30
        report_ws.cell(2, 1, f"编制单位：{company_name}")
        report_ws.cell(2, max(2, columns // 2), f"期间：{period}")
        report_ws.cell(2, columns, f"{form_no}    单位：元")
        report_ws.cell(2, columns).alignment = Alignment(horizontal="right")
        _headers(report_ws, 4, headers)
        report_ws.page_setup.orientation = "landscape" if landscape else "portrait"
        report_ws.page_setup.paperSize = report_ws.PAPERSIZE_A4
        report_ws.page_margins.left = 0.25
        report_ws.page_margins.right = 0.25
        report_ws.page_margins.top = 0.45
        report_ws.page_margins.bottom = 0.45

    def style_statement(report_ws, end_row: int, end_col: int, money_cols: Sequence[int]):
        border = Border(
            left=Side(style="thin", color="808080"),
            right=Side(style="thin", color="808080"),
            top=Side(style="thin", color="808080"),
            bottom=Side(style="thin", color="808080"),
        )
        for row in report_ws.iter_rows(min_row=5, max_row=end_row, max_col=end_col):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=9, color=TEXT)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right" if cell.column in money_cols else "left",
                    vertical="center", wrap_text=True,
                )
                if cell.column in money_cols:
                    cell.number_format = MONEY_FMT
        report_ws.print_area = f"A1:{get_column_letter(end_col)}{end_row}"

    # 资产负债表采用双栏项目布局。
    report_ws = sheets[report_names["balance"]]
    standard_header(
        report_ws, "资产负债表", report_names["caption"], 8,
        ["资产", "行次", "期末余额", "年初余额", "负债和所有者权益", "行次", "期末余额", "年初余额"],
        landscape=True,
    )
    asset_lines = list(range(1, 31))
    liability_lines = list(range(31, 54))
    for offset in range(max(len(asset_lines), len(liability_lines))):
        row_index = 5 + offset
        if offset < len(asset_lines):
            line = asset_lines[offset]
            report_ws.cell(row_index, 1, ws.cell(bs_map[line], 14).value)
            report_ws.cell(row_index, 2, line)
            report_ws.cell(row_index, 3, f"='报表取数底稿'!O{bs_map[line]}")
            report_ws.cell(row_index, 4, f"='报表取数底稿'!P{bs_map[line]}")
        if offset < len(liability_lines):
            line = liability_lines[offset]
            report_ws.cell(row_index, 5, ws.cell(bs_map[line], 14).value)
            report_ws.cell(row_index, 6, line)
            report_ws.cell(row_index, 7, f"='报表取数底稿'!O{bs_map[line]}")
            report_ws.cell(row_index, 8, f"='报表取数底稿'!P{bs_map[line]}")
    bs_report_end = 4 + len(asset_lines)
    style_statement(report_ws, bs_report_end, 8, (3, 4, 7, 8))
    _set_widths(report_ws, [30, 8, 15, 15, 32, 8, 15, 15])

    report_ws = sheets[report_names["profit"]]
    standard_header(
        report_ws, "利润表", report_names["caption"], 4,
        ["项目", "行次", "本年累计金额", "本月金额"],
    )
    for line in range(1, 33):
        row_index = line + 4
        report_ws.cell(row_index, 1, ws.cell(profit_map[line], 14).value)
        report_ws.cell(row_index, 2, line)
        report_ws.cell(row_index, 3, f"='报表取数底稿'!O{profit_map[line]}")
        report_ws.cell(row_index, 4, f"='报表取数底稿'!P{profit_map[line]}")
    style_statement(report_ws, 36, 4, (3, 4))
    _set_widths(report_ws, [56, 9, 18, 18])

    report_ws = sheets[report_names["cash_flow"]]
    standard_header(
        report_ws, "现金流量表", report_names["caption"], 4,
        ["项目", "行次", "本年累计金额", "本月金额"],
    )
    for line in range(1, 23):
        row_index = line + 4
        report_ws.cell(row_index, 1, ws.cell(cash_map[line], 14).value)
        report_ws.cell(row_index, 2, line)
        report_ws.cell(row_index, 3, f"='报表取数底稿'!O{cash_map[line]}")
        report_ws.cell(row_index, 4, f"='报表取数底稿'!P{cash_map[line]}")
    style_statement(report_ws, 26, 4, (3, 4))
    _set_widths(report_ws, [66, 9, 18, 18])


def _create_enterprise_ledger_sheets(sheets: Dict[str, object], store,
                                     company_name: str, period: str,
                                     subject_row_map: Dict[Tuple[str, str], int],
                                     voucher_end: int):
    """Create formula-linked general, subsidiary, and cash/bank ledgers."""
    period_vouchers = sorted(
        [row for row in store.list_vouchers() if row.get("period") == period],
        key=lambda row: (
            str(row.get("subject_code", "")), str(row.get("subject", "")),
            str(row.get("date", "")), str(row.get("voucher_no", "")),
            int(row.get("line_no", 0) or 0),
        ),
    )

    # 总分类账直接引用科目余额表，确保同一工作簿内口径唯一。
    ws = sheets["总分类账"]
    _sheet_base(ws, "A6")
    headers = [
        "科目编码", "科目名称", "期初借方", "期初贷方", "本期借方",
        "本期贷方", "期末借方", "期末贷方", "本期分录数",
    ]
    _title(ws, "总分类账", f"企业：{company_name} | 期间：{period}", len(headers))
    _headers(ws, 5, headers)
    general_map: Dict[str, int] = {}
    for row_index, ((code, subject), balance_row) in enumerate(
        subject_row_map.items(), start=6
    ):
        general_map[code] = row_index
        ws.cell(row_index, 1, code)
        ws.cell(row_index, 2, subject)
        for target_col, source_col in zip(range(3, 9), "CDEFGH"):
            ws.cell(
                row_index, target_col,
                f"='科目余额表'!{source_col}{balance_row}",
            )
        ws.cell(
            row_index, 9,
            (
                f'=COUNTIFS(\'记账凭证\'!$B$6:$B${voucher_end},"{period}",'
                f"'记账凭证'!$F$6:$F${voucher_end},A{row_index})"
            ),
        )
    general_end = max(6, len(subject_row_map) + 5)
    _style_data(ws, 6, general_end, len(headers), money_cols=(3, 4, 5, 6, 7, 8))
    _set_widths(ws, [13, 30, 15, 15, 15, 15, 15, 15, 12])
    _table(ws, 5, len(subject_row_map) + 5, len(headers), "GeneralLedger")
    if subject_row_map:
        total_row = len(subject_row_map) + 7
        ws.cell(total_row, 2, "合计").font = Font(name="微软雅黑", bold=True)
        for col in range(3, 9):
            letter = get_column_letter(col)
            ws.cell(total_row, col, f"=SUM({letter}6:{letter}{general_end})")
            ws.cell(total_row, col).number_format = MONEY_FMT
            ws.cell(total_row, col).font = Font(name="微软雅黑", bold=True)

    # 明细分类账按科目连续滚算余额，M列保存带方向的净余额并隐藏。
    ws = sheets["明细分类账"]
    _sheet_base(ws, "A6")
    headers = [
        "日期", "凭证号", "分录号", "摘要", "科目编码", "科目名称",
        "借方", "贷方", "余额方向", "余额", "来源", "发票号码", "净余额",
    ]
    _title(ws, "明细分类账", f"企业：{company_name} | 期间：{period}", len(headers))
    _headers(ws, 5, headers)
    for row_index, item in enumerate(period_vouchers, start=6):
        code = str(item.get("subject_code", ""))
        values = [
            _safe_date(item.get("date")), item.get("voucher_no", ""),
            item.get("line_no", ""), item.get("description", ""), code,
            item.get("subject", ""), _money(item.get("debit")),
            _money(item.get("credit")), None, None,
            item.get("source", ""), item.get("invoice_no", ""), None,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        opening_formula = (
            f"SUMIFS('总分类账'!$C$6:$C${general_end},'总分类账'!$A$6:$A${general_end},E{row_index})-"
            f"SUMIFS('总分类账'!$D$6:$D${general_end},'总分类账'!$A$6:$A${general_end},E{row_index})"
        )
        if row_index == 6:
            signed_formula = f"={opening_formula}+G{row_index}-H{row_index}"
        else:
            signed_formula = (
                f"=IF(E{row_index}=E{row_index-1},M{row_index-1},"
                f"{opening_formula})+G{row_index}-H{row_index}"
            )
        ws.cell(row_index, 13, signed_formula)
        ws.cell(
            row_index, 9,
            f'=IF(M{row_index}>0,"借",IF(M{row_index}<0,"贷","平"))',
        )
        ws.cell(row_index, 10, f"=ABS(M{row_index})")
    detail_end = max(6, len(period_vouchers) + 5)
    _style_data(
        ws, 6, detail_end, len(headers),
        money_cols=(7, 8, 10, 13), date_cols=(1,),
    )
    _set_widths(ws, [13, 15, 8, 32, 13, 28, 14, 14, 10, 15, 13, 18, 3])
    ws.column_dimensions["M"].hidden = True
    _table(ws, 5, len(period_vouchers) + 5, len(headers), "SubsidiaryLedger")

    # 现金银行日记账保留现金流分类和对方科目，便于银行及报表勾稽。
    cash_lines = [
        row for row in period_vouchers
        if str(row.get("subject_code", "")) in ("1001", "1002", "1012")
    ]
    cash_lines.sort(key=lambda row: (
        str(row.get("subject_code", "")), str(row.get("date", "")),
        str(row.get("voucher_no", "")), int(row.get("line_no", 0) or 0),
    ))
    cash_entries = {
        str(row.get("line_id", "")): row
        for row in store.cash_flow_summary(period)["month_entries"]
    }
    voucher_groups: Dict[str, List[Dict]] = {}
    for row in store.list_vouchers():
        if row.get("period") == period:
            voucher_groups.setdefault(str(row.get("voucher_no", "")), []).append(row)

    ws = sheets["现金银行日记账"]
    _sheet_base(ws, "A6")
    headers = [
        "日期", "凭证号", "摘要", "科目编码", "科目名称", "收入", "支出",
        "余额方向", "余额", "现金流项目", "分类来源", "对方科目", "净余额",
    ]
    _title(ws, "现金银行日记账", f"企业：{company_name} | 期间：{period}", len(headers))
    _headers(ws, 5, headers)
    for row_index, item in enumerate(cash_lines, start=6):
        code = str(item.get("subject_code", ""))
        entry = cash_entries.get(str(item.get("id", "")), {})
        counterpart_subjects = sorted({
            str(row.get("subject", ""))
            for row in voucher_groups.get(str(item.get("voucher_no", "")), [])
            if str(row.get("id", "")) != str(item.get("id", ""))
            and str(row.get("subject_code", "")) not in ("1001", "1002", "1012")
        })
        values = [
            _safe_date(item.get("date")), item.get("voucher_no", ""),
            item.get("description", ""), code, item.get("subject", ""),
            _money(item.get("debit")), _money(item.get("credit")), None, None,
            entry.get("category_label", "内部划转" if not entry else ""),
            entry.get("classification_source", "自动" if entry else "不适用"),
            "、".join(counterpart_subjects), None,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        opening_formula = (
            f"SUMIFS('总分类账'!$C$6:$C${general_end},'总分类账'!$A$6:$A${general_end},D{row_index})-"
            f"SUMIFS('总分类账'!$D$6:$D${general_end},'总分类账'!$A$6:$A${general_end},D{row_index})"
        )
        if row_index == 6:
            signed_formula = f"={opening_formula}+F{row_index}-G{row_index}"
        else:
            signed_formula = (
                f"=IF(D{row_index}=D{row_index-1},M{row_index-1},"
                f"{opening_formula})+F{row_index}-G{row_index}"
            )
        ws.cell(row_index, 13, signed_formula)
        ws.cell(
            row_index, 8,
            f'=IF(M{row_index}>0,"借",IF(M{row_index}<0,"贷","平"))',
        )
        ws.cell(row_index, 9, f"=ABS(M{row_index})")
    cash_end = max(6, len(cash_lines) + 5)
    _style_data(
        ws, 6, cash_end, len(headers), money_cols=(6, 7, 9, 13), date_cols=(1,),
    )
    _set_widths(ws, [13, 15, 30, 13, 25, 14, 14, 10, 15, 38, 13, 30, 3])
    ws.column_dimensions["M"].hidden = True
    _table(ws, 5, len(cash_lines) + 5, len(headers), "CashBankJournal")


def export_finance_workbook(store, output_path: Path,
                            period: Optional[str] = None) -> Path:
    """Export an auditable accounting and tax filing-support workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    settings = store.get_settings()
    company = settings["company"]
    tax = settings["tax"]
    all_vouchers = store.list_vouchers()
    report_period = period or datetime.now().strftime("%Y-%m")
    tax_view = store.tax_summary(report_period)
    invoice_months = set(tax_view.get("period", {}).get("months", [report_period]))
    invoices = [
        row for row in store.list_invoices()
        if str(row.get("invoice_date", ""))[:7] in invoice_months
    ]
    all_opening_balances = store.list_opening_balances()
    if period:
        opening_balances = store.opening_balances_for_period(period)
    else:
        opening_periods = sorted({
            row.get("period", "") for row in all_opening_balances if row.get("period")
        })
        opening_balances = (
            store.list_opening_balances(opening_periods[-1]) if opening_periods else []
        )
    opening_source_period = (
        opening_balances[0].get("period", "") if opening_balances else "未设置"
    )
    if period:
        voucher_start_period = (
            opening_source_period if opening_source_period != "未设置" else period
        )
        vouchers = [
            row for row in all_vouchers
            if voucher_start_period <= str(row.get("period", "")) <= period
        ]
        voucher_scope_label = f"{voucher_start_period} 至 {period}"
    else:
        vouchers = all_vouchers
        voucher_scope_label = "全部期间"
    vouchers = sorted(
        vouchers,
        key=lambda row: (
            str(row.get("date", "")), str(row.get("voucher_no", "")),
            int(row.get("line_no", 0) or 0),
        ),
    )
    bank_transactions = store.list_bank_transactions(period)
    payroll_records = store.list_payroll(period)
    fixed_assets = [
        row for row in store.list_fixed_assets()
        if not period or str(row.get("purchase_date", ""))[:7] <= period
    ]
    depreciation_rows = [
        row for row in store.depreciation_schedule(report_period)
        if not period or str(row.get("purchase_date", ""))[:7] <= period
    ]
    issues = store.validate(period)
    checklist = store.month_end_checklist(report_period)
    period_label = period or "全部期间"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    company_name = company.get("name") or "未设置企业/单位名称"

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    report_names = _enterprise_statement_names(store) if store.is_enterprise else {
        "balance": "科目余额表",
        "profit": "利润表辅助",
        "cash_flow": "银行对账",
        "caption": "内部管理报表",
    }
    sheet_names = ["月度总览", "使用说明", "科目目录", "期初余额", "记账凭证"]
    if store.profile_key == "enterprise":
        sheet_names.extend(["总分类账", "明细分类账", "现金银行日记账"])
    sheet_names.extend([
        "发票台账", "银行对账", "工资社保", "固定资产", "折旧明细", "科目余额表",
    ])
    if store.profile_key == "enterprise":
        sheet_names.extend([
            "报表取数底稿", report_names["balance"],
            report_names["profit"], report_names["cash_flow"],
        ])
    sheet_names.extend([
        "利润表辅助", "资产负债辅助", "税务期间", "纳税调整", "增值税测算",
        "个税测算", "印花税准备", "所得税预缴", "所得税测算", "年度汇算准备",
        "申报校验", "政策依据",
    ])
    sheets = {name: wb.create_sheet(name) for name in sheet_names}
    if store.profile_key == "enterprise":
        _create_enterprise_standard_reports(
            sheets, store, company_name, report_period,
        )

    # 使用说明与企业资料
    ws = sheets["使用说明"]
    _sheet_base(ws)
    _title(ws, "小企业月度报税准备工作簿", f"账套：{store.profile_label} | 期间：{period_label}", 6)
    note_rows = [
        ("工作簿用途", "用于整理期初余额、凭证、发票、银行对账、工资、固定资产和税费测算，便于自行复核与申报准备。"),
        ("重要声明", LEGAL_NOTICE_SUMMARY),
        ("政策参数快照", policy_snapshot_text(tax)),
        ("参数维护责任", "税率、免税阈值、优惠资格和政策截止日期由用户在系统设置中维护；软件不自动保证政策持续有效。"),
        ("数据范围", f"仅包含本软件中已保存的数据；科目余额采用 {opening_source_period} 的期初余额，并滚算至所选期间。"),
        ("更新方式", "在软件中修正凭证、发票或设置后重新导出，不建议直接覆盖明细工作表。"),
        ("生成时间", generated_at),
    ]
    if store.profile_key == "enterprise":
        note_rows.insert(
            2,
            (
                "标准财务报表",
                f"{report_names['caption']}报表引用“报表取数底稿”；自定义科目、重分类和披露事项仍需人工复核。",
            ),
        )
        note_rows.insert(
            3,
            (
                "账簿输出",
                "总分类账、明细分类账和现金银行日记账按所选期间生成，余额与科目余额表及凭证序时账勾稽。",
            ),
        )
    row = 4
    for label, value in note_rows:
        ws.cell(row, 1, label).font = Font(name="微软雅黑", bold=True, color=DARK)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row, 2, value).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 2).font = Font(name="微软雅黑", size=10, color=TEXT)
        row += 1
    row += 1
    ws.cell(row, 1, "企业/单位资料").fill = PatternFill("solid", fgColor=TEAL)
    ws.cell(row, 1).font = Font(name="微软雅黑", bold=True, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    profile_rows = [
        ("名称", company_name), ("统一社会信用代码", company.get("credit_code", "")),
        ("纳税人类型", company.get("taxpayer_type", "")), ("行业", company.get("industry", "")),
        ("法定代表人/负责人", company.get("legal_representative", "")),
        ("财务联系人", company.get("finance_contact", "")),
        ("注册地址", company.get("registered_address", "")),
        ("开户银行", company.get("bank_name", "")), ("银行账号", company.get("bank_account", "")),
    ]
    for label, value in profile_rows:
        ws.cell(row, 1, label).font = Font(name="微软雅黑", bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        display_value = value
        if label == "银行账号" and str(value).isdigit():
            account_text = str(value)
            display_value = " ".join(
                account_text[index:index + 4]
                for index in range(0, len(account_text), 4)
            )
        ws.cell(row, 2, display_value)
        if label in ("统一社会信用代码", "银行账号"):
            ws.cell(row, 2).number_format = "@"
            ws.cell(row, 2).quotePrefix = True
            ws.cell(row, 2).alignment = Alignment(horizontal="left")
        row += 1
    _set_widths(ws, [20, 20, 18, 18, 18, 18])

    # 独立66科目目录：词库只做语义映射，本表是唯一科目主档。
    ws = sheets["科目目录"]
    _sheet_base(ws, "A6")
    _title(
        ws, "《小企业会计准则》66科目目录",
        f"启用模板：{settings.get('accounting', {}).get('account_template', '服务业')} | 主档与语义词库分离",
        7,
    )
    account_headers = ["顺序", "科目编码", "科目名称", "类别", "余额方向", "当前启用", "核算说明"]
    _headers(ws, 5, account_headers)
    enabled_codes = set(store.enabled_account_codes())
    for row_index, account in enumerate(store.all_accounts(), start=6):
        values = [
            account.get("order", ""), str(account.get("code", "")), account.get("name", ""),
            account.get("class", ""), account.get("normal_balance", ""),
            "是" if str(account.get("code", "")) in enabled_codes else "否",
            account.get("usage", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    _style_data(ws, 6, 71, 7)
    _set_widths(ws, [9, 14, 28, 14, 12, 12, 54])
    _table(ws, 5, 71, 7, "OfficialAccountCatalog")

    # 期初余额
    ws = sheets["期初余额"]
    _sheet_base(ws, "A6")
    opening_headers = ["期初期间", "科目编码", "科目名称", "期初借方", "期初贷方", "备注"]
    _title(
        ws, "期初余额", f"采用期间：{opening_source_period} | 借贷双方应保持平衡",
        len(opening_headers),
    )
    _headers(ws, 5, opening_headers)
    for row_index, item in enumerate(opening_balances, start=6):
        code = str(item.get("subject_code", ""))
        subject = str(item.get("subject", ""))
        name = subject[len(code):].strip() if code and subject.startswith(code) else subject
        values = [
            item.get("period", ""), code, name,
            _money(item.get("debit_balance")), _money(item.get("credit_balance")),
            item.get("note", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    opening_end = max(6, len(opening_balances) + 5)
    _style_data(ws, 6, opening_end, len(opening_headers), money_cols=(4, 5))
    _set_widths(ws, [12, 14, 30, 16, 16, 32])
    _table(ws, 5, len(opening_balances) + 5, len(opening_headers), "OpeningBalances")
    if opening_balances:
        total_row = len(opening_balances) + 7
        ws.cell(total_row, 3, "合计").font = Font(name="微软雅黑", bold=True)
        for col in (4, 5):
            letter = get_column_letter(col)
            ws.cell(total_row, col, f"=SUM({letter}6:{letter}{len(opening_balances)+5})")
            ws.cell(total_row, col).number_format = MONEY_FMT
            ws.cell(total_row, col).font = Font(name="微软雅黑", bold=True)
        ws.cell(total_row, 6, f'=IF(ABS(D{total_row}-E{total_row})<0.01,"平衡","不平衡")')

    # 记账凭证
    ws = sheets["记账凭证"]
    _sheet_base(ws, "A6")
    voucher_headers = [
        "凭证日期", "期间", "凭证号", "分录号", "摘要", "科目编码", "科目名称",
        "借方金额", "贷方金额", "状态", "来源", "发票号码", "往来单位", "税额", "附件路径",
    ]
    _title(
        ws, "记账凭证序时账",
        f"企业/单位：{company_name} | 数据范围：{voucher_scope_label} | 报告期间：{period_label}",
        len(voucher_headers),
    )
    _headers(ws, 5, voucher_headers)
    for row_index, item in enumerate(vouchers, start=6):
        values = [
            _safe_date(item.get("date")), item.get("period", ""), item.get("voucher_no", ""),
            item.get("line_no", ""), item.get("description", ""), item.get("subject_code", ""),
            item.get("subject", ""), _money(item.get("debit")), _money(item.get("credit")),
            item.get("status", ""), item.get("source", ""), item.get("invoice_no", ""),
            item.get("counterparty", ""), _money(item.get("tax_amount")), item.get("attachment", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    voucher_end = max(6, len(vouchers) + 5)
    _style_data(ws, 6, voucher_end, len(voucher_headers), money_cols=(8, 9, 14), date_cols=(1,))
    _set_widths(ws, [13, 10, 15, 8, 30, 12, 24, 14, 14, 10, 10, 18, 22, 13, 36])
    _table(ws, 5, len(vouchers) + 5, len(voucher_headers), "VoucherJournal")
    if vouchers:
        total_row = len(vouchers) + 7
        ws.cell(total_row, 7, "本表合计").font = Font(name="微软雅黑", bold=True)
        ws.cell(total_row, 8, f"=SUM(H6:H{len(vouchers)+5})")
        ws.cell(total_row, 9, f"=SUM(I6:I{len(vouchers)+5})")
        ws.cell(total_row, 10, f"=IF(ABS(H{total_row}-I{total_row})<0.01,\"平衡\",\"不平衡\")")
        for col in (8, 9):
            ws.cell(total_row, col).number_format = MONEY_FMT
            ws.cell(total_row, col).font = Font(name="微软雅黑", bold=True)

    # 发票台账
    ws = sheets["发票台账"]
    _sheet_base(ws, "A6")
    invoice_headers = [
        "开票日期", "期间", "发票代码", "发票号码", "方向", "单据类型", "发票形式",
        "原蓝字发票号", "销售方", "购买方", "不含税金额", "税额", "价税合计",
        "可抵扣", "价税口径", "税务处理", "状态", "来源文件",
    ]
    _title(ws, "发票台账", f"企业/单位：{company_name} | 期间：{period_label}", len(invoice_headers))
    _headers(ws, 5, invoice_headers)
    for row_index, item in enumerate(invoices, start=6):
        values = [
            _safe_date(item.get("invoice_date")), str(item.get("invoice_date", ""))[:7],
            item.get("invoice_code", ""), item.get("invoice_no", ""),
            item.get("invoice_type", "进项"), item.get("document_type", "正常发票"),
            item.get("invoice_form", "普通发票"), item.get("original_invoice_no", ""),
            item.get("seller", ""), item.get("buyer", ""),
            _money(item.get("amount")), _money(item.get("tax_amount")),
            _money(item.get("total_amount")), "是" if item.get("deductible") else "否",
            item.get("price_tax_mode", "含税"), item.get("tax_treatment", "自动判断"),
            item.get("status", ""), item.get("file_path", item.get("attachment", "")),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    invoice_end = max(6, len(invoices) + 5)
    _style_data(ws, 6, invoice_end, len(invoice_headers), money_cols=(11, 12, 13), date_cols=(1,))
    _set_widths(ws, [13, 10, 18, 18, 9, 12, 14, 18, 24, 24, 14, 13, 14, 10, 10, 12, 10, 34])
    _table(ws, 5, len(invoices) + 5, len(invoice_headers), "InvoiceLedger")

    # 银行对账
    ws = sheets["银行对账"]
    _sheet_base(ws, "A6")
    bank_headers = [
        "交易日期", "期间", "方向", "金额", "摘要", "对方户名", "本方账号",
        "账户余额", "匹配凭证", "匹配状态",
    ]
    cash_entry_map = {}
    if store.profile_key == "enterprise":
        bank_headers.extend(["现金流项目", "分类来源"])
        cash_entry_map = {
            str(row.get("line_id", "")): row
            for row in store.cash_flow_summary(report_period)["month_entries"]
        }
    bank_headers.append("来源文件")
    _title(ws, "银行流水对账", f"期间：{period_label} | 未匹配流水需在申报前复核", len(bank_headers))
    _headers(ws, 5, bank_headers)
    for row_index, item in enumerate(bank_transactions, start=6):
        values = [
            _safe_date(item.get("date")), str(item.get("date", ""))[:7],
            item.get("direction", ""), _money(item.get("amount")), item.get("summary", ""),
            item.get("counterparty", ""), item.get("account", ""), _money(item.get("balance")),
            item.get("voucher_no", ""), item.get("status", "未匹配"),
        ]
        if store.profile_key == "enterprise":
            entry = cash_entry_map.get(str(item.get("voucher_line_id", "")))
            category_key = str(item.get("cash_flow_category", ""))
            manual_label = next(
                (
                    label for label in store.cash_flow_category_options(
                        item.get("direction", "")
                    )
                    if store.cash_flow_category_key(label) == category_key
                ),
                "",
            )
            category_label = manual_label or (
                entry.get("category_label", "") if entry else ""
            )
            if manual_label:
                classification_source = (
                    "手工" if item.get("voucher_line_id") else "手工（待匹配）"
                )
            elif entry:
                classification_source = (
                    "自动（待复核）" if entry.get("needs_review") else "自动"
                )
            else:
                classification_source = "待匹配"
            values.extend([category_label or "未分类", classification_source])
        values.append(item.get("source_file", ""))
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    bank_end = max(6, len(bank_transactions) + 5)
    _style_data(ws, 6, bank_end, len(bank_headers), money_cols=(4, 8), date_cols=(1,))
    if store.profile_key == "enterprise":
        _set_widths(ws, [13, 10, 9, 14, 32, 24, 22, 15, 15, 12, 38, 16, 38])
    else:
        _set_widths(ws, [13, 10, 9, 14, 32, 24, 22, 15, 15, 12, 38])
    _table(ws, 5, len(bank_transactions) + 5, len(bank_headers), "BankReconciliation")

    # 工资社保
    ws = sheets["工资社保"]
    _sheet_base(ws, "A6")
    payroll_headers = [
        "工资期间", "员工姓名", "应发工资", "个人社保", "个人公积金", "个人所得税",
        "实发工资", "单位社保", "单位公积金", "企业人工成本", "计提状态", "凭证号",
        "计提日期", "备注",
    ]
    _title(ws, "工资社保台账", f"期间：{period_label} | 实发工资和人工成本由公式复核", len(payroll_headers))
    _headers(ws, 5, payroll_headers)
    for row_index, item in enumerate(payroll_records, start=6):
        values = [
            item.get("period", ""), item.get("employee_name", ""),
            _money(item.get("gross_salary")), _money(item.get("social_personal")),
            _money(item.get("housing_personal")), _money(item.get("income_tax")), None,
            _money(item.get("social_company")), _money(item.get("housing_company")), None,
            item.get("status", ""), item.get("voucher_no", ""),
            _safe_date(item.get("pay_date")), item.get("note", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 7, f"=C{row_index}-D{row_index}-E{row_index}-F{row_index}")
        ws.cell(row_index, 10, f"=C{row_index}+H{row_index}+I{row_index}")
    payroll_end = max(6, len(payroll_records) + 5)
    _style_data(
        ws, 6, payroll_end, len(payroll_headers),
        money_cols=(3, 4, 5, 6, 7, 8, 9, 10), date_cols=(13,),
    )
    _set_widths(ws, [12, 14, 14, 13, 14, 13, 14, 13, 14, 16, 12, 15, 13, 30])
    _table(ws, 5, len(payroll_records) + 5, len(payroll_headers), "PayrollRegister")

    # 固定资产卡片
    ws = sheets["固定资产"]
    _sheet_base(ws, "A6")
    asset_headers = [
        "资产名称", "类别", "购置日期", "原值", "净残值率", "使用月数", "折旧起始期间",
        "月折旧额", "固定资产科目", "累计折旧科目", "费用科目", "状态", "备注",
    ]
    _title(ws, "固定资产卡片", f"截至：{report_period} | 采用直线法折旧", len(asset_headers))
    _headers(ws, 5, asset_headers)
    for row_index, item in enumerate(fixed_assets, start=6):
        values = [
            item.get("asset_name", ""), item.get("category", ""),
            _safe_date(item.get("purchase_date")), _money(item.get("original_cost")),
            float(item.get("residual_rate", 0)), int(item.get("useful_months", 0)),
            item.get("depreciation_start_period", ""), None,
            item.get("asset_subject", ""), item.get("depreciation_subject", ""),
            item.get("expense_subject", ""), item.get("status", ""), item.get("note", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 8, f"=ROUND(D{row_index}*(1-E{row_index})/F{row_index},2)")
        ws.cell(row_index, 5).number_format = PERCENT_FMT
    asset_end = max(6, len(fixed_assets) + 5)
    _style_data(ws, 6, asset_end, len(asset_headers), money_cols=(4, 8), date_cols=(3,))
    _set_widths(ws, [22, 14, 13, 15, 12, 11, 14, 14, 24, 24, 24, 11, 30])
    _table(ws, 5, len(fixed_assets) + 5, len(asset_headers), "FixedAssetRegister")

    # 折旧明细
    ws = sheets["折旧明细"]
    _sheet_base(ws, "A6")
    depreciation_headers = [
        "期间", "资产名称", "类别", "原值", "折旧月序", "月折旧额", "本期折旧",
        "累计折旧", "账面净值", "是否已计提", "凭证号", "费用科目", "累计折旧科目",
    ]
    _title(ws, "固定资产折旧明细", f"期间：{report_period} | 末月自动按剩余可折旧额补差", len(depreciation_headers))
    _headers(ws, 5, depreciation_headers)
    for row_index, item in enumerate(depreciation_rows, start=6):
        values = [
            report_period, item.get("asset_name", ""), item.get("category", ""),
            _money(item.get("original_cost")), item.get("month_number", 0),
            _money(item.get("monthly_depreciation")), _money(item.get("depreciation_amount")),
            _money(item.get("accumulated_depreciation")), _money(item.get("net_book_value")),
            "是" if item.get("posted") else "否", item.get("voucher_no", ""),
            item.get("expense_subject", ""), item.get("depreciation_subject", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    depreciation_end = max(6, len(depreciation_rows) + 5)
    _style_data(
        ws, 6, depreciation_end, len(depreciation_headers),
        money_cols=(4, 6, 7, 8, 9),
    )
    _set_widths(ws, [11, 22, 14, 15, 11, 14, 14, 15, 15, 12, 15, 24, 24])
    _table(ws, 5, len(depreciation_rows) + 5, len(depreciation_headers), "DepreciationDetail")

    # 科目余额表，发生额以凭证表为数据源，便于审计追踪。
    subjects: Dict[Tuple[str, str], None] = {}
    for item in vouchers:
        subjects[(str(item.get("subject_code", "")), str(item.get("subject", "")))] = None
    for item in opening_balances:
        subjects[(str(item.get("subject_code", "")), str(item.get("subject", "")))] = None
    subject_rows = sorted(subjects, key=lambda pair: (pair[0], pair[1]))
    ws = sheets["科目余额表"]
    _sheet_base(ws, "A6")
    balance_headers = [
        "科目编码", "科目名称", "期初借方", "期初贷方",
        "本期借方", "本期贷方", "期末借方", "期末贷方",
    ]
    _title(ws, "科目余额表", f"期间：{period_label} | 期初来源：{opening_source_period}", len(balance_headers))
    _headers(ws, 5, balance_headers)
    subject_row_map: Dict[Tuple[str, str], int] = {}
    for row_index, (code, subject) in enumerate(subject_rows, start=6):
        subject_row_map[(code, subject)] = row_index
        ws.cell(row_index, 1, code)
        ws.cell(row_index, 2, subject)
        subject_name = subject[len(code):].strip() if code and subject.startswith(code) else subject
        subject_name = subject_name.replace('"', '""')
        raw_debit = f'SUMIFS(\'期初余额\'!$D$6:$D${opening_end},\'期初余额\'!$B$6:$B${opening_end},$A{row_index},\'期初余额\'!$C$6:$C${opening_end},"{subject_name}")'
        raw_credit = f'SUMIFS(\'期初余额\'!$E$6:$E${opening_end},\'期初余额\'!$B$6:$B${opening_end},$A{row_index},\'期初余额\'!$C$6:$C${opening_end},"{subject_name}")'
        voucher_debit = f'SUMIFS(\'记账凭证\'!$H$6:$H${voucher_end},\'记账凭证\'!$F$6:$F${voucher_end},$A{row_index},\'记账凭证\'!$G$6:$G${voucher_end},$B{row_index}'
        voucher_credit = f'SUMIFS(\'记账凭证\'!$I$6:$I${voucher_end},\'记账凭证\'!$F$6:$F${voucher_end},$A{row_index},\'记账凭证\'!$G$6:$G${voucher_end},$B{row_index}'
        if period:
            prior_debit = voucher_debit + f',\'记账凭证\'!$B$6:$B${voucher_end},"<{period}")'
            prior_credit = voucher_credit + f',\'记账凭证\'!$B$6:$B${voucher_end},"<{period}")'
            current_debit = voucher_debit + f',\'记账凭证\'!$B$6:$B${voucher_end},"{period}")'
            current_credit = voucher_credit + f',\'记账凭证\'!$B$6:$B${voucher_end},"{period}")'
        else:
            prior_debit = prior_credit = "0"
            current_debit = voucher_debit + ")"
            current_credit = voucher_credit + ")"
        ws.cell(row_index, 3, f"=MAX(({raw_debit}+{prior_debit})-({raw_credit}+{prior_credit}),0)")
        ws.cell(row_index, 4, f"=MAX(({raw_credit}+{prior_credit})-({raw_debit}+{prior_debit}),0)")
        ws.cell(row_index, 5, f"={current_debit}")
        ws.cell(row_index, 6, f"={current_credit}")
        ws.cell(row_index, 7, f"=MAX(C{row_index}+E{row_index}-D{row_index}-F{row_index},0)")
        ws.cell(row_index, 8, f"=MAX(D{row_index}+F{row_index}-C{row_index}-E{row_index},0)")
    balance_end = max(6, len(subject_rows) + 5)
    _style_data(ws, 6, balance_end, len(balance_headers), money_cols=(3, 4, 5, 6, 7, 8))
    _set_widths(ws, [14, 30, 15, 15, 15, 15, 16, 16])
    _table(ws, 5, len(subject_rows) + 5, len(balance_headers), "SubjectBalance")
    if subject_rows:
        total_row = len(subject_rows) + 7
        ws.cell(total_row, 2, "合计").font = Font(name="微软雅黑", bold=True)
        for col in range(3, 9):
            letter = get_column_letter(col)
            ws.cell(total_row, col, f"=SUM({letter}6:{letter}{len(subject_rows)+5})")
            ws.cell(total_row, col).number_format = MONEY_FMT
            ws.cell(total_row, col).font = Font(name="微软雅黑", bold=True)

    if store.profile_key == "enterprise":
        _create_enterprise_ledger_sheets(
            sheets, store, company_name, report_period,
            subject_row_map, voucher_end,
        )

    revenue_cells, cost_cells, expense_cells = [], [], []
    asset_debit_cells, asset_credit_cells = [], []
    liability_credit_cells, liability_debit_cells = [], []
    equity_credit_cells, equity_debit_cells = [], []
    operating_revenue_codes = set(store.standard_code_set["operating_revenue"])
    if store.is_small_enterprise_standard:
        cost_codes = {"5401", "5402"}
        asset_cost_codes = {"4001", "4101", "4301", "4401", "4403"}
        equity_codes = {"3001", "3002", "3101", "3103", "3104"}
    else:
        cost_codes = {"6401", "6402"}
        asset_cost_codes = {"5001", "5101", "5201", "5301"}
        equity_codes = {"4001", "4002", "4101", "4103", "4104"}
    expense_codes = set(store.standard_code_set["profit_expense"]) - cost_codes
    for (code, subject), row_index in subject_row_map.items():
        if code in operating_revenue_codes:
            revenue_cells.append(f"F{row_index}")
        elif code in cost_codes:
            cost_cells.append(f"E{row_index}")
        elif code in expense_codes:
            expense_cells.append(f"E{row_index}")
        if code.startswith("1") or code in asset_cost_codes:
            asset_debit_cells.append(f"G{row_index}")
            asset_credit_cells.append(f"H{row_index}")
        elif code.startswith("2"):
            liability_credit_cells.append(f"H{row_index}")
            liability_debit_cells.append(f"G{row_index}")
        elif code in equity_codes:
            equity_credit_cells.append(f"H{row_index}")
            equity_debit_cells.append(f"G{row_index}")

    # 利润表辅助
    ws = sheets["利润表辅助"]
    _sheet_base(ws)
    _title(ws, "利润表申报辅助", f"期间：{period_label} | 金额单位：元", 4)
    _headers(ws, 5, ["项目", "本期金额", "数据来源", "复核说明"])
    if store.profile_key == "enterprise":
        profit_sheet = report_names["profit"]
        report_source = report_names["caption"]
        profit_rows = [
            ("营业收入", f"='{profit_sheet}'!D5", f"{report_source}行1", "核对销项发票及未开票收入"),
            ("营业成本", f"='{profit_sheet}'!D6", f"{report_source}行2", "核对成本结转完整性"),
            (
                "税金及期间费用",
                f"=SUM('{profit_sheet}'!D7,'{profit_sheet}'!D15,'{profit_sheet}'!D18,'{profit_sheet}'!D22)",
                f"{report_source}行3、11、14、18",
                "核对税前扣除凭证及限额",
            ),
            ("利润总额（辅助）", f"='{profit_sheet}'!D34", f"{report_source}行30", "含投资收益及营业外收支"),
        ]
    else:
        profit_rows = [
            ("营业收入", _formula_sum("科目余额表", revenue_cells), "收入类科目本期贷方发生额", "核对收入完整性"),
            ("营业成本", _formula_sum("科目余额表", cost_cells), "成本类科目本期借方发生额", "核对成本结转完整性"),
            ("期间费用", _formula_sum("科目余额表", expense_cells), "费用/支出类科目本期借方发生额", "核对凭证及限额"),
            ("利润总额（辅助）", "=B6-B7-B8", "本表公式", "不含未录入调整事项"),
        ]
    for row_index, values in enumerate(profit_rows, start=6):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 2).number_format = MONEY_FMT
    _style_data(ws, 6, 9, 4, money_cols=(2,))
    for col in range(1, 5):
        ws.cell(9, col).font = Font(name="微软雅黑", bold=True, color=DARK)
        ws.cell(9, col).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    _set_widths(ws, [26, 18, 34, 38])

    # 资产负债辅助
    ws = sheets["资产负债辅助"]
    _sheet_base(ws)
    _title(ws, "资产负债表支持数据", f"截至：{period_label} | 期初来源：{opening_source_period}", 4)
    _headers(ws, 5, ["项目", "期末辅助余额", "数据来源", "限制说明"])
    if store.profile_key == "enterprise":
        balance_sheet = report_names["balance"]
        profit_sheet = report_names["profit"]
        report_source = report_names["caption"]
        balance_support = [
            ("资产合计", f"='{balance_sheet}'!C34", f"{report_source}行30", "已包含借方余额重分类和备抵项目"),
            ("负债合计", f"='{balance_sheet}'!G21", f"{report_source}行47", "仅列负债科目贷方余额"),
            (
                "所有者权益科目合计",
                f"='{balance_sheet}'!G26-'{profit_sheet}'!C36",
                f"{report_source}行52减利润表行32",
                "不含尚未结转的本年利润",
            ),
            ("本年未结转净利润", f"='{profit_sheet}'!C36", f"{report_source}利润表行32", "期末结账时应按会计制度完成损益结转"),
            ("权益及未结转利润合计", "=B8+B9", "本表公式", f"应等于{report_source}资产负债表行52"),
            ("平衡差额", "=B6-B7-B10", "本表公式", "应为0；非0时不得作为申报定稿"),
        ]
    else:
        balance_support = [
            ("资产合计", _formula_net("科目余额表", asset_debit_cells, asset_credit_cells), "资产类借方余额减贷方余额", "累计折旧等备抵科目已作为抵减项"),
            ("负债合计", _formula_net("科目余额表", liability_credit_cells, liability_debit_cells), "负债类贷方余额减借方余额", "核对应付、税费和借款明细"),
            ("所有者权益科目合计", _formula_net("科目余额表", equity_credit_cells, equity_debit_cells), "权益类贷方余额减借方余额", "不含尚未结转的本期利润"),
            ("本期未结转利润", "='利润表辅助'!B9", "利润表辅助", "期末结账时应按会计制度完成损益结转"),
            ("权益及未结转利润合计", "=B8+B9", "本表公式", "用于申报准备阶段的勾稽复核"),
            ("平衡差额", "=B6-B7-B10", "本表公式", "为0不代表完整财务报表已编制完成"),
        ]
    for row_index, values in enumerate(balance_support, start=6):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 2).number_format = MONEY_FMT
    _style_data(ws, 6, 11, 4, money_cols=(2,))
    ws.conditional_formatting.add(
        "B11", CellIsRule(operator="notBetween", formula=["-0.01", "0.01"], fill=PatternFill("solid", fgColor=LIGHT_RED))
    )
    _set_widths(ws, [26, 18, 34, 42])

    # 税务期间与资格检查
    ws = sheets["税务期间"]
    _sheet_base(ws)
    vat_period_info = tax_view.get("period", {})
    cit_view = tax_view.get("cit", {})
    cit_period_info = cit_view.get("period", {})
    _title(
        ws, "税务期间与支持范围",
        "本工具仅支持“小规模纳税人 + 小型微利企业”；资格失败时停止自动给出税额",
        5,
    )
    _headers(ws, 5, ["检查项目", "实际值", "要求/上限", "结果", "说明"])
    period_rows = [
        ("增值税期间", vat_period_info.get("key", ""), "按月或按季", "通过", f"{vat_period_info.get('start_month', '')} 至 {vat_period_info.get('end_month', '')}"),
        ("企业所得税预缴期间", cit_period_info.get("key", ""), "按月或按季", "通过", f"{cit_period_info.get('start_month', '')} 至 {cit_period_info.get('end_month', '')}"),
    ]
    for check in tax_view.get("scope", {}).get("checks", []):
        period_rows.append((
            check.get("item", ""), check.get("actual", ""), check.get("limit", ""),
            "通过" if check.get("passed") else "不通过", tax_view.get("scope", {}).get("message", ""),
        ))
    for check in cit_view.get("eligibility", {}).get("checks", []):
        period_rows.append((
            check.get("item", ""), check.get("actual", ""), check.get("limit", ""),
            "通过" if check.get("passed") else "不通过", "小型微利企业资格复核",
        ))
    for row_index, values in enumerate(period_rows, start=6):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 4).fill = PatternFill(
            "solid", fgColor=LIGHT_GREEN if values[3] == "通过" else LIGHT_RED
        )
    _style_data(ws, 6, len(period_rows) + 5, 5)
    _set_widths(ws, [28, 24, 24, 12, 54])

    # 纳税调整台账
    annual_view = store.annual_cit_summary(report_period[:4])
    all_adjustments = list(annual_view.get("adjustments", []))
    ws = sheets["纳税调整"]
    _sheet_base(ws, "A6")
    _title(ws, "纳税调整台账", f"年度：{report_period[:4]} | 在软件财税工作台维护", 6)
    adjustment_headers = ["期间", "税种", "调整类别", "方向", "金额", "依据/备注"]
    _headers(ws, 5, adjustment_headers)
    for row_index, item in enumerate(all_adjustments, start=6):
        values = [
            item.get("period", ""), item.get("tax_type", ""), item.get("category", ""),
            item.get("direction", ""), _money(item.get("amount")),
            item.get("basis") or item.get("note", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
    adjustment_end = max(6, len(all_adjustments) + 5)
    _style_data(ws, 6, adjustment_end, 6, money_cols=(5,))
    _set_widths(ws, [12, 16, 24, 20, 16, 52])
    _table(ws, 5, len(all_adjustments) + 5, 6, "TaxAdjustments")

    # 增值税测算
    ws = sheets["增值税测算"]
    _sheet_base(ws)
    _title(
        ws, "小规模纳税人增值税及附加税费测算",
        f"税务期间：{vat_period_info.get('key', '')}（{vat_period_info.get('start_month', '')} 至 {vat_period_info.get('end_month', '')}）",
        4,
    )
    _headers(ws, 5, ["项目", "金额/税率", "计算依据", "复核事项"])
    invoice_count_formula = f'COUNTIFS(\'发票台账\'!$E$6:$E${invoice_end},"销项")'
    sales_formula = (
        f'=IF({invoice_count_formula}=0,\'利润表辅助\'!B6,'
        f'SUMIFS(\'发票台账\'!$K$6:$K${invoice_end},\'发票台账\'!$E$6:$E${invoice_end},"销项",'
        f'\'发票台账\'!$P$6:$P${invoice_end},"<>不征税"))'
    )
    non_exempt_formula = (
        f'=SUMIFS(\'发票台账\'!$K$6:$K${invoice_end},\'发票台账\'!$G$6:$G${invoice_end},"增值税专用发票")+'
        f'SUMIFS(\'发票台账\'!$K$6:$K${invoice_end},\'发票台账\'!$P$6:$P${invoice_end},"不得免税")-'
        f'SUMIFS(\'发票台账\'!$K$6:$K${invoice_end},\'发票台账\'!$G$6:$G${invoice_end},"增值税专用发票",'
        f'\'发票台账\'!$P$6:$P${invoice_end},"不得免税")'
    )
    vat_rows = [
        ("价税分离后销售额", sales_formula, "销项发票、红字发票、未开票收入", "无销项台账时回退到凭证收入"),
        ("不得免税销售额", non_exempt_formula, "专票或人工标记不得免税", "阈值内仍按简易计税率计税"),
        ("明确免税项目销售额", f'=SUMIFS(\'发票台账\'!$K$6:$K${invoice_end},\'发票台账\'!$P$6:$P${invoice_end},"免税项目")', "发票台账税务处理", "不含阈值免税"),
        ("免税销售额阈值", float(tax_view.get("vat", {}).get("threshold", 0)), "系统设置", "月10万元/季30万元为当前预设，可修改"),
        ("达到阈值免税条件", '=IF(MAX(B6-B8,0)<=B9,"是","否")', "本表公式", "专票和不得免税销售额除外"),
        ("应纳增值税（测算）", '=IF(B10="是",MAX(B7,0),MAX(B6-B8,0))*B14', "本表公式", "仅支持小规模纳税人"),
        ("附加税费比例", float(tax.get("surcharge_rate", 0)), "系统设置", "按所在地优惠及税种维护"),
        ("附加税费（测算）", "=B11*B12", "本表公式", "以实际申报结果为准"),
        ("简易计税率", float(tax.get("vat_rate", 0)), "系统设置", "现行预设1%，政策变化时由用户修改"),
    ]
    for row_index, values in enumerate(vat_rows, start=6):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        if row_index in (12, 14):
            ws.cell(row_index, 2).number_format = PERCENT_FMT
            ws.cell(row_index, 2).font = Font(name="微软雅黑", color=INPUT_BLUE)
            ws.cell(row_index, 2).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        elif row_index != 10:
            ws.cell(row_index, 2).number_format = MONEY_FMT
    ws["B12"].comment = Comment("来源：系统设置。附加税费可能受地区及优惠政策影响。", "User")
    ws["B14"].comment = Comment("来源：系统设置。政策变化时由用户自行复核并修改。", "User")
    _style_data(ws, 6, 14, 4, money_cols=(2,))
    _set_widths(ws, [28, 18, 34, 42])

    # 个税累计预扣准备表
    iit_view = store.individual_income_tax_summary(report_period)
    ws = sheets["个税测算"]
    _sheet_base(ws, "A6")
    _title(ws, "个人所得税累计预扣准备表", f"期间：{report_period} | 基本减除费用来自系统设置", 12)
    iit_headers = ["员工", "累计收入", "累计基本减除", "累计社保公积金", "累计应纳税所得额", "税率", "速算扣除数", "累计应纳税额", "已预扣", "本期测算", "本期已录", "差额"]
    _headers(ws, 5, iit_headers)
    for row_index, item in enumerate(iit_view.get("rows", []), start=6):
        values = [
            item.get("employee_name", ""), _money(item.get("cumulative_income")),
            _money(item.get("basic_deduction")), _money(item.get("other_deductions")), None,
            float(item.get("rate", 0)), _money(item.get("quick_deduction")), None,
            _money(item.get("prior_withheld")), None, _money(item.get("current_recorded")), None,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 5, f"=MAX(B{row_index}-C{row_index}-D{row_index},0)")
        ws.cell(row_index, 8, f"=MAX(E{row_index}*F{row_index}-G{row_index},0)")
        ws.cell(row_index, 10, f"=MAX(H{row_index}-I{row_index},0)")
        ws.cell(row_index, 12, f"=K{row_index}-J{row_index}")
        ws.cell(row_index, 6).number_format = PERCENT_FMT
    iit_end = max(6, len(iit_view.get("rows", [])) + 5)
    _style_data(ws, 6, iit_end, 12, money_cols=(2, 3, 4, 5, 7, 8, 9, 10, 11, 12))
    _set_widths(ws, [16, 14, 16, 17, 18, 10, 14, 16, 14, 14, 14, 14])
    _table(ws, 5, len(iit_view.get("rows", [])) + 5, 12, "IITPreparation")
    iit_total_row = iit_end + 2
    ws.cell(iit_total_row, 1, "合计").font = Font(name="微软雅黑", bold=True)
    for col in (10, 11, 12):
        letter = get_column_letter(col)
        ws.cell(iit_total_row, col, f"=SUM({letter}6:{letter}{iit_end})")
        ws.cell(iit_total_row, col).number_format = MONEY_FMT
        ws.cell(iit_total_row, col).font = Font(name="微软雅黑", bold=True)

    # 印花税准备表
    stamp_view = store.stamp_duty_summary(report_period)
    ws = sheets["印花税准备"]
    _sheet_base(ws, "A6")
    _title(ws, "印花税申报准备表", f"期间：{stamp_view.get('period', {}).get('key', '')} | 减征参数可修改", 9)
    stamp_headers = ["期间", "税目", "合同/凭证号", "对方", "计税金额", "税率", "减征前税额", "减征后比例", "应纳税额"]
    _headers(ws, 5, stamp_headers)
    for row_index, item in enumerate(stamp_view.get("items", []), start=6):
        values = [
            item.get("period", ""), item.get("item", ""), item.get("contract_no", ""),
            item.get("counterparty", ""), _money(item.get("taxable_amount")),
            float(item.get("rate", 0)), None, float(stamp_view.get("relief_rate", 1)), None,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 7, f"=E{row_index}*F{row_index}")
        ws.cell(row_index, 9, f"=G{row_index}*H{row_index}")
        ws.cell(row_index, 6).number_format = PERCENT_FMT
        ws.cell(row_index, 8).number_format = PERCENT_FMT
    stamp_end = max(6, len(stamp_view.get("items", [])) + 5)
    _style_data(ws, 6, stamp_end, 9, money_cols=(5, 7, 9))
    _set_widths(ws, [12, 20, 18, 24, 15, 10, 15, 14, 15])
    _table(ws, 5, len(stamp_view.get("items", [])) + 5, 9, "StampDutyPreparation")
    stamp_total_row = stamp_end + 2
    ws.cell(stamp_total_row, 4, "合计").font = Font(name="微软雅黑", bold=True)
    ws.cell(stamp_total_row, 9, f"=SUM(I6:I{stamp_end})")
    ws.cell(stamp_total_row, 9).number_format = MONEY_FMT
    ws.cell(stamp_total_row, 9).font = Font(name="微软雅黑", bold=True)

    # 企业所得税预缴准备表
    ws = sheets["所得税预缴"]
    _sheet_base(ws)
    _title(ws, "小型微利企业所得税预缴准备表", f"期间：{cit_period_info.get('key', '')} | 累计口径", 4)
    _headers(ws, 5, ["项目", "金额/税率", "计算依据", "复核事项"])
    cit_end_month = str(cit_period_info.get("end_month", report_period))
    adjust_formula = lambda direction: (
        f'=SUMIFS(\'纳税调整\'!$E$6:$E${adjustment_end},\'纳税调整\'!$B$6:$B${adjustment_end},"企业所得税",'
        f'\'纳税调整\'!$D$6:$D${adjustment_end},"{direction}",\'纳税调整\'!$A$6:$A${adjustment_end},"<={cit_end_month}")'
    )
    prepayment_rows = [
        ("本年累计会计利润", _money(cit_view.get("accounting_profit")), "账簿累计", "核对收入成本费用完整性"),
        ("纳税调整增加额", adjust_formula("调增"), "纳税调整台账", "无票支出及限额事项等"),
        ("纳税调整减少额", adjust_formula("调减"), "纳税调整台账", "按现行政策复核"),
        ("弥补以前年度亏损", adjust_formula("弥补以前年度亏损"), "纳税调整台账", "需有可弥补余额依据"),
        ("累计应纳税所得额", "=MAX(B6+B7-B8-B9,0)", "本表公式", "资格阈值也按本行复核"),
        ("有效税率", float(tax.get("cit_rate", 0)), "系统设置", "小型微利企业当前预设5%"),
        ("累计应纳所得税", '=IF(COUNTIF(\'税务期间\'!D10:D13,"不通过")=0,B10*B11,0)', "本表公式", "资格失败时停止自动给税额"),
        ("已预缴所得税", adjust_formula("已预缴所得税"), "纳税调整台账", "录入以前季度已预缴数"),
        ("本期应补所得税", "=MAX(B12-B13,0)", "本表公式", "以电子税务局结果为准"),
    ]
    for row_index, values in enumerate(prepayment_rows, start=6):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 2).number_format = PERCENT_FMT if row_index == 11 else MONEY_FMT
    ws["B11"].font = Font(name="微软雅黑", color=INPUT_BLUE)
    ws["B11"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    _style_data(ws, 6, 14, 4, money_cols=(2,))
    _set_widths(ws, [30, 18, 34, 46])

    # 兼容摘要页，保留原有工作簿链接位置。
    ws = sheets["所得税测算"]
    _sheet_base(ws)
    _title(ws, "企业所得税预缴测算", f"期间：{period_label} | 参数来自系统设置", 4)
    _headers(ws, 5, ["项目", "金额/税率", "计算依据", "复核事项"])
    cit_rows = [
        ("会计利润总额", "='所得税预缴'!B6", "所得税预缴", "本年累计"),
        ("纳税调整增加额", "='所得税预缴'!B7", "纳税调整台账", "软件内维护"),
        ("纳税调整减少额", "='所得税预缴'!B8", "纳税调整台账", "软件内维护"),
        ("应纳税所得额（辅助）", "='所得税预缴'!B10", "所得税预缴", "含弥补亏损"),
        ("有效测算税率", "='所得税预缴'!B11", "系统设置", "资格需通过"),
        ("应纳企业所得税（测算）", "='所得税预缴'!B14", "所得税预缴", "扣除已预缴税额"),
    ]
    for row_index, values in enumerate(cit_rows, start=6):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        if row_index == 10:
            ws.cell(row_index, 2).number_format = PERCENT_FMT
        else:
            ws.cell(row_index, 2).number_format = MONEY_FMT
        if row_index in (7, 8, 10):
            ws.cell(row_index, 2).font = Font(name="微软雅黑", color=INPUT_BLUE)
            ws.cell(row_index, 2).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    ws["B10"].comment = Comment("来源：系统设置。用于测算，不代表法定税率或最终优惠税率。", "User")
    _style_data(ws, 6, 11, 4, money_cols=(2,))
    _set_widths(ws, [30, 18, 34, 44])

    # 年度汇算准备表
    ws = sheets["年度汇算准备"]
    _sheet_base(ws)
    _title(ws, "企业所得税年度汇算准备表", f"年度：{report_period[:4]} | 非电子税务局直传模板", 4)
    _headers(ws, 5, ["项目", "金额/税率", "计算依据", "复核事项"])
    annual_adjust = lambda direction: (
        f'=SUMIFS(\'纳税调整\'!$E$6:$E${adjustment_end},\'纳税调整\'!$B$6:$B${adjustment_end},"企业所得税",'
        f'\'纳税调整\'!$D$6:$D${adjustment_end},"{direction}")'
    )
    annual_rows = [
        ("年度会计利润", _money(annual_view.get("accounting_profit")), "全年账簿", "完成12月结账后复核"),
        ("纳税调整增加额", annual_adjust("调增"), "纳税调整台账", "核对业务招待费、无票支出等"),
        ("纳税调整减少额", annual_adjust("调减"), "纳税调整台账", "核对优惠和免税收入"),
        ("弥补以前年度亏损", annual_adjust("弥补以前年度亏损"), "纳税调整台账", "核对可弥补余额"),
        ("年度应纳税所得额", "=MAX(B6+B7-B8-B9,0)", "本表公式", "不得替代年度纳税申报表"),
        ("有效税率", float(tax.get("cit_rate", 0)), "系统设置", "资格变化时停止使用预设"),
        ("年度应纳所得税", "=B10*B11", "本表公式", "以电子税务局为准"),
        ("全年已预缴所得税", annual_adjust("已预缴所得税"), "纳税调整台账", "按完税记录核对"),
        ("预计应补（退）税额", "=B12-B13", "本表公式", "负数仅表示可能退税，需主管机关确认"),
    ]
    for row_index, values in enumerate(annual_rows, start=6):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 2).number_format = PERCENT_FMT if row_index == 11 else MONEY_FMT
    ws["B11"].font = Font(name="微软雅黑", color=INPUT_BLUE)
    ws["B11"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    _style_data(ws, 6, 14, 4, money_cols=(2,))
    _set_widths(ws, [30, 18, 34, 48])

    # 申报校验
    ws = sheets["申报校验"]
    _sheet_base(ws, "A6")
    _title(ws, "申报前基础数据校验", f"期间：{period_label} | 请先处理错误和警告", 5)
    _headers(ws, 5, ["级别", "代码", "问题说明", "数量", "处理建议"])
    suggestions = {
        "COMPANY_PROFILE": "在系统设置中补齐企业资料",
        "UNBALANCED": "在手工入账或批量导入中修正凭证",
        "DUPLICATE_INVOICE": "核对发票代码、号码并删除重复记录",
        "RED_INVOICE_REFERENCE": "在财税工作台补充原蓝字发票号码",
        "PRICE_TAX_MISMATCH": "重新保存发票，由价税分离引擎校正金额",
        "UNSUPPORTED_TAX_SCOPE": "本版本仅支持小规模纳税人和小型微利企业",
        "SMALL_PROFIT_QUALIFICATION": "核对所得额、人数、资产和限制行业条件",
        "MISSING_INVOICE": "补充发票或其他合规税前扣除凭证",
        "OPENING_UNBALANCED": "在基础账务中补齐或修正借贷双方期初余额",
        "BANK_UNRECONCILED": "在基础账务中完成银行流水自动或手工匹配",
        "PAYROLL_UNPOSTED": "核对工资社保后生成计提凭证",
        "DEPRECIATION_UNPOSTED": "核对固定资产卡片后生成本期折旧凭证",
        "CASH_FLOW_REVIEW": "在基础账务的银行对账页确认现金流量项目",
        "CASH_FLOW_TIE": "核对货币资金凭证及现金流量分类，确保净变动勾稽",
        "TRIAL_BALANCE": "核对期初余额和凭证，确保期末借贷余额平衡",
        "OK": "基础校验通过；仍须人工核对现行申报表及主管机关口径",
    }
    for row_index, issue in enumerate(issues, start=6):
        values = [
            issue.get("level", ""), issue.get("code", ""), issue.get("message", ""),
            issue.get("count", 0), suggestions.get(issue.get("code", ""), "人工复核"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        fill = LIGHT_GREEN if issue.get("level") == "通过" else (
            LIGHT_RED if issue.get("level") == "错误" else LIGHT_YELLOW
        )
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row_index, 1).font = Font(name="微软雅黑", bold=True)
    issue_end = max(6, len(issues) + 5)
    _style_data(ws, 6, issue_end, 5)
    _set_widths(ws, [10, 24, 64, 10, 40])
    _table(ws, 5, len(issues) + 5, 5, "ValidationIssues")

    # 政策依据
    ws = sheets["政策依据"]
    _sheet_base(ws, "A6")
    _title(ws, "政策资料索引", LEGAL_NOTICE_SUMMARY, 5)
    _headers(ws, 5, ["资料标题", "发布机构", "引用/访问日期", "网址", "适用提示"])
    policy_sources = settings.get("policy_sources", [])
    for row_index, source in enumerate(policy_sources, start=6):
        values = [
            source.get("title", ""), source.get("publisher", ""), source.get("date", ""),
            source.get("url", ""), "请打开原文并确认现行有效性及具体适用条件",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 4).hyperlink = source.get("url", "")
        ws.cell(row_index, 4).style = "Hyperlink"
    policy_end = max(6, len(policy_sources) + 5)
    _style_data(ws, 6, policy_end, 5)
    _set_widths(ws, [46, 18, 16, 64, 44])
    _table(ws, 5, len(policy_sources) + 5, 5, "PolicySources")

    # 一人公司月度总览：先看状态和关键数字，再下钻到明细。
    ws = sheets["月度总览"]
    _sheet_base(ws, "A15")
    _title(
        ws, "一人公司月度报税准备总览",
        f"企业：{company_name} | 期间：{period_label} | 生成时间：{generated_at}", 8,
    )
    error_count = sum(1 for issue in issues if issue.get("level") == "错误")
    warning_count = sum(1 for issue in issues if issue.get("level") == "警告")
    if not checklist["ready"]:
        status_text = f"暂不建议申报：仍有 {checklist['blocking_count']} 项关账检查需要处理"
        status_fill = LIGHT_RED
    elif error_count:
        status_text = f"暂不建议申报：申报校验仍有 {error_count} 项错误"
        status_fill = LIGHT_RED
    elif warning_count:
        status_text = f"关账已完成，但仍有 {warning_count} 项申报警告需要人工复核"
        status_fill = LIGHT_YELLOW
    else:
        status_text = "基础校验通过，可导出申报准备底稿；不代表申报口径、合规性或受理结果已确认"
        status_fill = LIGHT_GREEN
    ws.merge_cells("A4:H4")
    ws["A4"] = status_text
    ws["A4"].font = Font(name="微软雅黑", size=11, bold=True, color=TEXT)
    ws["A4"].fill = PatternFill("solid", fgColor=status_fill)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 28
    ws.merge_cells("A3:H3")
    ws["A3"] = LEGAL_NOTICE_SUMMARY
    ws["A3"].font = Font(name="微软雅黑", size=9, color="7A4E00")
    ws["A3"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 36

    top_metrics = (
        (1, "营业收入", "='利润表辅助'!B6"),
        (3, "成本及期间费用", "='利润表辅助'!B7+'利润表辅助'!B8"),
        (5, "利润总额", "='利润表辅助'!B9"),
        (
            7, "预计申报税额合计",
            f"=A11+C11+E11+'印花税准备'!I{stamp_total_row}+'个税测算'!J{iit_total_row}",
        ),
    )
    tax_metrics = (
        (1, "预计增值税", "='增值税测算'!B11"),
        (3, "预计附加税费", "='增值税测算'!B13"),
        (5, "预计企业所得税", "='所得税测算'!B11"),
    )
    for col, label, formula in top_metrics:
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.merge_cells(start_row=7, start_column=col, end_row=8, end_column=col + 1)
        ws.cell(6, col, label)
        ws.cell(7, col, formula)
        ws.cell(6, col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(6, col).font = Font(name="微软雅黑", bold=True, color=WHITE)
        ws.cell(6, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(7, col).font = Font(name="微软雅黑", size=14, bold=True, color=DARK)
        ws.cell(7, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(7, col).number_format = MONEY_FMT
    for col, label, formula in tax_metrics:
        ws.merge_cells(start_row=10, start_column=col, end_row=10, end_column=col + 1)
        ws.merge_cells(start_row=11, start_column=col, end_row=12, end_column=col + 1)
        ws.cell(10, col, label)
        ws.cell(11, col, formula)
        ws.cell(10, col).fill = PatternFill("solid", fgColor=TEAL)
        ws.cell(10, col).font = Font(name="微软雅黑", bold=True, color=WHITE)
        ws.cell(10, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(11, col).font = Font(name="微软雅黑", size=12, bold=True, color=DARK)
        ws.cell(11, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(11, col).number_format = MONEY_FMT

    exact_period_vouchers = [
        row for row in all_vouchers if str(row.get("period", "")) == report_period
    ]
    voucher_count = len({row.get("voucher_no") for row in exact_period_vouchers})
    ws.merge_cells("G10:H10")
    ws.merge_cells("G11:H12")
    ws["G10"] = "本期数据量"
    ws["G10"].fill = PatternFill("solid", fgColor=TEAL)
    ws["G10"].font = Font(name="微软雅黑", bold=True, color=WHITE)
    ws["G10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G11"] = (
        f"{voucher_count} 张凭证 | {len(invoices)} 张发票 | "
        f"{len(bank_transactions)} 条银行流水"
    )
    ws["G11"].font = Font(name="微软雅黑", size=10, bold=True, color=DARK)
    ws["G11"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A14:H14")
    ws["A14"] = "月末关账清单"
    ws["A14"].fill = PatternFill("solid", fgColor=DARK)
    ws["A14"].font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
    ws["A14"].alignment = Alignment(horizontal="left", vertical="center")
    _headers(ws, 15, ["状态", "检查项", "检查说明", "", "", "", "打开位置", ""])
    action_sheets = {
        "企业资料": "使用说明", "期初余额": "期初余额", "记账凭证": "记账凭证",
        "银行对账": "银行对账", "工资社保": "工资社保",
        "固定资产折旧": "折旧明细", "现金流量分类": "银行对账",
        "税务资格与期间": "税务期间", "税费计提": "所得税预缴",
        "损益结转": "利润表辅助", "试算平衡与报表勾稽": "科目余额表",
    }
    checklist_start = 16
    for row_index, item in enumerate(checklist["items"], start=checklist_start):
        ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=6)
        ws.merge_cells(start_row=row_index, start_column=7, end_row=row_index, end_column=8)
        ws.cell(row_index, 1, item.get("status", ""))
        ws.cell(row_index, 2, item.get("item", ""))
        ws.cell(row_index, 3, item.get("detail", ""))
        target_sheet = action_sheets.get(item.get("item", ""), "申报校验")
        ws.cell(row_index, 7, f"打开{target_sheet}")
        ws.cell(row_index, 7).hyperlink = f"#'{target_sheet}'!A1"
        ws.cell(row_index, 7).style = "Hyperlink"
        status = item.get("status", "")
        ws.cell(row_index, 1).fill = PatternFill(
            "solid", fgColor=LIGHT_GREEN if status == "通过" else (
                LIGHT_RED if status == "待处理" else LIGHT_YELLOW
            )
        )
        ws.cell(row_index, 1).font = Font(name="微软雅黑", bold=True, color=TEXT)
        for col in range(1, 9):
            ws.cell(row_index, col).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_index].height = 30

    next_row = checklist_start + len(checklist["items"]) + 1
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=8)
    ws.cell(next_row, 1, "申报准备顺序")
    ws.cell(next_row, 1).fill = PatternFill("solid", fgColor=TEAL)
    ws.cell(next_row, 1).font = Font(name="微软雅黑", bold=True, color=WHITE)
    for offset, (label, target_sheet) in enumerate((
        ("1. 先处理“申报校验”中的错误和警告", "申报校验"),
        ("2. 核对会小企财务报表与科目余额表", report_names["balance"]),
        ("3. 核对增值税、附加税费和企业所得税测算", "增值税测算"),
        ("4. 登录电子税务局，按主管机关表单逐项填报并保存回执", "所得税测算"),
    ), start=1):
        row_index = next_row + offset
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 1).hyperlink = f"#'{target_sheet}'!A1"
        ws.cell(row_index, 1).style = "Hyperlink"
        ws.cell(row_index, 1).alignment = Alignment(vertical="center")
    _set_widths(ws, [13, 22, 18, 18, 18, 18, 17, 17])

    for ws in wb.worksheets:
        ws.oddHeader.center.text = f"&B{company_name} - {ws.title}"
        ws.oddFooter.left.text = "免费财税申报辅助材料 | 申报前须人工复核"
        ws.oddFooter.right.text = "第 &P 页，共 &N 页"
        ws.print_options.horizontalCentered = True
        ws.auto_filter.ref = ws.auto_filter.ref

    wb.save(output_path)
    return output_path
