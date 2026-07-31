#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import invoice drafts from tax-system Excel query exports."""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency error is reported to the user
    xlrd = None


BASE_SHEET_ALIASES = ("发票基础信息", "发票信息", "基础信息")
DETAIL_SHEET_ALIASES = ("信息汇总表", "发票明细", "明细信息")
VOID_STATUSES = {"作废", "已作废", "异常", "失控", "红冲作废"}


class InvoiceExcelImportError(ValueError):
    """Raised when a workbook is not a supported invoice export."""


class _XlsSheetAdapter:
    """Expose the small openpyxl worksheet surface used by this importer."""

    def __init__(self, sheet, datemode: int):
        self._sheet = sheet
        self._datemode = datemode
        self.title = sheet.name

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        start = max(int(min_row or 1) - 1, 0)
        stop = self._sheet.nrows if max_row is None else min(int(max_row), self._sheet.nrows)
        for row_index in range(start, stop):
            values = []
            for column_index in range(self._sheet.ncols):
                cell = self._sheet.cell(row_index, column_index)
                value = cell.value
                if xlrd is not None and cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, self._datemode)
                    except (TypeError, ValueError, xlrd.XLDateError):
                        pass
                values.append(value)
            yield tuple(values)


class _XlsWorkbookAdapter:
    def __init__(self, book):
        self._book = book
        self.sheetnames = list(book.sheet_names())
        self.worksheets = [
            _XlsSheetAdapter(book.sheet_by_name(name), book.datemode)
            for name in self.sheetnames
        ]
        self._worksheets_by_name = {sheet.title: sheet for sheet in self.worksheets}

    def __getitem__(self, name):
        return self._worksheets_by_name[name]

    def close(self):
        release = getattr(self._book, "release_resources", None)
        if callable(release):
            release()


def _open_workbook(source: Path):
    if source.suffix.lower() == ".xls":
        if xlrd is None:
            raise InvoiceExcelImportError(
                "当前安装缺少旧版Excel读取组件，请安装完整更新后重试"
            )
        return _XlsWorkbookAdapter(xlrd.open_workbook(str(source), on_demand=True))
    return load_workbook(source, read_only=True, data_only=True)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return format(value, ".0f")
        return format(value, ".15g")
    return str(value).strip()


def _money(value: Any) -> float:
    text = _text(value).replace(",", "").replace("￥", "").replace("¥", "")
    if not text:
        return 0.0
    try:
        return float(Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return 0.0


def _date_text(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = _text(value)
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if not match:
        return ""
    try:
        return date(*(int(part) for part in match.groups())).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def _normalized_tax_id(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value)).upper()


def _header_map(row: Iterable[Any]) -> Dict[str, int]:
    return {_text(value).replace("\n", "").replace(" ", ""): index for index, value in enumerate(row)}


def _value(row: Tuple[Any, ...], headers: Dict[str, int], *names: str) -> Any:
    for name in names:
        index = headers.get(name.replace(" ", ""))
        if index is not None and index < len(row):
            return row[index]
    return None


def _invoice_key(row: Tuple[Any, ...], headers: Dict[str, int]) -> Optional[Tuple[str, str]]:
    digital = _text(_value(row, headers, "数电发票号码", "全电发票号码"))
    if digital:
        return "digital", digital
    code = _text(_value(row, headers, "发票代码"))
    number = _text(_value(row, headers, "发票号码"))
    if code or number:
        return code, number
    return None


def _find_sheet(workbook, aliases: Iterable[str], required_headers: Iterable[str]):
    for alias in aliases:
        if alias in workbook.sheetnames:
            return workbook[alias]
    required = set(required_headers)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        first = next(rows, ())
        if required.issubset(_header_map(first)):
            return sheet
    return None


def _sheet_row_count(sheet) -> int:
    """Return a best-effort row count for both openpyxl and xlrd adapters."""
    value = getattr(sheet, "max_row", None)
    if value is None:
        value = getattr(getattr(sheet, "_sheet", None), "nrows", None)
    try:
        return max(int(value or 0), 1)
    except (TypeError, ValueError):
        return 1


def _rows(
    sheet,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
    stage: str = "",
):
    iterator = sheet.iter_rows(values_only=True)
    headers = _header_map(next(iterator, ()))
    total = _sheet_row_count(sheet)
    for row_number, row in enumerate(iterator, start=2):
        if progress_callback and (row_number == 2 or row_number % 25 == 0 or row_number >= total):
            progress_callback(min(row_number, total), total, stage)
        if any(value not in (None, "") for value in row):
            yield row_number, row, headers


def _tax_categories(descriptions: Iterable[str]) -> List[str]:
    values = []
    seen = set()
    for description in descriptions:
        for category in re.findall(r"\*([^*]{1,30})\*", description):
            cleaned = category.strip()
            if cleaned and cleaned not in seen:
                values.append(cleaned)
                seen.add(cleaned)
    return values


def _is_non_postable_status(status: str) -> bool:
    compact = str(status or "").replace(" ", "")
    return (
        compact in VOID_STATUSES
        or "作废" in compact
        or "异常" in compact
        or "失控" in compact
        or compact.startswith("已红冲")
    )


def read_tax_invoice_workbook(
    source: Path,
    company_tax_id: str = "",
    company_industry: str = "",
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> List[Dict[str, Any]]:
    """Return one review draft per invoice from a tax-system xls/xlsx/xlsm export."""
    source = Path(source).resolve()
    if source.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
        raise InvoiceExcelImportError("仅支持税务系统导出的 .xls、.xlsx 或 .xlsm 文件")
    try:
        workbook = _open_workbook(source)
    except Exception as exc:
        raise InvoiceExcelImportError(f"无法读取Excel文件：{exc}") from exc

    base_sheet = _find_sheet(
        workbook, BASE_SHEET_ALIASES, ("开票日期", "价税合计")
    )
    detail_sheet = _find_sheet(
        workbook, DETAIL_SHEET_ALIASES, ("开票日期", "货物或应税劳务名称")
    )
    if base_sheet is None and detail_sheet is None:
        workbook.close()
        raise InvoiceExcelImportError(
            "未找到“发票基础信息”或“信息汇总表”，请选择税务系统的全量发票查询导出文件"
        )

    details: Dict[Tuple[str, str], Dict[str, Any]] = defaultdict(
        lambda: {"descriptions": [], "amount": 0.0, "tax": 0.0, "total": 0.0}
    )
    if detail_sheet is not None:
        for _row_number, row, headers in _rows(
            detail_sheet, progress_callback, "正在读取发票明细"
        ):
            key = _invoice_key(row, headers)
            if key is None:
                continue
            description = _text(_value(row, headers, "货物或应税劳务名称", "项目名称"))
            if description and description not in details[key]["descriptions"]:
                details[key]["descriptions"].append(description)
            details[key]["amount"] += _money(_value(row, headers, "金额"))
            details[key]["tax"] += _money(_value(row, headers, "税额"))
            details[key]["total"] += _money(_value(row, headers, "价税合计"))

    using_detail_as_base = base_sheet is None
    base_rows = (
        list(_rows(base_sheet, progress_callback, "正在读取发票基础信息"))
        if base_sheet is not None
        else []
    )
    if not base_rows and detail_sheet is not None:
        base_rows = list(
            _rows(detail_sheet, progress_callback, "正在整理发票基础信息")
        )

    normalized_company_id = _normalized_tax_id(company_tax_id)
    drafts: List[Dict[str, Any]] = []
    seen_keys = set()
    total_base_rows = max(len(base_rows), 1)
    for base_index, (row_number, row, headers) in enumerate(base_rows, start=1):
        if progress_callback and (base_index == 1 or base_index % 25 == 0 or base_index == total_base_rows):
            progress_callback(base_index, total_base_rows, "正在整理发票记录")
        key = _invoice_key(row, headers)
        if key is None or key in seen_keys:
            continue
        seen_keys.add(key)
        detail = details.get(key, {})
        digital_no = _text(_value(row, headers, "数电发票号码", "全电发票号码"))
        invoice_code = _text(_value(row, headers, "发票代码"))
        invoice_no = digital_no or _text(_value(row, headers, "发票号码"))
        seller_tax_id = _normalized_tax_id(_value(row, headers, "销方识别号", "销售方识别号"))
        buyer_tax_id = _normalized_tax_id(_value(row, headers, "购方识别号", "购买方识别号"))
        seller = _text(_value(row, headers, "销方名称", "销售方名称"))
        buyer = _text(_value(row, headers, "购方名称", "购买方名称"))
        descriptions = list(detail.get("descriptions", []))
        direct_description = _text(_value(row, headers, "货物或应税劳务名称", "项目名称"))
        if direct_description and direct_description not in descriptions:
            descriptions.append(direct_description)
        description = "；".join(descriptions) or f"{seller or '销售方'}开具的发票"

        base_amount = _money(_value(row, headers, "金额"))
        base_tax = _money(_value(row, headers, "税额"))
        base_total = _money(_value(row, headers, "价税合计"))
        net_amount = (
            _money(detail.get("amount"))
            if using_detail_as_base and detail else base_amount
        )
        tax_amount = (
            _money(detail.get("tax"))
            if using_detail_as_base and detail else base_tax
        )
        total_amount = (
            _money(detail.get("total"))
            if using_detail_as_base and detail else base_total
        )
        if not total_amount:
            total_amount = round(net_amount + tax_amount, 2)

        positive_text = _text(_value(row, headers, "是否正数发票"))
        is_red = total_amount < 0 or positive_text in {"否", "N", "No", "0", "红字"}
        invoice_status = _text(_value(row, headers, "发票状态")) or "正常"
        risk_level = _text(_value(row, headers, "发票风险等级"))
        warnings = []
        non_postable = _is_non_postable_status(invoice_status)
        if non_postable:
            warnings.append(f"发票状态为“{invoice_status}”，不可入账")
        if is_red:
            warnings.append("检测到红字/负数发票，已按绝对金额进入复核，入账前请核对借贷方向")
        if risk_level and risk_level not in {"正常", "低风险", "无风险"}:
            warnings.append(f"发票风险等级为“{risk_level}”，请人工核验")

        invoice_type = "进项"
        direction = "借方"
        if normalized_company_id and normalized_company_id == seller_tax_id:
            invoice_type = "销项"
            direction = "贷方"
        elif normalized_company_id and normalized_company_id == buyer_tax_id:
            invoice_type = "进项"
        elif normalized_company_id:
            warnings.append("购销双方税号均与当前企业税号不一致，请核对发票方向")

        drafts.append({
            "file_name": f"{source.name} · {invoice_no or row_number}",
            "filepath": str(source),
            "source_type": "tax_excel",
            "source_row": row_number,
            "status": "不可入账" if non_postable else "待处理",
            "non_postable": non_postable,
            "amount": abs(total_amount),
            "net_amount": abs(net_amount),
            "tax_amount": abs(tax_amount),
            "signed_total_amount": total_amount,
            "signed_net_amount": net_amount,
            "signed_tax_amount": tax_amount,
            "description": description,
            "item_descriptions": descriptions,
            "tax_categories": _tax_categories(descriptions),
            "company_industry": company_industry,
            "invoice_date": _date_text(_value(row, headers, "开票日期")),
            "invoice_code": invoice_code,
            "invoice_no": invoice_no,
            "seller": seller,
            "seller_tax_id": seller_tax_id,
            "buyer": buyer,
            "buyer_tax_id": buyer_tax_id,
            "invoice_type": invoice_type,
            "document_type": "红字发票" if is_red else "正常发票",
            "invoice_form": _text(_value(row, headers, "发票票种")) or "普通发票",
            "invoice_status": invoice_status,
            "risk_level": risk_level,
            "direction": direction,
            "matched_subject": "",
            "match_score": 0.0,
            "match_type": "",
            "confidence": 1.0,
            "needs_review": bool(warnings),
            "warnings": warnings,
        })

    workbook.close()
    if not drafts:
        raise InvoiceExcelImportError("Excel中没有找到带发票号码的有效数据行")
    return drafts
