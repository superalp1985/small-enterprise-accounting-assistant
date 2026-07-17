#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import json
import tkinter as tk
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any, Callable, Dict, Iterable, List, Optional

from openpyxl import load_workbook

import logger as L
from modules.vocabulary_module import load_vocab


BG = "#F0F0F0"
BLUE = "#0078D4"
DARK = "#003087"
WHITE = "#FFFFFF"
GREEN = "#107C10"
RED = "#D83B01"
ORANGE = "#E67E22"
GRAY = "#D0D0D0"
YELLOW = "#FFF4CE"
FONT = ("微软雅黑", 10)
FONT_B = ("微软雅黑", 10, "bold")
FONT_T = ("微软雅黑", 14, "bold")
FONT_S = ("微软雅黑", 9)


def make_btn(parent, text, command, color=BLUE, width=10):
    return tk.Button(
        parent, text=text, command=command, bg=color, fg=WHITE,
        font=FONT_B, relief="flat", padx=8, pady=4, width=width,
        activebackground=DARK, activeforeground=WHITE, cursor="hand2",
    )


def _number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    text = str(value).replace(",", "").replace("¥", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def _read_tabular(path: Path) -> List[Dict[str, Any]]:
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:] if any(value not in (None, "") for value in row)
        ]

    content = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            content = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if content is None:
        raise ValueError("无法识别文件编码")
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    return list(csv.DictReader(content.splitlines(), delimiter=delimiter))


def _pick(row: Dict[str, Any], *headers: str):
    normalized = {str(key).strip().replace(" ", ""): value for key, value in row.items()}
    for header in headers:
        key = header.replace(" ", "")
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    return ""


class BasicAccountingModule(tk.Frame):
    """Opening balances, bank reconciliation, payroll, and fixed assets."""

    def __init__(self, parent, config, store, authenticated=False,
                 on_data_changed: Optional[Callable] = None):
        super().__init__(parent, bg=BG)
        self.parent = parent
        self.config = config
        self.store = store
        self.authenticated = authenticated
        self.on_data_changed = on_data_changed
        self.enterprise_mode = self.store.profile_key == "enterprise"
        self.subject_options = self._load_subjects()
        current_period = date.today().strftime("%Y-%m")
        self.opening_period_var = tk.StringVar(value=f"{date.today().year}-01")
        self.bank_period_var = tk.StringVar(value=current_period)
        self.payroll_period_var = tk.StringVar(value=current_period)
        self.asset_period_var = tk.StringVar(value=current_period)
        self.status_var = tk.StringVar(value="就绪")
        self._build_ui()
        self.reload_from_store()

    def _load_subjects(self) -> List[str]:
        vocab = load_vocab(
            self.config.vocab_path,
            getattr(self.config, "account_catalog_path", None),
        )
        enabled_codes = set(self.store.enabled_account_codes())
        catalog_subjects = [
            f"{row.get('code', '')} {row.get('name', '')}".strip()
            for row in self.store.enabled_accounts()
        ]
        detail_subjects = [
            row["subject"] for row in vocab
            if row.get("subject") and str(row.get("subject_code", "")) in enabled_codes
        ]
        return list(dict.fromkeys(catalog_subjects + detail_subjects))

    def _build_ui(self):
        outer = tk.LabelFrame(
            self, text=" 基础账务 ", font=FONT_T, bg=BG, fg=DARK,
            bd=1, relief="groove",
        )
        outer.pack(fill="both", expand=True, pady=6)
        notebook = ttk.Notebook(outer)
        notebook.pack(fill="both", expand=True, padx=12, pady=10)
        self.opening_tab = tk.Frame(notebook, bg=BG)
        self.bank_tab = tk.Frame(notebook, bg=BG)
        self.payroll_tab = tk.Frame(notebook, bg=BG)
        self.asset_tab = tk.Frame(notebook, bg=BG)
        notebook.add(self.opening_tab, text="期初余额")
        notebook.add(self.bank_tab, text="银行对账")
        notebook.add(self.payroll_tab, text="工资社保")
        notebook.add(self.asset_tab, text="固定资产")
        self._build_opening_tab()
        self._build_bank_tab()
        self._build_payroll_tab()
        self._build_asset_tab()
        tk.Label(
            outer, textvariable=self.status_var, font=FONT_S, bg=GRAY,
            fg="#444", anchor="w", padx=10, pady=5,
        ).pack(fill="x", padx=12, pady=(0, 10))

    def _period_toolbar(self, parent, variable, refresh_command):
        bar = tk.Frame(parent, bg=BG, pady=8)
        bar.pack(fill="x", padx=10)
        tk.Label(bar, text="期间：", font=FONT_B, bg=BG).pack(side="left")
        entry = ttk.Combobox(bar, textvariable=variable, width=12, font=FONT)
        entry.pack(side="left", padx=(0, 6))
        entry.bind("<Return>", lambda event: refresh_command())
        entry.bind("<<ComboboxSelected>>", lambda event: refresh_command())
        return bar, entry

    def _tree(self, parent, columns, widths):
        frame = tk.Frame(parent, bg=BG)
        frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        for column, width in zip(columns, widths):
            tree.heading(column, text=column)
            tree.column(column, width=width, anchor="center")
        yscroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        return tree

    def _build_opening_tab(self):
        bar, self.opening_period_combo = self._period_toolbar(
            self.opening_tab, self.opening_period_var, self._refresh_opening
        )
        make_btn(bar, "新增", self._add_opening, color=GREEN, width=8).pack(side="left", padx=3)
        make_btn(bar, "编辑", self._edit_opening, width=8).pack(side="left", padx=3)
        make_btn(bar, "删除", self._delete_opening, color=RED, width=8).pack(side="left", padx=3)
        make_btn(bar, "导入Excel/CSV", self._import_opening, color=BLUE, width=14).pack(side="left", padx=3)
        make_btn(bar, "刷新", self._refresh_opening, color="#666", width=8).pack(side="right", padx=3)
        self.opening_total_var = tk.StringVar()
        tk.Label(
            self.opening_tab, textvariable=self.opening_total_var, font=FONT_B,
            bg=YELLOW, fg="#5A4500", anchor="w", padx=10, pady=7,
        ).pack(fill="x", padx=10, pady=(0, 8))
        columns = ("期间", "科目", "借方余额", "贷方余额", "备注")
        self.opening_tree = self._tree(self.opening_tab, columns, (90, 280, 130, 130, 260))
        self.opening_tree.bind("<Double-1>", lambda event: self._edit_opening())

    def _build_bank_tab(self):
        bar, self.bank_period_combo = self._period_toolbar(
            self.bank_tab, self.bank_period_var, self._refresh_bank
        )
        make_btn(bar, "导入流水", self._import_bank, color=GREEN, width=10).pack(side="left", padx=3)
        make_btn(bar, "自动匹配", self._auto_reconcile, color=BLUE, width=10).pack(side="left", padx=3)
        make_btn(bar, "手工匹配", self._manual_reconcile, color=BLUE, width=10).pack(side="left", padx=3)
        make_btn(bar, "取消匹配", self._unmatch_bank, color=ORANGE, width=10).pack(side="left", padx=3)
        if self.enterprise_mode:
            make_btn(
                bar, "现金流项目", self._set_cash_flow_category,
                color=BLUE, width=11,
            ).pack(side="left", padx=3)
        make_btn(bar, "删除", self._delete_bank, color=RED, width=8).pack(side="left", padx=3)
        make_btn(bar, "刷新", self._refresh_bank, color="#666", width=8).pack(side="right", padx=3)
        self.bank_summary_var = tk.StringVar()
        tk.Label(
            self.bank_tab, textvariable=self.bank_summary_var, font=FONT_B,
            bg=YELLOW, fg="#5A4500", anchor="w", padx=10, pady=7,
        ).pack(fill="x", padx=10, pady=(0, 8))
        if self.enterprise_mode:
            columns = (
                "日期", "方向", "金额", "摘要", "对方户名", "匹配凭证",
                "现金流项目", "分类来源", "状态",
            )
            widths = (100, 70, 110, 240, 180, 125, 290, 110, 90)
        else:
            columns = ("日期", "方向", "金额", "摘要", "对方户名", "匹配凭证", "状态")
            widths = (100, 70, 110, 260, 200, 130, 90)
        self.bank_tree = self._tree(self.bank_tab, columns, widths)
        if self.enterprise_mode:
            self.bank_tree.bind("<Double-1>", lambda event: self._set_cash_flow_category())

    def _build_payroll_tab(self):
        bar, self.payroll_period_combo = self._period_toolbar(
            self.payroll_tab, self.payroll_period_var, self._refresh_payroll
        )
        make_btn(bar, "新增", self._add_payroll, color=GREEN, width=8).pack(side="left", padx=3)
        make_btn(bar, "编辑", self._edit_payroll, width=8).pack(side="left", padx=3)
        make_btn(bar, "删除", self._delete_payroll, color=RED, width=8).pack(side="left", padx=3)
        make_btn(bar, "生成计提凭证", self._post_payroll, color=BLUE, width=14).pack(side="left", padx=3)
        make_btn(bar, "刷新", self._refresh_payroll, color="#666", width=8).pack(side="right", padx=3)
        self.payroll_summary_var = tk.StringVar()
        tk.Label(
            self.payroll_tab, textvariable=self.payroll_summary_var, font=FONT_B,
            bg=YELLOW, fg="#5A4500", anchor="w", padx=10, pady=7,
        ).pack(fill="x", padx=10, pady=(0, 8))
        columns = (
            "员工", "应发工资", "个人社保", "个人公积金", "个税",
            "实发工资", "单位社保", "单位公积金", "状态", "凭证号",
        )
        self.payroll_tree = self._tree(
            self.payroll_tab, columns, (100, 100, 95, 100, 85, 100, 95, 100, 80, 130)
        )
        self.payroll_tree.bind("<Double-1>", lambda event: self._edit_payroll())

    def _build_asset_tab(self):
        bar, self.asset_period_combo = self._period_toolbar(
            self.asset_tab, self.asset_period_var, self._refresh_assets
        )
        make_btn(bar, "新增资产", self._add_asset, color=GREEN, width=10).pack(side="left", padx=3)
        make_btn(bar, "编辑", self._edit_asset, width=8).pack(side="left", padx=3)
        make_btn(bar, "删除", self._delete_asset, color=RED, width=8).pack(side="left", padx=3)
        make_btn(bar, "一键生成折旧凭证", self._post_depreciation, color=BLUE, width=16).pack(
            side="left", padx=3
        )
        make_btn(bar, "刷新", self._refresh_assets, color="#666", width=8).pack(side="right", padx=3)
        self.asset_summary_var = tk.StringVar()
        tk.Label(
            self.asset_tab, textvariable=self.asset_summary_var, font=FONT_B,
            bg=YELLOW, fg="#5A4500", anchor="w", padx=10, pady=7,
        ).pack(fill="x", padx=10, pady=(0, 8))
        columns = (
            "资产名称", "类别", "购置日期", "原值", "使用月数", "月折旧",
            "本期折旧", "累计折旧", "账面净值", "状态", "凭证号",
        )
        self.asset_tree = self._tree(
            self.asset_tab, columns, (150, 90, 100, 105, 85, 100, 100, 105, 105, 75, 125)
        )
        self.asset_tree.bind("<Double-1>", lambda event: self._edit_asset())

    def _selected_id(self, tree, label):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("提示", f"请先选择{label}")
            return None
        return selected[0]

    def _valid_period(self, variable):
        value = variable.get().strip()
        try:
            datetime.strptime(value, "%Y-%m")
        except ValueError:
            messagebox.showwarning("期间格式", "期间应为 YYYY-MM 格式")
            return None
        return value

    def _record_dialog(self, title: str, fields: Iterable, initial: Dict[str, Any],
                       save: Callable[[Dict[str, Any]], None], geometry="660x620"):
        window = tk.Toplevel(self)
        window.title(title)
        window.configure(bg=BG)
        window.geometry(geometry)
        window.transient(self.winfo_toplevel())
        window.grab_set()
        variables = {}
        for row, field in enumerate(fields):
            key, label, kind = field
            tk.Label(window, text=label, font=FONT_B, bg=BG).grid(
                row=row, column=0, sticky="w", padx=18, pady=7
            )
            value = initial.get(key, "")
            variable = tk.StringVar(value=str(value))
            variables[key] = variable
            if isinstance(kind, (list, tuple)):
                widget = ttk.Combobox(
                    window, textvariable=variable, values=list(kind),
                    state="readonly", font=FONT,
                )
            else:
                widget = tk.Entry(window, textvariable=variable, font=FONT, relief="solid", bd=1)
            widget.grid(row=row, column=1, sticky="ew", padx=(0, 18), pady=7)
        window.columnconfigure(1, weight=1)

        def submit():
            payload = {key: variable.get().strip() for key, variable in variables.items()}
            try:
                save(payload)
            except Exception as exc:
                messagebox.showerror("保存失败", str(exc), parent=window)
                return
            window.destroy()

        footer = tk.Frame(window, bg=BG)
        footer.grid(row=len(list(fields)), column=0, columnspan=2, sticky="e", padx=18, pady=16)
        make_btn(footer, "保存", submit, color=GREEN, width=10).pack(side="left", padx=4)
        make_btn(footer, "取消", window.destroy, color="#666", width=9).pack(side="left", padx=4)

    def _add_opening(self):
        self._opening_dialog({"period": self.opening_period_var.get(), "direction": "借方"})

    def _edit_opening(self):
        record_id = self._selected_id(self.opening_tree, "期初余额")
        if not record_id:
            return
        record = next((row for row in self.store.list_opening_balances() if row["id"] == record_id), None)
        if not record:
            return
        initial = {
            **record,
            "direction": "借方" if record.get("debit_balance") else "贷方",
            "amount": record.get("debit_balance") or record.get("credit_balance"),
        }
        self._opening_dialog(initial)

    def _opening_dialog(self, initial):
        record_id = initial.get("id", "")
        fields = [
            ("period", "期初期间", None),
            ("subject", "会计科目", self.subject_options),
            ("direction", "余额方向", ["借方", "贷方"]),
            ("amount", "期初余额", None),
            ("note", "备注", None),
        ]

        def save(payload):
            amount = _number(payload["amount"])
            self.store.upsert_opening_balance({
                "id": record_id, "period": payload["period"], "subject": payload["subject"],
                "debit_balance": amount if payload["direction"] == "借方" else 0,
                "credit_balance": amount if payload["direction"] == "贷方" else 0,
                "note": payload["note"],
            })
            L.log("保存期初余额", f"{payload['period']}/{payload['subject']}/{amount}")
            self._refresh_opening()
            self._notify()

        self._record_dialog("期初余额", fields, initial, save, "680x420")

    def _delete_opening(self):
        record_id = self._selected_id(self.opening_tree, "期初余额")
        if record_id and messagebox.askyesno("确认删除", "删除选中的期初余额？"):
            try:
                self.store.delete_opening_balance(record_id)
                self._refresh_opening()
                self._notify()
            except Exception as exc:
                messagebox.showerror("删除失败", str(exc))

    def _import_opening(self):
        path = filedialog.askopenfilename(
            parent=self, title="导入期初余额",
            filetypes=[("表格文件", "*.xlsx *.csv *.tsv")],
        )
        if not path:
            return
        try:
            rows = _read_tabular(Path(path))
            imported = 0
            for row in rows:
                code = str(_pick(row, "科目编码", "科目代码", "编码")).strip()
                name = str(_pick(row, "科目名称", "科目", "名称")).strip()
                subject = name if name.startswith(code) else f"{code} {name}".strip()
                self.store.upsert_opening_balance({
                    "period": str(_pick(row, "期间", "期初期间") or self.opening_period_var.get())[:7],
                    "subject": subject,
                    "debit_balance": _number(_pick(row, "借方余额", "期初借方", "借方")),
                    "credit_balance": _number(_pick(row, "贷方余额", "期初贷方", "贷方")),
                    "note": str(_pick(row, "备注", "说明")),
                })
                imported += 1
            self._refresh_opening()
            self._notify()
            messagebox.showinfo("导入完成", f"已导入或更新 {imported} 条期初余额")
        except Exception as exc:
            messagebox.showerror("导入失败", str(exc))

    def _refresh_opening(self):
        period = self._valid_period(self.opening_period_var)
        if not period:
            return
        for item in self.opening_tree.get_children():
            self.opening_tree.delete(item)
        records = self.store.list_opening_balances(period)
        for row in records:
            self.opening_tree.insert("", "end", iid=row["id"], values=(
                row["period"], row["subject"], f"¥{row['debit_balance']:,.2f}",
                f"¥{row['credit_balance']:,.2f}", row.get("note", ""),
            ))
        totals = self.store.opening_balance_totals(period)
        status = "平衡" if abs(totals["difference"]) < 0.01 else "不平衡"
        self.opening_total_var.set(
            f"借方合计 ¥{totals['debit']:,.2f}    贷方合计 ¥{totals['credit']:,.2f}    "
            f"差额 ¥{totals['difference']:,.2f}    {status}"
        )

    def _import_bank(self):
        path = filedialog.askopenfilename(
            parent=self, title="导入银行流水",
            filetypes=[("表格文件", "*.xlsx *.csv *.tsv")],
        )
        if not path:
            return
        try:
            source_rows = _read_tabular(Path(path))
            records = []
            for row in source_rows:
                income = _number(_pick(row, "收入金额", "贷方发生额", "收入"))
                expense = _number(_pick(row, "支出金额", "借方发生额", "支出"))
                direction = str(_pick(row, "方向", "收支方向", "交易方向")).strip()
                amount = _number(_pick(row, "金额", "交易金额"))
                if income > 0:
                    direction, amount = "收入", income
                elif expense > 0:
                    direction, amount = "支出", expense
                elif direction not in ("收入", "支出"):
                    direction = "收入" if amount >= 0 else "支出"
                    amount = abs(amount)
                records.append({
                    "date": str(_pick(row, "交易日期", "记账日期", "日期"))[:10],
                    "direction": direction, "amount": amount,
                    "summary": str(_pick(row, "摘要", "用途", "交易摘要", "备注")),
                    "counterparty": str(_pick(row, "对方户名", "对方名称", "交易对手")),
                    "account": str(_pick(row, "本方账号", "账号", "银行账号")),
                    "balance": _number(_pick(row, "余额", "账户余额")),
                    "source_file": str(path),
                })
            result = self.store.import_bank_transactions(records)
            self._refresh_bank()
            self._notify()
            messagebox.showinfo(
                "导入完成", f"新增 {result['imported']} 条，跳过重复 {result['skipped']} 条"
            )
        except Exception as exc:
            messagebox.showerror("导入失败", str(exc))

    def _auto_reconcile(self):
        period = self._valid_period(self.bank_period_var)
        if not period:
            return
        try:
            result = self.store.auto_reconcile_bank_transactions(period)
            self._refresh_bank()
            self._notify()
            messagebox.showinfo(
                "对账完成", f"本次自动匹配 {result['matched']} 条，未匹配 {result['unmatched']} 条"
            )
        except Exception as exc:
            messagebox.showerror("对账失败", str(exc))

    def _manual_reconcile(self):
        record_id = self._selected_id(self.bank_tree, "银行流水")
        if not record_id:
            return
        voucher_no = simpledialog.askstring("手工匹配", "请输入凭证号：", parent=self)
        if not voucher_no:
            return
        try:
            self.store.set_bank_match(record_id, voucher_no)
            self._refresh_bank()
            self._notify()
        except Exception as exc:
            messagebox.showerror("匹配失败", str(exc))

    def _unmatch_bank(self):
        record_id = self._selected_id(self.bank_tree, "银行流水")
        if not record_id:
            return
        try:
            self.store.set_bank_match(record_id, "")
            self._refresh_bank()
            self._notify()
        except Exception as exc:
            messagebox.showerror("取消失败", str(exc))

    def _set_cash_flow_category(self):
        record_id = self._selected_id(self.bank_tree, "银行流水")
        if not record_id:
            return
        record = next(
            (row for row in self.store.list_bank_transactions() if row["id"] == record_id),
            None,
        )
        if not record:
            return
        options = self.store.cash_flow_category_options(record.get("direction", ""))
        current_key = str(record.get("cash_flow_category", ""))
        current_label = next(
            (
                label for label in options
                if self.store.cash_flow_category_key(label) == current_key
            ),
            "自动判断",
        )

        def save(payload):
            label = payload["category"]
            category = "" if label == "自动判断" else self.store.cash_flow_category_key(label)
            self.store.set_bank_cash_flow_category(record_id, category)
            self._refresh_bank()
            self._notify()

        self._record_dialog(
            "现金流量项目分类",
            [("category", "现金流量项目", ["自动判断", *options])],
            {"category": current_label},
            save,
            "760x180",
        )

    def _delete_bank(self):
        selected = self.bank_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择银行流水")
            return
        if messagebox.askyesno("确认删除", f"删除选中的 {len(selected)} 条银行流水？"):
            try:
                self.store.delete_bank_transactions(selected)
                self._refresh_bank()
                self._notify()
            except Exception as exc:
                messagebox.showerror("删除失败", str(exc))

    def _refresh_bank(self):
        period = self._valid_period(self.bank_period_var)
        if not period:
            return
        for item in self.bank_tree.get_children():
            self.bank_tree.delete(item)
        records = self.store.list_bank_transactions(period)
        entry_map = {}
        if self.enterprise_mode:
            entry_map = {
                str(row.get("line_id", "")): row
                for row in self.store.cash_flow_summary(period)["month_entries"]
            }
        for row in records:
            values = [
                row["date"], row["direction"], f"¥{row['amount']:,.2f}",
                row.get("summary", ""), row.get("counterparty", ""),
                row.get("voucher_no", ""),
            ]
            if self.enterprise_mode:
                entry = entry_map.get(str(row.get("voucher_line_id", "")))
                category_key = str(row.get("cash_flow_category", ""))
                manual_label = next(
                    (
                        label for label in self.store.cash_flow_category_options(
                            row.get("direction", "")
                        )
                        if self.store.cash_flow_category_key(label) == category_key
                    ),
                    "",
                )
                category_label = manual_label or (
                    entry.get("category_label", "") if entry else ""
                )
                if manual_label:
                    source = "手工" if row.get("voucher_line_id") else "手工（待匹配）"
                elif entry:
                    source = "自动（待复核）" if entry.get("needs_review") else "自动"
                else:
                    source = "待匹配"
                values.extend([category_label or "未分类", source])
            values.append(row.get("status", "未匹配"))
            self.bank_tree.insert("", "end", iid=row["id"], values=values)
        income = sum(row["amount"] for row in records if row["direction"] == "收入")
        expense = sum(row["amount"] for row in records if row["direction"] == "支出")
        unmatched = sum(1 for row in records if not row.get("voucher_no"))
        summary = (
            f"收入 ¥{income:,.2f}    支出 ¥{expense:,.2f}    净流入 ¥{income-expense:,.2f}    "
            f"未匹配 {unmatched} 条"
        )
        if self.enterprise_mode:
            review_count = sum(1 for row in entry_map.values() if row.get("needs_review"))
            summary += f"    现金流待复核 {review_count} 条"
        self.bank_summary_var.set(summary)

    def _add_payroll(self):
        self._payroll_dialog({"period": self.payroll_period_var.get(), "pay_date": date.today().isoformat()})

    def _edit_payroll(self):
        record_id = self._selected_id(self.payroll_tree, "工资记录")
        if not record_id:
            return
        record = next((row for row in self.store.list_payroll() if row["id"] == record_id), None)
        if record:
            self._payroll_dialog(record)

    def _payroll_dialog(self, initial):
        record_id = initial.get("id", "")
        fields = [
            ("period", "工资期间", None), ("employee_name", "员工姓名", None),
            ("gross_salary", "应发工资", None), ("social_personal", "个人社保", None),
            ("housing_personal", "个人公积金", None), ("income_tax", "个人所得税", None),
            ("social_company", "单位社保", None), ("housing_company", "单位公积金", None),
            ("pay_date", "计提日期", None), ("note", "备注", None),
        ]

        def save(payload):
            self.store.upsert_payroll({
                "id": record_id, **payload,
                "gross_salary": _number(payload["gross_salary"]),
                "social_personal": _number(payload["social_personal"]),
                "housing_personal": _number(payload["housing_personal"]),
                "income_tax": _number(payload["income_tax"]),
                "social_company": _number(payload["social_company"]),
                "housing_company": _number(payload["housing_company"]),
            })
            self.payroll_period_var.set(payload["period"])
            self._refresh_payroll()
            self._notify()

        self._record_dialog("工资社保记录", fields, initial, save, "700x640")

    def _delete_payroll(self):
        record_id = self._selected_id(self.payroll_tree, "工资记录")
        if record_id and messagebox.askyesno("确认删除", "删除选中的工资记录？"):
            try:
                self.store.delete_payroll(record_id)
                self._refresh_payroll()
                self._notify()
            except Exception as exc:
                messagebox.showerror("删除失败", str(exc))

    def _post_payroll(self):
        record_id = self._selected_id(self.payroll_tree, "工资记录")
        if not record_id:
            return
        if not messagebox.askyesno("生成凭证", "按当前工资社保数据生成计提凭证？"):
            return
        try:
            voucher_no = self.store.post_payroll_voucher(record_id)
            L.log("生成工资凭证", voucher_no)
            self._refresh_payroll()
            self._notify()
            messagebox.showinfo("生成完成", f"凭证号：{voucher_no}")
        except Exception as exc:
            messagebox.showerror("生成失败", str(exc))

    def _refresh_payroll(self):
        period = self._valid_period(self.payroll_period_var)
        if not period:
            return
        for item in self.payroll_tree.get_children():
            self.payroll_tree.delete(item)
        records = self.store.list_payroll(period)
        for row in records:
            self.payroll_tree.insert("", "end", iid=row["id"], values=(
                row["employee_name"], f"¥{row['gross_salary']:,.2f}",
                f"¥{row['social_personal']:,.2f}", f"¥{row['housing_personal']:,.2f}",
                f"¥{row['income_tax']:,.2f}", f"¥{row['net_salary']:,.2f}",
                f"¥{row['social_company']:,.2f}", f"¥{row['housing_company']:,.2f}",
                row.get("status", ""), row.get("voucher_no", ""),
            ))
        gross = sum(row["gross_salary"] for row in records)
        net = sum(row["net_salary"] for row in records)
        company_cost = sum(
            row["gross_salary"] + row["social_company"] + row["housing_company"]
            for row in records
        )
        self.payroll_summary_var.set(
            f"应发合计 ¥{gross:,.2f}    实发合计 ¥{net:,.2f}    企业人工成本 ¥{company_cost:,.2f}"
        )

    def _add_asset(self):
        purchase = date.today()
        next_month = purchase.month % 12 + 1
        next_year = purchase.year + (1 if purchase.month == 12 else 0)
        self._asset_dialog({
            "category": "办公设备", "purchase_date": purchase.isoformat(),
            "original_cost": "", "residual_rate_percent": "5",
            "useful_months": "36", "depreciation_start_period": f"{next_year:04d}-{next_month:02d}",
            "asset_subject": "1601 固定资产", "depreciation_subject": "1602 累计折旧",
            "expense_subject": self.store.management_expense_subject,
            "status": "使用中",
        })

    def _edit_asset(self):
        record_id = self._selected_id(self.asset_tree, "固定资产")
        if not record_id:
            return
        record = next((row for row in self.store.list_fixed_assets() if row["id"] == record_id), None)
        if record:
            initial = {**record, "residual_rate_percent": record["residual_rate"] * 100}
            self._asset_dialog(initial)

    def _asset_dialog(self, initial):
        record_id = initial.get("id", "")
        fields = [
            ("asset_name", "资产名称", None),
            ("category", "资产类别", ["办公设备", "机器设备", "运输工具", "电子设备", "其他"]),
            ("purchase_date", "购置日期", None), ("original_cost", "资产原值", None),
            ("residual_rate_percent", "预计净残值率（%）", None),
            ("useful_months", "预计使用月数", None),
            ("depreciation_start_period", "折旧起始期间", None),
            ("asset_subject", "固定资产科目", self.subject_options),
            ("depreciation_subject", "累计折旧科目", self.subject_options),
            ("expense_subject", "折旧费用科目", self.subject_options),
            ("status", "资产状态", ["使用中", "停用", "已处置"]),
            ("note", "备注", None),
        ]

        def save(payload):
            self.store.upsert_fixed_asset({
                "id": record_id, **payload,
                "original_cost": _number(payload["original_cost"]),
                "residual_rate": _number(payload["residual_rate_percent"]) / 100,
                "useful_months": int(_number(payload["useful_months"])),
            })
            self._refresh_assets()
            self._notify()

        self._record_dialog("固定资产", fields, initial, save, "760x760")

    def _delete_asset(self):
        record_id = self._selected_id(self.asset_tree, "固定资产")
        if record_id and messagebox.askyesno("确认删除", "删除选中的固定资产？"):
            try:
                self.store.delete_fixed_asset(record_id)
                self._refresh_assets()
                self._notify()
            except Exception as exc:
                messagebox.showerror("删除失败", str(exc))

    def _post_depreciation(self):
        period = self._valid_period(self.asset_period_var)
        if not period:
            return
        if not messagebox.askyesno("生成凭证", f"生成 {period} 固定资产折旧凭证？"):
            return
        try:
            voucher_no = self.store.post_depreciation_voucher(period)
            L.log("生成折旧凭证", voucher_no)
            self._refresh_assets()
            self._notify()
            messagebox.showinfo("生成完成", f"凭证号：{voucher_no}")
        except Exception as exc:
            messagebox.showerror("生成失败", str(exc))

    def _refresh_assets(self):
        period = self._valid_period(self.asset_period_var)
        if not period:
            return
        for item in self.asset_tree.get_children():
            self.asset_tree.delete(item)
        schedule = self.store.depreciation_schedule(period)
        for row in schedule:
            self.asset_tree.insert("", "end", iid=row["id"], values=(
                row["asset_name"], row["category"], row["purchase_date"],
                f"¥{row['original_cost']:,.2f}", row["useful_months"],
                f"¥{row['monthly_depreciation']:,.2f}", f"¥{row['depreciation_amount']:,.2f}",
                f"¥{row['accumulated_depreciation']:,.2f}", f"¥{row['net_book_value']:,.2f}",
                row["status"], row.get("voucher_no", ""),
            ))
        original = sum(row["original_cost"] for row in schedule)
        current = sum(row["depreciation_amount"] for row in schedule)
        accumulated = sum(row["accumulated_depreciation"] for row in schedule)
        unposted = sum(1 for row in schedule if row["depreciation_amount"] > 0 and not row["posted"])
        self.asset_summary_var.set(
            f"资产原值 ¥{original:,.2f}    本期折旧 ¥{current:,.2f}    "
            f"累计折旧 ¥{accumulated:,.2f}    待计提 {unposted} 项"
        )

    def _period_values(self):
        values = {date.today().strftime("%Y-%m")}
        values.update(row.get("period", "") for row in self.store.list_vouchers())
        values.update(row.get("period", "") for row in self.store.list_opening_balances())
        values.update(row.get("period", "") for row in self.store.list_payroll())
        values.update(str(row.get("date", ""))[:7] for row in self.store.list_bank_transactions())
        return sorted((value for value in values if len(value) == 7), reverse=True)

    def reload_from_store(self):
        values = self._period_values()
        for combo in (
            self.opening_period_combo, self.bank_period_combo,
            self.payroll_period_combo, self.asset_period_combo,
        ):
            combo.configure(values=values)
        self._refresh_opening()
        self._refresh_bank()
        self._refresh_payroll()
        self._refresh_assets()

    def _notify(self):
        if self.on_data_changed:
            self.on_data_changed()

    def set_authenticated(self, active: bool, operator: str = ""):
        self.authenticated = active

    def pack(self, **kwargs):
        super().pack(**kwargs)
        self.reload_from_store()
